"""
Kiwi-Mem — 带记忆系统的 LLM 转发网关
=============================================
让你的 AI 拥有长期记忆。

工作原理：
1. 接收客户端（Kelivo / ChatBox / 任何 OpenAI 兼容客户端）的消息
2. 自动搜索数据库中的相关记忆，注入 system prompt
3. 转发给 LLM API（支持 OpenRouter / OpenAI / 任何兼容接口）
4. 后台自动存储对话 + 用 AI 提取新记忆

环境变量 MEMORY_ENABLED=false 时退化为纯转发网关（第一阶段）。
"""

import os
import json
import base64
import binascii
import hashlib
import uuid
import asyncio
import copy
import math
import httpx
from contextlib import asynccontextmanager
from datetime import datetime, timezone, timedelta
from fastapi import FastAPI, Request, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware

from database import (
    init_tables, close_pool, get_pool, save_message, delete_latest_assistant_message, search_memories, save_memory,
    track_memory_recall, touch_permanent_memories, search_scenes,
    get_all_memories_count, get_recent_memories, get_recent_messages, get_latest_conversation_session_id, delete_memory,
    batch_delete_memories_guarded, clear_all_memories, update_memory, check_memory_duplicate,
    search_archived_memories, get_memory_versions,
    ingest_memory_event, ingest_memory_events, get_memory_handoff,
    list_memory_event_conversations, get_memory_event_conversation_page, search_memory_events,
    migrate_embeddings, backfill_permanent_memory_embeddings, get_embedding_stats,
    # v5.3 时间有效期 + 矛盾检测
    invalidate_memory, create_memory_edge, detect_contradictions,
    get_all_providers, get_provider, create_provider, update_provider, delete_provider,
    get_provider_models, get_all_saved_models, get_enabled_provider_models, add_provider_model, update_provider_model, delete_provider_model,
    resolve_provider_for_model,
    get_all_categories, create_category, update_category, delete_category, match_category_by_name,
    get_system_prompt_from_db, set_system_prompt_in_db,
    # v4.1 云端同步
    sync_get_conversations, sync_get_conversation, sync_upsert_conversation, sync_delete_conversation,
    sync_upsert_messages, sync_get_projects, sync_upsert_project, sync_delete_project, sync_import_all,
    # v4.2 提醒系统
    create_reminder, get_reminders, update_reminder, delete_reminder, get_due_reminders, fire_reminder,
)
from config import (
    SYNC_SETTING_KEYS,
    get_all_config, set_config, get_config, get_config_int, get_config_bool, get_config_float,
)
from memory_extractor import extract_memories
from memory_identity import (
    ArchiveIdentifierConfigError,
    MemoryClientConfigError,
    RECALL_IDENTITY_INSTRUCTION,
    archive_identifier_key_fingerprint,
    authenticate_memory_client,
    is_opaque_archive_identifier,
    prepare_generated_memory,
    validate_autobiographical_memory,
)
from mcp_server import get_mcp_app, get_calendar_mcp_app, mcp_memory, mcp_calendar
from web_search import web_search, format_results_for_prompt, get_engine_list
from mcp_client import get_tools_for_servers, call_tool, call_tools_batch, clear_tool_cache
from anthropic_adapter import (
    to_anthropic_request, to_anthropic_headers, get_anthropic_url,
    from_anthropic_response, anthropic_stream_to_openai,
)
from access_control import (
    AdminTokenConfigError,
    authenticate_admin_token,
    authenticate_internal_control,
    bearer_token,
    is_control_plane_path,
    is_shared_identity_entry_path,
)

# ============================================================
# 配置项 —— 全部从环境变量读取，部署时在云平台面板里设置
# ============================================================

# 你的 API Key（OpenRouter / OpenAI / 其他兼容服务）
API_KEY = os.getenv("API_KEY", "")

# API 地址（改这个就能切换不同的 LLM 服务商）
API_BASE_URL = os.getenv("API_BASE_URL", "https://openrouter.ai/api/v1/chat/completions")

# 默认模型（如果客户端没指定就用这个）
DEFAULT_MODEL = os.getenv("DEFAULT_MODEL", "anthropic/claude-sonnet-4-6")

# 网关端口
PORT = int(os.getenv("PORT", "8080"))

# 记忆系统开关（数据库出问题时可以临时关掉）
MEMORY_ENABLED = os.getenv("MEMORY_ENABLED", "true" if os.getenv("DATABASE_URL") else "false").lower() == "true"
CHAT_ARCHIVE_ENABLED = os.getenv("CHAT_ARCHIVE_ENABLED", "false").lower() == "true"

# 每次注入的最大记忆条数
MAX_MEMORIES_INJECT = int(os.getenv("MAX_MEMORIES_INJECT", "15"))

# 记忆提取间隔：每隔几轮对话提取一次记忆（默认3轮）
MEMORY_EXTRACT_INTERVAL = int(os.getenv("MEMORY_EXTRACT_INTERVAL", "3"))


# ============================================================
# 动态配置读取（v3.1）
# ============================================================
# 配置优先级：数据库 > 环境变量 > 默认值
# 以上三个变量保留作为启动时/数据库不可用时的降级值
# 运行时通过以下函数读取最新配置

async def get_memory_enabled() -> bool:
    """读取记忆开关（动态）"""
    try:
        return await get_config_bool("memory_enabled", fallback=MEMORY_ENABLED)
    except Exception:
        return MEMORY_ENABLED


async def get_chat_archive_enabled() -> bool:
    """Read the opt-in immutable transcript archive switch."""
    try:
        return await get_config_bool("chat_archive_enabled", fallback=CHAT_ARCHIVE_ENABLED)
    except Exception:
        return CHAT_ARCHIVE_ENABLED

async def get_max_inject() -> int:
    """读取注入条数（动态）"""
    try:
        return await get_config_int("max_inject", fallback=MAX_MEMORIES_INJECT)
    except Exception:
        return MAX_MEMORIES_INJECT

async def get_locked_inject_ratio() -> float:
    """读取锁定记忆保底比例（0~1），越界配置按边界夹住。"""
    try:
        ratio = await get_config_float("locked_inject_ratio", fallback=0.2)
    except Exception:
        ratio = 0.2
    return max(0.0, min(1.0, float(ratio)))

async def get_extract_interval() -> int:
    """读取提取间隔（动态）"""
    try:
        return await get_config_int("extract_interval", fallback=MEMORY_EXTRACT_INTERVAL)
    except Exception:
        return MEMORY_EXTRACT_INTERVAL

# 额外的请求头（有些 API 需要，比如 OpenRouter 需要 Referer）
EXTRA_REFERER = os.getenv("EXTRA_REFERER", "https://ai-memory-gateway.local")
EXTRA_TITLE = os.getenv("EXTRA_TITLE", "Kiwi-Mem")
AIHUBMIX_EXTENDED_CACHE_TTL_BETA = "extended-cache-ttl-2025-04-11"






# ============================================================
# 对话计数器（控制记忆提取频率）
# ============================================================

_conversation_counter = 0
_counter_lock = asyncio.Lock()


# ============================================================
# 后台任务引用（防止 GC 回收）
# ============================================================

_background_tasks: set = set()


def _spawn_background_task(coro):
    """启动后台任务并保留引用，避免在执行中被 GC 回收。"""
    task = asyncio.create_task(coro)
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)
    return task


async def _backfill_permanent_embeddings_background():
    """启动后轻量回填老锁定记忆的 embedding，失败只降级为日志。"""
    try:
        result = await backfill_permanent_memory_embeddings()
        if isinstance(result, dict) and result.get("total", 0):
            print(f"🔒 锁定记忆向量回填结果：{result}")
    except Exception as e:
        print(f"⚠️  锁定记忆向量回填失败（不影响启动）: {type(e).__name__}: {e}")


# ============================================================
# 人设加载
# ============================================================

def load_system_prompt():
    """从 system_prompt.txt 文件读取人设内容（降级方案）"""
    prompt_path = os.path.join(os.path.dirname(__file__), "system_prompt.txt")
    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if content:
                return content
    except FileNotFoundError:
        pass
    print("ℹ️  未找到 system_prompt.txt 或文件为空，将不注入 system prompt")
    return ""


# 文件版作为启动降级值
_FILE_SYSTEM_PROMPT = load_system_prompt()
# 运行时变量（可被数据库覆盖）
SYSTEM_PROMPT = _FILE_SYSTEM_PROMPT

if SYSTEM_PROMPT:
    print(f"✅ 人设已加载（文件），长度：{len(SYSTEM_PROMPT)} 字符")
else:
    print("ℹ️  无人设，纯转发模式")


async def get_active_system_prompt() -> str:
    """获取当前生效的 system prompt（数据库优先，文件降级）"""
    try:
        db_prompt = await get_system_prompt_from_db()
        if db_prompt is not None:
            return db_prompt
    except Exception:
        pass
    return _FILE_SYSTEM_PROMPT


# ============================================================
# 应用生命周期管理
# ============================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用启动时初始化数据库和MCP，关闭时断开连接"""
    digest_task = None
    dream_check_task = None
    
    if MEMORY_ENABLED:
        try:
            await init_tables()
            pool = await get_pool()
            async with pool.acquire() as conn:
                count = await conn.fetchval("""
                    SELECT COUNT(*)
                    FROM memories
                    WHERE COALESCE(memory_type, 'fragment') != 'dream_deleted'
                """)
            print(f"✅ 记忆系统已启动，当前活跃记忆数量：{count}")
            print(f"📊 记忆提取间隔：每 {MEMORY_EXTRACT_INTERVAL} 轮对话提取一次")
            
            # v5.6：首次启动时将出厂默认 prompt 写入 config 表（空值才写入）
            try:
                factory = _get_factory_prompts()
                seeded = 0
                for key, default_text in factory.items():
                    existing = await get_config(key)
                    if not existing:
                        await set_config(key, default_text)
                        seeded += 1
                if seeded > 0:
                    print(f"📝 首次启动：写入了 {seeded} 个默认 prompt 到配置表")
            except Exception as e:
                print(f"⚠️  默认 prompt 初始化失败: {e}")

            _spawn_background_task(_backfill_permanent_embeddings_background())
            
            # 启动每日记忆整理调度器
            from daily_digest import daily_digest_scheduler
            digest_task = asyncio.create_task(daily_digest_scheduler())
            
            # 启动自动 Dream 检查器（每小时检查24h无活动）
            from dream import auto_dream_scheduler
            dream_check_task = asyncio.create_task(auto_dream_scheduler())
            
        except Exception as e:
            print(f"⚠️  数据库初始化失败: {e}")
            print("⚠️  记忆系统将不可用，但网关仍可正常转发")
    else:
        print("ℹ️  记忆系统已关闭（设置 MEMORY_ENABLED=true 开启）")
    
    # 启动 MCP session managers（v5.4：两个模块）
    async with mcp_memory.session_manager.run():
        async with mcp_calendar.session_manager.run():
            print("✅ MCP server 已启动（/memory/mcp + /calendar/mcp）")

            # v6.3：仅当 tool_drawer_enabled=true 时初始化工具抽屉
            # （默认 false，开源用户行为不变；抽屉打开后通过向量路由按需展开工具）
            try:
                drawer_enabled = await get_config_bool("tool_drawer_enabled", fallback=False)
            except Exception:
                drawer_enabled = False
            if drawer_enabled:
                try:
                    from tool_drawer import init_drawer
                    await init_drawer()
                    print("✅ 工具抽屉已初始化")
                except Exception as e:
                    print(f"⚠️ 工具抽屉初始化失败: {e}（降级为传统模式）")

            yield
    
    if digest_task:
        digest_task.cancel()
    if dream_check_task:
        dream_check_task.cancel()
    if MEMORY_ENABLED:
        await close_pool()


app = FastAPI(title="Kiwi-Mem", version="1.3.0", lifespan=lifespan)


# ============================================================
# CORS 配置 — 从环境变量读取允许的域名
# ============================================================

# CORS 白名单：通过环境变量配置，逗号分隔
# 示例：CORS_ORIGINS=https://your-frontend.example.com,http://localhost:5173
_cors_env = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://localhost:5174")
_cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def require_control_plane_auth(request: Request, call_next):
    """Fail closed on operator routes and every externally callable transport."""
    path = request.url.path

    if is_shared_identity_entry_path(path):
        # Browser CORS preflights contain no credentials or application data;
        # the subsequent request still has to authenticate.
        if request.method == "OPTIONS":
            return await call_next(request)
        try:
            context = authenticate_memory_client(bearer_token(request.headers.get("authorization", "")))
        except MemoryClientConfigError as exc:
            return JSONResponse(status_code=503, content={"detail": str(exc)})
        if context is None:
            return JSONResponse(
                status_code=401,
                content={"detail": "memory client bearer token required"},
                headers={"WWW-Authenticate": "Bearer"},
            )
        # The mounted transports currently share one canonical memory space.
        # Keeping the context on request.state avoids ever reflecting the door
        # name in response payloads while leaving room for future scoped tools.
        request.state.memory_client_context = context
        return await call_next(request)

    if not is_control_plane_path(path):
        return await call_next(request)

    internal_token = request.headers.get("x-kiwi-internal-control", "")
    if authenticate_internal_control(internal_token):
        return await call_next(request)

    try:
        authenticated = authenticate_admin_token(bearer_token(request.headers.get("authorization", "")))
    except AdminTokenConfigError as exc:
        return JSONResponse(status_code=503, content={"detail": str(exc)})
    if not authenticated:
        return JSONResponse(
            status_code=401,
            content={"detail": "admin bearer token required"},
            headers={"WWW-Authenticate": "Bearer"},
        )
    return await call_next(request)


def _memory_client_context(request: Request):
    """Authenticate a door while keeping identity assignment server-owned."""
    authorization = request.headers.get("authorization", "")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token.strip():
        raise HTTPException(
            status_code=401,
            detail="memory client bearer token required",
            headers={"WWW-Authenticate": "Bearer"},
        )
    try:
        context = authenticate_memory_client(token.strip())
    except MemoryClientConfigError:
        raise HTTPException(status_code=503, detail="memory client authentication is not configured")
    if context is None:
        raise HTTPException(
            status_code=401,
            detail="invalid memory client bearer token",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return context


def _reject_memory_identity_override(payload: dict) -> None:
    reserved = {"assistant_identity_id", "memory_space_id", "source_client"}
    attempted = sorted(reserved.intersection(payload))
    if attempted:
        raise HTTPException(
            status_code=422,
            detail=f"identity fields are assigned by the server: {', '.join(attempted)}",
        )


def _required_memory_text(payload: dict, field: str, max_length: int) -> str:
    value = str(payload.get(field) or "").strip()
    if not value:
        raise HTTPException(status_code=422, detail=f"{field} must not be empty")
    if len(value) > max_length:
        raise HTTPException(status_code=422, detail=f"{field} exceeds {max_length} characters")
    return value


_PUBLIC_ARCHIVE_CONVERSATION_FIELDS = (
    "conversation_id",
    "first_event_at",
    "last_event_at",
    "message_count",
    "user_message_count",
    "assistant_message_count",
    "preview",
)
_PUBLIC_ARCHIVE_EVENT_FIELDS = (
    "conversation_id",
    "role",
    "content",
    "occurred_at",
)


def _public_archive_record(row, allowed_fields: tuple[str, ...]) -> dict:
    """Project audit rows onto the identity-neutral public archive contract."""
    source = dict(row)
    return {field: source.get(field) for field in allowed_fields}


def _assert_current_archive_rows(rows) -> None:
    """Never serialize legacy/raw room identifiers through a public response."""
    for row in rows:
        source = dict(row)
        conversation_id = source.get("conversation_id")
        event_id = source.get("event_id")
        if conversation_id and not is_opaque_archive_identifier(conversation_id, "conversation"):
            raise HTTPException(status_code=503, detail="legacy archive identifiers require migration")
        if event_id and not is_opaque_archive_identifier(event_id, "event"):
            raise HTTPException(status_code=503, detail="legacy archive identifiers require migration")


def _required_memory_content(payload: dict, field: str = "content", max_length: int = 200000) -> str:
    """Validate raw evidence without trimming any verbatim whitespace."""
    raw = payload.get(field)
    value = "" if raw is None else str(raw)
    if not value.strip():
        raise HTTPException(status_code=422, detail=f"{field} must not be empty")
    if len(value) > max_length:
        raise HTTPException(status_code=422, detail=f"{field} exceeds {max_length} characters")
    return value


def _parse_memory_event_time(value):
    if value in (None, ""):
        return datetime.now(timezone.utc)
    if isinstance(value, datetime):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
        except ValueError:
            raise HTTPException(status_code=422, detail="occurred_at must be an ISO-8601 timestamp")
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _encode_archive_cursor(occurred_at: datetime, item_id: str) -> str:
    payload = json.dumps(
        {"at": occurred_at.astimezone(timezone.utc).isoformat(), "id": str(item_id)},
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return base64.urlsafe_b64encode(payload).decode("ascii").rstrip("=")


def _decode_archive_cursor(value: str) -> tuple[datetime | None, str | None]:
    cursor = str(value or "").strip()
    if not cursor:
        return None, None
    try:
        padded = cursor + "=" * (-len(cursor) % 4)
        payload = json.loads(base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8"))
        occurred_at = _parse_memory_event_time(payload.get("at"))
        item_id = str(payload.get("id") or "")
        if not item_id:
            raise ValueError("missing cursor id")
        return occurred_at, item_id
    except (ValueError, TypeError, json.JSONDecodeError, UnicodeDecodeError, binascii.Error):
        raise HTTPException(status_code=422, detail="invalid archive cursor")


def _prepare_memory_event_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="event must be a JSON object")
    _reject_memory_identity_override(payload)
    role = str(payload.get("role") or "").strip().lower()
    if role not in {"user", "assistant", "system"}:
        raise HTTPException(status_code=422, detail="role must be user, assistant, or system")
    metadata = payload.get("metadata", {})
    if not isinstance(metadata, dict):
        raise HTTPException(status_code=422, detail="metadata must be a JSON object")
    return {
        "event_id": _required_memory_text(payload, "event_id", 200),
        "conversation_id": _required_memory_text(payload, "conversation_id", 200),
        "role": role,
        "content": _required_memory_content(payload),
        "occurred_at": _parse_memory_event_time(payload.get("occurred_at")),
        "metadata": metadata,
    }

# ============================================================
# 模板变量替换
# ============================================================

TZ_CST = timezone(timedelta(hours=8))  # 东八区

def replace_template_variables(text: str, context: dict = None) -> str:
    """
    替换 system prompt / skill prompt 中的模板变量。
    支持的变量：
      {cur_datetime}    → 2026-03-24 14:30:00
      {cur_date}        → 2026-03-24
      {cur_time}        → 14:30:00
      {cur_weekday}     → 星期一
      {model_name}      → deepseek/deepseek-chat-v3-0324
      {user_name}       → 用户昵称
      {assistant_name}  → AI名字
    """
    if not text or '{' not in text:
        return text

    now = datetime.now(TZ_CST)
    weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    ctx = context or {}

    replacements = {
        '{cur_datetime}':   now.strftime('%Y-%m-%d %H:%M:%S'),
        '{cur_date}':       now.strftime('%Y-%m-%d'),
        '{cur_time}':       now.strftime('%H:%M:%S'),
        '{cur_weekday}':    weekdays[now.weekday()],
        '{model_name}':     ctx.get('model_name', ''),
        '{user_name}':      ctx.get('user_name', ''),
        '{assistant_name}': ctx.get('assistant_name', ''),
    }

    for key, val in replacements.items():
        # 即便 val 是空字符串也要替换，否则 {user_name} 等占位符会原样残留进 prompt
        if key in text and isinstance(val, str):
            text = text.replace(key, val)

    return text


# ============================================================
# 记忆注入
# ============================================================

def _format_scene_field(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except Exception:
            return value.strip()
    if isinstance(value, list):
        parts = []
        for item in value:
            if item is None:
                continue
            if isinstance(item, str):
                text = item.strip()
            elif isinstance(item, dict):
                text = str(item.get("content") or item.get("text") or "").strip()
                if not text:
                    text = json.dumps(item, ensure_ascii=False)
            else:
                text = str(item).strip()
            if text:
                parts.append(text)
        return "；".join(parts)
    if isinstance(value, dict):
        text = str(value.get("content") or value.get("text") or "").strip()
        return text or json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _filter_valid_foresight(value):
    """Filter out expired foresight items at runtime without changing storage."""
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except Exception:
            return value
    if not isinstance(value, list):
        return value

    today = datetime.now(TZ_CST).date()
    kept = []
    for item in value:
        if not isinstance(item, dict):
            kept.append(item)
            continue
        vu = item.get("valid_until")
        if not vu:
            kept.append(item)
            continue
        try:
            vu_date = datetime.fromisoformat(str(vu).strip().replace("Z", "+00:00")).date()
        except Exception:
            try:
                vu_date = datetime.strptime(str(vu).strip()[:10], "%Y-%m-%d").date()
            except Exception:
                kept.append(item)
                continue
        if vu_date >= today:
            kept.append(item)
    return kept


_COMPRESSED_SUMMARY_PREFIX = "[之前的对话摘要]"
_RUNTIME_CONTEXT_BEGIN = "[以下为系统自动注入的实时上下文，并非用户发送]"
_RUNTIME_CONTEXT_END = "[实时上下文结束，以下是用户消息]"


def _content_text_for_prompt(content) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text") or ""))
        return "\n".join(p for p in parts if p)
    if content is None:
        return ""
    return str(content)


def _is_compressed_summary_content(content) -> bool:
    return _content_text_for_prompt(content).lstrip().startswith(_COMPRESSED_SUMMARY_PREFIX)


def _is_runtime_context_system_content(content) -> bool:
    text = _content_text_for_prompt(content).lstrip()
    return text.startswith("[系统提醒]")


def _is_compressed_summary_message(message: dict) -> bool:
    return message.get("role") == "user" and _is_compressed_summary_content(message.get("content"))


def _count_real_user_messages(messages: list) -> int:
    return sum(
        1 for message in messages
        if message.get("role") == "user" and not _is_compressed_summary_message(message)
    )


def _append_runtime_context_section(prompt_meta: dict, text: str) -> None:
    text = (text or "").strip()
    if text:
        prompt_meta.setdefault("runtime_context_sections", []).append(text)


def _merge_frontend_system_prompt(enhanced_prompt: str, frontend_sys: str) -> str:
    frontend_sys = (frontend_sys or "").strip()
    if not frontend_sys:
        return enhanced_prompt

    # Legacy frontends may still send the compression summary as system.
    # It must be normalized into a user message before this merge step.
    if _is_compressed_summary_content(frontend_sys):
        return enhanced_prompt

    if "<!-- CACHE_B2 -->" in enhanced_prompt:
        b2_marker = "<!-- CACHE_B2 -->"
        b2_pos = enhanced_prompt.index(b2_marker)
        return (
            enhanced_prompt[:b2_pos]
            + "\n\n" + frontend_sys
            + "\n\n"
            + enhanced_prompt[b2_pos:]
        )

    return enhanced_prompt + "\n\n" + frontend_sys


def _normalize_frontend_system_messages(messages: list, template_ctx: dict, prompt_meta: dict) -> None:
    normalized = []
    legacy_summaries = []

    for msg in messages:
        if msg.get("role") == "system":
            system_text = msg.get("content") or ""
            if _is_runtime_context_system_content(system_text):
                runtime_system_text = replace_template_variables(system_text, template_ctx)
                _append_runtime_context_section(prompt_meta, runtime_system_text)
                continue
            if _is_compressed_summary_content(system_text):
                summary_text = replace_template_variables(system_text, template_ctx)
                legacy_summaries.append({"role": "user", "content": summary_text})
                continue
        normalized.append(msg)

    if legacy_summaries:
        insert_at = 0
        while insert_at < len(normalized) and normalized[insert_at].get("role") == "system":
            insert_at += 1
        normalized[insert_at:insert_at] = legacy_summaries

    if len(normalized) != len(messages) or legacy_summaries:
        messages[:] = normalized


def _inject_runtime_context_into_last_user(messages: list, sections: list) -> bool:
    clean_sections = [str(section).strip() for section in (sections or []) if str(section or "").strip()]
    if not clean_sections:
        return False

    injected_text = (
        _RUNTIME_CONTEXT_BEGIN
        + "\n"
        + "\n\n".join(clean_sections)
        + "\n"
        + _RUNTIME_CONTEXT_END
    )

    for message in reversed(messages):
        if message.get("role") != "user" or _is_compressed_summary_message(message):
            continue

        content = message.get("content", "")
        if isinstance(content, list):
            message["content"] = [{"type": "text", "text": injected_text}] + content
        elif isinstance(content, str):
            message["content"] = injected_text + "\n\n" + content
        elif content is None:
            message["content"] = injected_text
        else:
            message["content"] = injected_text + "\n\n" + str(content)
        return True

    return False


def _normalize_prompt_cache_ttl(value) -> str:
    ttl = str(value or "1h").strip().lower()
    return "1h" if ttl == "1h" else "5m"


def _cache_control_block(ttl: str = "1h") -> dict:
    block = {"type": "ephemeral"}
    if _normalize_prompt_cache_ttl(ttl) == "1h":
        block["ttl"] = "1h"
    return block


def _content_blocks_have_type(content, blocked_types: set) -> bool:
    if not isinstance(content, list):
        return False
    for block in content:
        if isinstance(block, dict) and block.get("type") in blocked_types:
            return True
    return False


def _message_has_image_content(content) -> bool:
    return _content_blocks_have_type(content, {"image", "image_url"})


def _message_has_tool_content(message: dict) -> bool:
    if message.get("tool_calls"):
        return True
    return _content_blocks_have_type(message.get("content"), {"tool_use", "tool_result"})


def _is_message_cache_candidate(message: dict) -> bool:
    if message.get("role") != "assistant":
        return False
    if _message_has_tool_content(message) or _message_has_image_content(message.get("content")):
        return False
    content = message.get("content")
    if isinstance(content, str):
        return bool(content.strip())
    if isinstance(content, list):
        return any(isinstance(block, dict) and block.get("type") == "text" and str(block.get("text") or "").strip() for block in content)
    return False


def _add_cache_control_to_message(message: dict, cache_ttl: str = "5m") -> bool:
    content = message.get("content")
    if isinstance(content, str):
        text = content.strip()
        if not text:
            return False
        message["content"] = [{"type": "text", "text": content, "cache_control": _cache_control_block(cache_ttl)}]
        return True

    if isinstance(content, list):
        blocks = copy.deepcopy(content)
        for block in reversed(blocks):
            if isinstance(block, dict) and block.get("type") == "text" and str(block.get("text") or "").strip():
                block["cache_control"] = _cache_control_block(cache_ttl)
                message["content"] = blocks
                return True
    return False


def _clear_message_cache_control(message: dict) -> None:
    content = message.get("content")
    if not isinstance(content, list):
        return
    for block in content:
        if isinstance(block, dict):
            block.pop("cache_control", None)


def _apply_messages_cache_breakpoint(messages: list, is_regenerate: bool = False, cache_ttl: str = "5m"):
    if is_regenerate:
        return None, "regenerate"
    if len(messages) < 2:
        return None, "not_enough_messages"

    for idx in range(len(messages) - 2, -1, -1):
        msg = messages[idx]
        if not _is_message_cache_candidate(msg):
            continue
        if _add_cache_control_to_message(msg, cache_ttl=cache_ttl):
            return idx, "applied"
    return None, "no_eligible_assistant"


def _apply_tools_cache_breakpoint(tools: list, tool_map: dict, cache_ttl: str = "5m") -> bool:
    if not tools or not tool_map:
        return False

    last_meta_idx = None
    for idx, tool in enumerate(tools):
        name = ((tool or {}).get("function") or {}).get("name")
        if tool_map.get(name, {}).get("type") == "meta":
            last_meta_idx = idx
            continue
        break

    if last_meta_idx is None:
        return False

    marked_tool = copy.deepcopy(tools[last_meta_idx])
    marked_tool["cache_control"] = _cache_control_block(cache_ttl)
    tools[last_meta_idx] = marked_tool
    return True


def _count_cache_control_entries_in_messages(messages: list) -> int:
    total = 0
    for msg in messages or []:
        content = msg.get("content")
        if isinstance(content, list):
            total += sum(1 for block in content if isinstance(block, dict) and block.get("cache_control"))
    return total


def _count_cache_control_entries_in_tools(tools: list) -> int:
    return sum(1 for tool in (tools or []) if isinstance(tool, dict) and tool.get("cache_control"))


def _memory_similarity(memory: dict) -> float:
    return float(memory.get("similarity") or 0)


def _memory_similarity_sort_key(memory: dict):
    """按 RRF/综合 score 排序，保留混合检索的原始竞争语义。"""
    return (
        float(memory.get("score") or 0),
        _memory_similarity(memory),
        int(memory.get("id") or 0),
    )


def _is_semantic_locked_memory(memory: dict, semantic_threshold: float) -> bool:
    return bool(memory.get("is_permanent")) and _memory_similarity(memory) >= semantic_threshold


def _rebalance_locked_memory_candidates(memories: list, max_inject: int, locked_ratio: float, semantic_threshold: float) -> list:
    """
    在已过搜索阈值的候选池内，为锁定记忆做注入名额保底。

    ratio=0 时等价于纯 score 竞争；保底只是下限，超出保底的锁定记忆会和普通碎片继续竞争。
    """
    if not memories or max_inject <= 0:
        return []

    locked_ratio = max(0.0, min(1.0, float(locked_ratio or 0)))
    if locked_ratio <= 0:
        return sorted(memories, key=_memory_similarity_sort_key, reverse=True)[:max_inject]

    reserved = min(max_inject, math.ceil(max_inject * locked_ratio))
    locked = [m for m in memories if _is_semantic_locked_memory(m, semantic_threshold)]
    if not locked:
        return sorted(memories, key=_memory_similarity_sort_key, reverse=True)[:max_inject]

    locked_sorted = sorted(locked, key=_memory_similarity_sort_key, reverse=True)
    guaranteed = locked_sorted[:reserved]
    guaranteed_ids = {m.get("id") for m in guaranteed if m.get("id") is not None}
    remaining = [m for m in memories if m.get("id") not in guaranteed_ids]
    remaining_slots = max_inject - len(guaranteed)
    competitive = sorted(remaining, key=_memory_similarity_sort_key, reverse=True)[:remaining_slots]
    return guaranteed + competitive


async def build_system_prompt_with_memories(user_message: str, user_msg_count: int = 1, project_id: str = None, conversation_id: str = None, is_regenerate: bool = False) -> tuple:
    """
    构建带记忆的 system prompt（v5.5 日历层级注入 + v5.8 项目注入 + 缓存优化）
    
    返回 (prompt_str, metadata_dict)，metadata 可能包含 handoff 信息。
    
    注入顺序（为 Prompt Caching 优化——不变的在前，变化的在后）：
    ── 静态区（同一天内不变，命中缓存 1/10 价格）──
    1. 人设
    2. 用户画像
    3. 锁定记忆（很少变）
    4. 日历层级注入（一天内不变）
    5. 项目指令（静态，整个项目内不变）
    ── 运行时注入区（每轮变化，放入最后一条 user 消息，不写回历史）──
    6. 语义搜索碎片（每轮根据用户消息重新搜索，含项目记忆）
    7. 项目文件相关片段（语义搜索）
    8. Dream 犯困提示
    """
    active_prompt = await get_active_system_prompt()

    # ---- 情绪标记指示（静态，系统级指令）----
    # 情绪隐藏标记的生产端：指示聊天模型在高情绪轮打 <!--emotion:x--> 标记，
    # 由 detect_emotion_from_response 解析。属系统机制而非人设内容，故代码注入。
    active_prompt += ("\n\n【情绪标记】当你判断本轮对话情绪浓度明显时（用户表达强烈情绪、脆弱袒露、重要情感时刻），"
                      "在回复的最末尾附加隐藏标记 <!--emotion:高--> 或 <!--emotion:中-->。日常对话不要加。"
                      "该标记对用户不可见，供记忆系统识别情绪锚点使用。")
    active_prompt += ("\n\n【做梦标记】当用户明确同意让你去做梦、整理记忆、去睡觉、去休息时，"
                      "在回复的最末尾附加隐藏标记 <!--dream:trigger-->。"
                      "只在用户主动要求或同意时才加，不要自行决定触发。"
                      "此标记可以和情绪标记同时存在，两者互不影响。")
    prompt_meta = {}
    
    # ---- ① 用户画像（静态，一天不变）----
    try:
        user_profile = await get_config("user_profile")
        if user_profile:
            active_prompt += f"\n\n【用户画像】\n{user_profile}"
    except Exception as e:
        print(f"⚠️  用户画像读取失败: {e}")
    
    if not await get_memory_enabled():
        return active_prompt, prompt_meta
    
    try:
        # ---- ② 锁定记忆：阶段 5 起不再全量注入 ----
        # 锁定记忆改为参与语义搜索；命中后进入运行时注入区，避免静态 system 前缀被频繁重写。

        # ── 缓存断点1：人设 + 画像（几乎不变）──
        active_prompt += "\n\n<!-- CACHE_B1 -->"
        
        # ---- ③ 日历层级注入（静态，一天内不变）----
        try:
            calendar_enabled = await get_config("calendar_inject_enabled")
            if calendar_enabled is None or str(calendar_enabled).lower() != 'false':
                from database import get_calendar_for_injection
                cal_entries = await get_calendar_for_injection(lookback_days=365)
                if cal_entries:
                    cal_lines = []
                    for entry in cal_entries:
                        label = entry.get("label", "")
                        # 优先用 digest（模型注入版），没有就用 summary（兜底），都没有就跳过
                        text = entry.get("digest") or entry.get("summary") or ""
                        if text:
                            cal_lines.append(f"📅 {label}：{text}")
                    if cal_lines:
                        cal_text = "\n".join(cal_lines)
                        active_prompt += f"\n\n【近期日历（从大到小的层级记忆，越远越概括）】\n{cal_text}"
                        print(f"📅 日历注入了 {len(cal_lines)} 条层级记忆")
        except Exception as e:
            print(f"⚠️  日历注入失败: {e}")
        
        # ---- ④ 项目指令注入（静态，整个项目内不变）----
        if project_id:
            try:
                from database import get_project_by_id
                proj = await get_project_by_id(project_id)
                if proj and proj.get("instructions"):
                    active_prompt += f"\n\n【项目指令】\n{proj['instructions']}"
                    print(f"📂 注入了项目指令（项目: {proj.get('name', project_id)}）")
            except Exception as e:
                print(f"⚠️  项目指令注入失败: {e}")
        
        # ---- v6.3：工具抽屉目录（静态，仅在抽屉开启时注入）----
        drawer_enabled_inject = await get_config_bool("tool_drawer_enabled", fallback=False)
        if drawer_enabled_inject:
            try:
                from tool_drawer import get_directory_text
                active_prompt += get_directory_text()
            except Exception as e:
                print(f"⚠️  工具目录注入失败: {e}")

        # ── 缓存断点2：日历 + 项目指令 + 工具目录（一天变一次）──
        active_prompt += "\n\n<!-- CACHE_B2 -->"

        # ---- 当前时间（动态区，分钟级）----
        # 注意：必须放在 CACHE_B2 之后。DeepSeek 没有显式 cache_control，
        # 完全依赖前缀字节一致来命中自动缓存——时间放在静态区会从这一行
        # 起整段前缀失效（issue #8）。分钟级足够日常使用，也让相邻请求的
        # prompt 更稳定、便于日志对比。
        _now_cst = datetime.now(TZ_CST)
        _weekdays_cn = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        _append_runtime_context_section(
            prompt_meta,
            f"【当前时间】{_now_cst.strftime('%Y-%m-%d %H:%M')} {_weekdays_cn[_now_cst.weekday()]}",
        )

        # ---- ⑤ 语义搜索碎片（动态，每轮变化）----
        inject_limit = max(0, await get_max_inject())
        locked_ratio = await get_locked_inject_ratio()
        candidate_limit = inject_limit * 3 if inject_limit > 0 else 0
        # 先只搜索，不立刻增加召回计数；等确认真正写进 prompt 后再补计数。
        memories, _query_emb = await search_memories(
            user_message,
            limit=candidate_limit,
            project_id=project_id,
            return_embedding=True,
            track_recall=False,
        )
        if _query_emb:
            prompt_meta["user_embedding"] = _query_emb
        
        # 加载热度参数（v5.4：可配置阈值）
        from database import get_heat_params
        heat_params = await get_heat_params()
        th_high = heat_params["threshold_high"]
        th_medium = heat_params["threshold_medium"]
        semantic_threshold = await get_config_float("semantic_threshold", fallback=0.25)
        
        # v5.6：中热度摘要截断字数（可配置）
        truncate_len = await get_config_int("heat_medium_truncate", fallback=60)
        
        semantic_locked_candidate_count = sum(
            1 for m in memories
            if _is_semantic_locked_memory(m, semantic_threshold) and m.get("id") is not None
        )
        raw_memory_count = len(memories)
        memories = _rebalance_locked_memory_candidates(memories, inject_limit, locked_ratio, semantic_threshold)
        locked_hit_ids = [
            m["id"] for m in memories
            if m.get("is_permanent") and m.get("id") is not None
        ]
        if locked_hit_ids:
            try:
                touched = await touch_permanent_memories(locked_hit_ids)
                if touched:
                    print(f"📌 注入了 {touched} 条锁定记忆（已刷新 last_accessed）")
            except Exception as e:
                print(f"locked memory touch failed (chat continues): {type(e).__name__}: {e}")
        if raw_memory_count > inject_limit or semantic_locked_candidate_count:
            print(
                f"📌 锁定保底重组：候选 {raw_memory_count} 条，语义锁定候选 {semantic_locked_candidate_count} 条，"
                f"ratio={locked_ratio:.2f}，最终注入候选 {len(memories)}/{inject_limit}"
            )
        
        if memories:
            memory_lines = []
            injected_ids = []
            for mem in memories:
                title = mem.get("title", "")
                heat = mem.get("heat", 1.0)
                date_tag = ""
                if mem.get("created_at"):
                    try:
                        dt = mem["created_at"]
                        if hasattr(dt, "strftime"):
                            date_tag = f"[{dt.strftime('%Y-%m-%d')}]"
                        else:
                            date_tag = f"[{str(dt)[:10]}]"
                    except Exception:
                        pass
                
                cat_name = mem.get("category_name", "")
                cat_tag = f"({cat_name})" if cat_name else ""
                is_locked = bool(mem.get("is_permanent"))
                locked_tag = "【用户标记为重要】" if is_locked else ""
                
                # v5.4 热度分档注入（阈值可配置）
                # 防御性冗余：calculate_heat 已对 is_permanent 返回 1.0，
                # 此处为注入层的行为兜底（防未来改动 heat 计算时漏掉锁定分支），勿删。
                if is_locked or heat > th_high:
                    if title:
                        memory_lines.append(f"- {date_tag}{cat_tag}{locked_tag}【{title}】{mem['content']}")
                    else:
                        memory_lines.append(f"- {date_tag}{cat_tag}{locked_tag} {mem['content']}")
                    if mem.get("id") is not None and not is_locked:
                        injected_ids.append(mem["id"])
                elif heat > th_medium:
                    if title:
                        brief = mem['content'][:truncate_len] + "…" if len(mem['content']) > truncate_len else mem['content']
                        memory_lines.append(f"- {date_tag}{cat_tag}{locked_tag}【{title}】{brief}（印象模糊）")
                    else:
                        brief = mem['content'][:truncate_len] + "…" if len(mem['content']) > truncate_len else mem['content']
                        memory_lines.append(f"- {date_tag}{cat_tag}{locked_tag} {brief}（印象模糊）")
                    if mem.get("id") is not None and not is_locked:
                        injected_ids.append(mem["id"])
                
            memory_text = "\n".join(memory_lines)
            
            if memory_lines:
                _append_runtime_context_section(
                    prompt_meta,
                    f"【从过往对话中检索到的相关记忆】\n以下是与当前话题可能相关的历史信息，自然地融入对话中，不要刻意提起'我记得'：\n{memory_text}",
                )
                skipped = len(memories) - len(memory_lines)
                skip_msg = f"（跳过 {skipped} 条低热度）" if skipped > 0 else ""
                print(f"📚 注入了 {len(memory_lines)} 条相关记忆{skip_msg}（热度分档注入）")
            else:
                print(f"📚 搜到 {len(memories)} 条记忆但全部热度过低，跳过注入")

            if injected_ids:
                try:
                    await track_memory_recall(injected_ids, user_message, conversation_id=conversation_id)
                except Exception as e:
                    print(f"⚠️  记忆召回追踪失败（不影响聊天）: {type(e).__name__}: {e}")
        
        # ---- Dream 场景整合认知（动态，复用本轮 query embedding，不额外生成 embedding）----
        if _query_emb:
            try:
                scene_enabled = await get_config_bool("scene_inject_enabled", fallback=True)
                if scene_enabled:
                    scene_limit = max(0, await get_config_int("scene_inject_limit", fallback=2))
                    scene_min_sim = await get_config_float("scene_inject_min_sim", fallback=0.5)
                    if scene_limit > 0:
                        scenes = await search_scenes(_query_emb, limit=scene_limit, min_sim=scene_min_sim)
                        if scenes:
                            limited_scenes = scenes[:scene_limit]
                            scene_lines = []
                            for scene in limited_scenes:
                                title = (scene.get("title") or f"场景 #{scene.get('id')}").strip()
                                facts = _format_scene_field(scene.get("atomic_facts"))
                                foresight = _format_scene_field(_filter_valid_foresight(scene.get("foresight")))
                                scene_lines.append(f"◈ {title}")
                                scene_lines.append(f"  事实：{facts}")
                                if foresight:
                                    scene_lines.append(f"  前瞻：{foresight}")
                            _append_runtime_context_section(
                                prompt_meta,
                                "【记忆深处的整合认知（梦境整理沉淀）】\n" + "\n".join(scene_lines),
                            )

                            print(f"🧩 注入了 {len(limited_scenes)} 个相关记忆场景")
            except Exception as e:
                print(f"⚠️  场景搜索注入失败（不影响聊天）: {type(e).__name__}: {e}")

        # ---- ⑥ 项目文件相关片段（动态，每轮根据用户消息搜索）----
        if project_id:
            try:
                from database import search_file_chunks
                file_chunks = await search_file_chunks(project_id, user_message, limit=6)
                if file_chunks:
                    chunk_lines = []
                    for chunk in file_chunks:
                        chunk_lines.append(f"📎 [{chunk['file_name']}] {chunk['content']}")
                    chunk_text = "\n".join(chunk_lines)
                    _append_runtime_context_section(
                        prompt_meta,
                        f"【项目文件中的相关内容】\n{chunk_text}",
                    )
                    print(f"📂 注入了 {len(file_chunks)} 条文件片段")
            except Exception as e:
                print(f"⚠️  文件搜索失败: {e}")
        
        # ---- ⑦ Dream 犯困提示（动态）----
        try:
            from dream import get_drowsy_prompt
            drowsy = await get_drowsy_prompt()
            if drowsy:
                _append_runtime_context_section(prompt_meta, drowsy)
                print(f"😴 注入了犯困提示")
        except Exception:
            pass
        
        # ---- ⑥ 无缝换窗 v2 ----
        try:
            handoff_on = await get_config_bool("handoff_enabled", fallback=True)
            if handoff_on and user_msg_count == 1 and not is_regenerate:
                from database import get_handoff_source, get_handoff_data

                source = await get_handoff_source(
                    exclude_conversation_id=conversation_id,
                    project_id=project_id,
                )
                if source:
                    tail_count = await get_config_int("handoff_tail_count", fallback=6)
                    data = await get_handoff_data(source["id"], tail_count=tail_count)

                    existing_summary = None
                    msgs_to_compress = []

                    if data["has_divider"]:
                        if data["divider_summary"]:
                            existing_summary = data["divider_summary"]
                        msgs_to_compress = data["uncompressed"]
                    elif data["comp_summary"]:
                        existing_summary = data["comp_summary"]
                    else:
                        msgs_to_compress = data["all_messages"]

                    handoff_summary = None
                    if existing_summary and not msgs_to_compress:
                        handoff_summary = existing_summary
                    elif existing_summary and msgs_to_compress:
                        if len(msgs_to_compress) <= tail_count:
                            handoff_summary = existing_summary
                        else:
                            handoff_summary = await _compress_for_handoff(existing_summary, msgs_to_compress)
                            if not handoff_summary:
                                handoff_summary = existing_summary  # 压缩失败/超时：退回已有概要，强于只带尾巴
                    elif msgs_to_compress:
                        if len(msgs_to_compress) <= 10:
                            handoff_summary = "\n".join(
                                f"{'用户' if m['role'] == 'user' else '助手'}: {(m.get('content') or '')[:500]}"
                                for m in msgs_to_compress
                            )
                        else:
                            handoff_summary = await _compress_for_handoff(None, msgs_to_compress)

                    tail_lines = []
                    for m in data["tail_messages"]:
                        role_label = "用户" if m["role"] == "user" else "助手"
                        content = m.get("content", "")
                        if len(content) > 500:
                            content = content[:500] + "…（截断）"
                        if content:
                            tail_lines.append(f"{role_label}: {content}")
                    tail_text = "\n".join(tail_lines)

                    summary_parts = []
                    if handoff_summary:
                        summary_parts.append(f"[全程概要]\n{handoff_summary}")
                    if tail_text:
                        summary_parts.append(f"[结尾原文]\n{tail_text}")
                    full_summary = "\n\n".join(summary_parts).strip()

                    if full_summary:
                        title_hint = source.get("title", "") or ""
                        usage_rule = "仅供了解上下文背景。用户延续话题时自然参考，开启新话题时安静忽略，不要主动提起上一窗口。"
                        _append_runtime_context_section(
                            prompt_meta,
                            f"【上一个对话衔接（{title_hint}）】\n{usage_rule}\n{full_summary}",
                        )
                        prompt_meta["handoff"] = {
                            "version": 2,
                            "sourceId": source["id"],
                            "sourceTitle": title_hint,
                            "summary": full_summary,
                            "status": "full" if handoff_summary else "tail_only",
                        }
                        status = "概要+原文" if handoff_summary else "仅原文"
                        print(f"🔗 无缝换窗 v2：注入{status}，来源={source['id']}")
        except Exception as e:
            print(f"⚠️  无缝换窗 v2 失败: {e}")
        
        return active_prompt, prompt_meta
        
    except Exception as e:
        print(f"⚠️  记忆检索失败: {e}，使用纯人设")
        return active_prompt, prompt_meta


# ============================================================
# 后台记忆处理
# ============================================================

# ============================================================
# 情绪检测（v5.2 热度系统）
# ============================================================

# 用户消息中的情绪关键词（规则引擎兜底）
EMOTION_HIGH_KEYWORDS = [
    "抱抱", "贴贴", "亲亲", "呜", "哭", "崩溃", "撑不住", "好难过",
    "好开心", "好幸福", "我爱你", "谢谢你", "你真好", "好想你",
    "对不起", "害怕", "不想活", "好累", "受不了", "心疼",
    "太好了", "我好高兴", "感动", "哭了",
]
EMOTION_MEDIUM_KEYWORDS = [
    "难过", "开心", "紧张", "焦虑", "生气", "委屈", "失落",
    "高兴", "感谢", "抱歉", "担心", "烦", "郁闷", "不舒服",
]


def detect_emotion_from_user_msg(text: str) -> str:
    """从用户消息检测情绪级别（规则引擎）"""
    if not text:
        return "normal"
    for kw in EMOTION_HIGH_KEYWORDS:
        if kw in text:
            return "high"
    for kw in EMOTION_MEDIUM_KEYWORDS:
        if kw in text:
            return "medium"
    return "normal"


def detect_emotion_from_response(text: str) -> str:
    """从模型回复中解析隐藏的情绪标记 <!--emotion:高-->"""
    if not text:
        return "normal"
    import re
    match = re.search(r'<!--\s*emotion\s*[:：]\s*(高|中|low|medium|high)\s*-->', text)
    if match:
        level = match.group(1)
        if level in ("高", "high"):
            return "high"
        elif level in ("中", "medium"):
            return "medium"
    return "normal"


def strip_emotion_tag(text: str) -> str:
    """滤掉回复里的情绪隐藏标记（任何值），用于存储和提取前的清洗。
    情绪解析须在调用本函数之前完成（解析依赖原始带标记文本）。"""
    if not text:
        return text
    import re
    return re.sub(r'<!--\s*emotion\s*[:：][^>]*-->', '', text).rstrip()


def detect_dream_trigger(text: str) -> bool:
    """检测模型回复中是否有 <!--dream:trigger--> 标记"""
    if not text:
        return False
    import re
    return bool(re.search(r'<!--\s*dream\s*[:：]\s*trigger\s*-->', text))


def strip_dream_tag(text: str) -> str:
    """滤掉回复里的 Dream 触发标记，用于存储前的清洗"""
    if not text:
        return text
    import re
    return re.sub(r'<!--\s*dream\s*[:：]\s*trigger\s*-->', '', text).rstrip()


def merge_emotion_levels(user_level: str, response_level: str) -> str:
    """取两个情绪级别的高值"""
    order = {"normal": 0, "medium": 1, "high": 2}
    return max([user_level, response_level], key=lambda x: order.get(x, 0))


def emotion_to_weight(level: str) -> int:
    """情绪级别转数字权重（0-10）"""
    return {"high": 8, "medium": 4, "normal": 0}.get(level, 0)


# 主动记忆触发词 —— 用户说了这些词就立刻提取，不等计数器
MEMORY_TRIGGER_WORDS = ["记住", "记下", "帮我记", "请记", "别忘了", "不要忘记", "你要记得", "记一下"]


async def process_memories_background(session_id: str, user_msg: str, assistant_msg: str, model: str, emotion_level: str = "normal", project_id: str = None, is_regenerate: bool = False):
    """
    后台异步：存储对话 + 按间隔提取记忆（不阻塞主流程）
    
    v2.4 改进：
    - 提取前的对比范围改为「搜索相关 + 最近记忆」组合，覆盖种子记忆
    - 存储前逐条做去重检测，防止冗余写入
    
    v2.5 改进：
    - 检测主动记忆触发词，命中时立即提取，不等计数器
    - 不重置计数器，不干扰正常的定时提取节奏
    
    v3.7 改进：
    - 提取时只从当前 session 捞最近 N 轮完整对话（而不是混合所有窗口）
    - N = extract_interval，攒几轮就提取几轮
    
    v5.2 改进：
    - 接受 emotion_level 参数，传递给 save_memory 的 emotional_weight
    
    v5.8 改进：
    - 接受 project_id 参数，项目内对话提取的记忆自动打上 project_id 标签
    """
    global _conversation_counter

    # 情绪解析已在调用方用原始文本完成（emotion_level 已传入），
    # 这里把标记滤掉，保证存储和提取用的都是干净文本。
    assistant_msg = strip_emotion_tag(assistant_msg)
    assistant_msg = strip_dream_tag(assistant_msg)

    try:
        # 对话始终保存
        if is_regenerate:
            await delete_latest_assistant_message(session_id)
        else:
            await save_message(session_id, "user", user_msg, model)
        await save_message(session_id, "assistant", assistant_msg, model)

        # 项目对话默认不提取碎片（对话已保存，但不走记忆提取流程）
        # 未来做"碎片进全局"开关后，这里加条件判断
        if project_id:
            print(f"📂 项目对话，跳过记忆提取（project_id={project_id}）")
            return {"action": "skip_project", "project_id": project_id}

        # 检测用户是否主动要求记忆
        force_extract = any(kw in user_msg for kw in MEMORY_TRIGGER_WORDS)

        # 使用锁保护计数器，防止并发请求导致重复提取或跳过
        should_extract = False
        async with _counter_lock:
            if not is_regenerate:
                _conversation_counter += 1
            extract_interval = await get_extract_interval()
            if _conversation_counter < extract_interval and not force_extract:
                print(f"💬 对话已保存（{_conversation_counter}/{extract_interval}轮后提取记忆）")
                return {"action": "skip", "counter": _conversation_counter, "interval": extract_interval}

            if force_extract:
                print(f"🎯 检测到主动记忆请求，立即提取（计数器保持 {_conversation_counter}/{extract_interval}）")
            else:
                _conversation_counter = 0
                print(f"🧠 达到提取间隔（{extract_interval}轮），开始提取记忆...")
            should_extract = True
        
        # ===== v2.4 改进：组合式获取已有记忆 =====
        # 用当前对话内容搜索相关记忆（能覆盖到种子记忆）
        # track_recall=False: 这里是去重对比，不是用户聊天，不应该增加召回计数
        related = await search_memories(user_msg, limit=50, track_recall=False, project_id=project_id)
        related_contents = [r["content"] for r in related]
        
        # 再补充最近的记忆（防止遗漏新存的）
        recent = await get_recent_memories(limit=30, project_id=project_id)
        recent_contents = [r["content"] for r in recent]
        
        # 合并去重
        seen = set()
        existing_contents = []
        for content in related_contents + recent_contents:
            if content not in seen:
                seen.add(content)
                existing_contents.append(content)
        
        print(f"📋 对比范围：{len(existing_contents)} 条已有记忆（搜索相关 {len(related_contents)} + 最近 {len(recent_contents)}，去重后 {len(existing_contents)}）")
        
        # 只读取当前 session。全局“最近 N 条”会在多窗口并行时把互不相干的
        # 对话拼成一段从未发生过的经历。
        recent_msgs = await get_recent_messages(session_id, limit=extract_interval * 2)
        
        if recent_msgs:
            messages_for_extraction = [
                {"role": row["role"], "content": row["content"]}
                for row in recent_msgs
            ]
            print(f"📨 提取范围：当前会话最近 {len(messages_for_extraction)} 条消息（约 {len(messages_for_extraction)//2} 轮对话）")
        else:
            # 降级：如果数据库查不到，至少用当前这一轮
            messages_for_extraction = [
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ]
            print(f"📨 提取范围：当前 1 轮对话（降级）")
        
        # 获取可用分类名（用于自动归类）
        try:
            all_cats = await get_all_categories()
            cat_names = [c["name"] for c in all_cats]
        except Exception:
            cat_names = []
        
        # 读取数据库中的模型和提示词配置
        from config import get_config
        db_memory_model = await get_config("default_memory_model")
        db_memory_prompt = await get_config("prompt_memory_extract")
        
        new_memories = await extract_memories(
            messages_for_extraction,
            existing_memories=existing_contents,
            categories=cat_names,
            model_override=db_memory_model if db_memory_model else None,
            prompt_override=db_memory_prompt if db_memory_prompt else None,
            emotion_level=emotion_level,
        )
        
        # 过滤垃圾记忆（不靠模型自觉，硬过滤）
        META_BLACKLIST = [
            "记忆库", "记忆系统", "检索", "没有被记录", "没有被提取",
            "记忆遗漏", "尚未被记录", "写入不完整", "检索功能",
            "系统没有返回", "关键词匹配", "语义匹配", "语义检索",
            "阈值", "数据库", "seed", "导入", "部署",
            "bug", "debug", "端口", "网关",
        ]
        
        filtered_memories = []
        for mem in new_memories:
            content = mem["content"]
            if any(kw in content for kw in META_BLACKLIST):
                print(f"🚫 过滤掉meta记忆: {content[:60]}...")
                continue
            filtered_memories.append(mem)
        
        # ===== v5.3 改进：去重 + 矛盾检测（共用一次搜索）=====
        saved_count = 0
        skipped_count = 0
        contradiction_count = 0
        saved_items = []  # 收集保存的记忆内容（供事件 payload 使用）

        for mem in filtered_memories:
            # 去重检测（v5.4：传标题，标题不同时不误杀；v5.8：按 project_id 作用域去重）
            is_dup, similar_results = await check_memory_duplicate(mem["content"], new_title=mem.get("title", ""), project_id=project_id)

            if is_dup:
                skipped_count += 1
                continue

            # 按作用域过滤矛盾候选：项目碎片只能替代同项目的旧碎片，全局只看全局
            if project_id:
                scoped_results = [m for m in similar_results if m.get("project_id") == project_id]
            else:
                scoped_results = [m for m in similar_results if m.get("project_id") is None]

            # 矛盾检测（v5.3：复用去重搜索结果，不额外调 embedding API）
            contradicted_ids = detect_contradictions(
                mem.get("title", ""), mem["content"], scoped_results
            )

            # 自动匹配分类
            cat_id = None
            cat_hint = mem.get("category", "")
            if cat_hint:
                cat_id = await match_category_by_name(cat_hint)

            # 保存新记忆（v5.3：返回 ID，用于创建 supersedes edge）
            new_id = await save_memory(
                content=mem["content"],
                importance=mem["importance"],
                source_session=session_id,
                title=mem.get("title", ""),
                category_id=cat_id,
                source="ai_extracted",
                emotional_weight=mem.get("emotional_weight", 0) or emotion_to_weight(emotion_level),
                project_id=project_id,
                memory_kind=mem.get("memory_kind"),
                enforce_identity=True,
            )
            saved_count += 1
            saved_items.append({"title": mem.get("title", ""), "content": mem["content"][:120]})

            # 处理矛盾：标旧记忆失效 + 创建 supersedes edge
            if contradicted_ids and new_id:
                for old_id in contradicted_ids:
                    if not await invalidate_memory(old_id, reason=f"被新记忆 #{new_id} 替代"):
                        print(f"⚠️ 记忆 #{old_id} 失效未命中，仍建 supersedes 边（矛盾计数可能偏高）")
                    await create_memory_edge(
                        new_id, "memory", old_id, "memory", "supersedes",
                        reason="提取时自动检测到信息更新", created_by="extractor"
                    )
                    contradiction_count += 1

        if saved_count > 0 or skipped_count > 0:
            total = await get_all_memories_count()
            contra_msg = f"，{contradiction_count} 条旧记忆被替代" if contradiction_count > 0 else ""
            print(f"💾 提取结果：{saved_count} 条新记忆已保存，{skipped_count} 条重复已跳过{contra_msg}，总计 {total} 条")
            return {"action": "extract", "saved": saved_count, "skipped": skipped_count, "contradictions": contradiction_count, "total": total, "items": saved_items}
        else:
            print(f"💭 本轮对话未产生新记忆")
            return {"action": "extract", "saved": 0, "skipped": 0, "contradictions": 0, "total": await get_all_memories_count(), "items": []}
            
    except Exception as e:
        print(f"⚠️  后台记忆处理失败: {e}")
        return {"action": "error", "error": str(e)}


# ============================================================
# API 接口
# ============================================================

@app.get("/memory/v1/whoami")
async def memory_whoami(request: Request):
    """Confirm the canonical identity without exposing which door authenticated."""
    context = _memory_client_context(request)
    return {
        "assistant_identity_id": context.assistant_identity_id,
        "memory_space_id": context.memory_space_id,
    }


@app.post("/memory/v1/events/ingest")
async def memory_ingest_event(request: Request):
    """Append raw dialogue evidence; retries are idempotent by event_id."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    event = _prepare_memory_event_payload(payload)
    try:
        result = await ingest_memory_event(
            event_id=event["event_id"],
            source_client=context.client_id,
            conversation_id=event["conversation_id"],
            role=event["role"],
            content=event["content"],
            occurred_at=event["occurred_at"],
            metadata=event["metadata"],
            memory_space_id=context.memory_space_id,
            assistant_identity_id=context.assistant_identity_id,
        )
    except ArchiveIdentifierConfigError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    # Source metadata stays available for audit in PostgreSQL, never in recalled prose.
    return {"created": bool(result.get("created")), "event_id": result["event_id"]}


@app.post("/memory/v1/events/ingest-batch")
async def memory_ingest_event_batch(request: Request):
    """Atomically replay up to 500 archived messages after an offline period."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)
    raw_events = payload.get("events")
    if not isinstance(raw_events, list) or not raw_events:
        raise HTTPException(status_code=422, detail="events must be a non-empty array")
    if len(raw_events) > 500:
        raise HTTPException(status_code=422, detail="events exceeds batch limit 500")
    events = [_prepare_memory_event_payload(event) for event in raw_events]
    if len({event["event_id"] for event in events}) != len(events):
        raise HTTPException(status_code=422, detail="event_id must be unique within a batch")
    try:
        results = await ingest_memory_events(
            events,
            source_client=context.client_id,
            memory_space_id=context.memory_space_id,
            assistant_identity_id=context.assistant_identity_id,
        )
    except ArchiveIdentifierConfigError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    created = sum(1 for result in results if result.get("created"))
    return {
        "received": len(results),
        "created": created,
        "duplicates": len(results) - created,
        "event_ids": [result["event_id"] for result in results],
    }


@app.get("/memory/v1/archive/conversations")
async def memory_archive_conversations(
    request: Request,
    limit: int = 50,
    cursor: str = "",
):
    """List complete archived conversations without exposing entry metadata."""
    context = _memory_client_context(request)
    limit = max(1, min(int(limit), 200))
    if "source_client" in request.query_params:
        raise HTTPException(status_code=422, detail="source_client is internal audit metadata")
    before_at, before_id = _decode_archive_cursor(cursor)
    rows = await list_memory_event_conversations(
        memory_space_id=context.memory_space_id,
        limit=limit + 1,
        before_at=before_at,
        before_conversation_id=before_id,
    )
    _assert_current_archive_rows(rows)
    has_more = len(rows) > limit
    visible = rows[:limit]
    next_cursor = None
    if has_more and visible:
        next_cursor = _encode_archive_cursor(visible[-1]["last_event_at"], visible[-1]["conversation_id"])
    conversations = []
    for row in visible:
        item = _public_archive_record(row, _PUBLIC_ARCHIVE_CONVERSATION_FIELDS)
        item["preview"] = str(item.get("preview") or "")[:500]
        conversations.append(item)
    return {"conversations": conversations, "next_cursor": next_cursor}


@app.get("/memory/v1/archive/conversations/{conversation_id}")
async def memory_archive_conversation(
    conversation_id: str,
    request: Request,
    limit: int = 100,
    cursor: str = "",
    include_system: bool = False,
):
    """Read one transcript page; rows are immutable and never rewritten by Dream."""
    context = _memory_client_context(request)
    conversation_id = str(conversation_id or "").strip()
    if not conversation_id or len(conversation_id) > 200:
        raise HTTPException(status_code=422, detail="invalid conversation_id")
    if not is_opaque_archive_identifier(conversation_id, "conversation"):
        raise HTTPException(status_code=422, detail="conversation_id must be an opaque archive reference")
    limit = max(1, min(int(limit), 500))
    before_at, before_id = _decode_archive_cursor(cursor)
    rows = await get_memory_event_conversation_page(
        conversation_id,
        memory_space_id=context.memory_space_id,
        limit=limit + 1,
        before_at=before_at,
        before_event_id=before_id,
        include_system=include_system,
    )
    _assert_current_archive_rows(rows)
    has_more = len(rows) > limit
    visible = rows[:limit]
    next_cursor = None
    if has_more and visible:
        next_cursor = _encode_archive_cursor(visible[-1]["occurred_at"], visible[-1]["event_id"])
    return {
        "conversation_id": conversation_id,
        "events": [
            _public_archive_record(row, _PUBLIC_ARCHIVE_EVENT_FIELDS)
            for row in reversed(visible)
        ],
        "page_order": "oldest_to_newest",
        "next_cursor": next_cursor,
    }


@app.post("/memory/v1/archive/search")
async def memory_archive_search(request: Request):
    """Search exact words inside the immutable original transcript archive."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)
    query = _required_memory_text(payload, "query", 10000)
    try:
        limit = max(1, min(int(payload.get("limit", 50)), 200))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="limit must be an integer")
    role = str(payload.get("role") or "").strip().lower() or None
    if role not in {None, "user", "assistant", "system"}:
        raise HTTPException(status_code=422, detail="role must be user, assistant, or system")
    before_at, before_id = _decode_archive_cursor(str(payload.get("cursor") or ""))
    conversation_filter = str(payload.get("conversation_id") or "").strip() or None
    if conversation_filter and len(conversation_filter) > 200:
        raise HTTPException(status_code=422, detail="conversation_id exceeds 200 characters")
    if conversation_filter and not is_opaque_archive_identifier(conversation_filter, "conversation"):
        raise HTTPException(status_code=422, detail="conversation_id must be an opaque archive reference")
    rows = await search_memory_events(
        query,
        memory_space_id=context.memory_space_id,
        limit=limit + 1,
        before_at=before_at,
        before_event_id=before_id,
        conversation_id=conversation_filter,
        role=role,
    )
    _assert_current_archive_rows(rows)
    has_more = len(rows) > limit
    visible = rows[:limit]
    next_cursor = None
    if has_more and visible:
        next_cursor = _encode_archive_cursor(visible[-1]["occurred_at"], visible[-1]["event_id"])
    return {
        "matches": [
            _public_archive_record(row, _PUBLIC_ARCHIVE_EVENT_FIELDS)
            for row in visible
        ],
        "next_cursor": next_cursor,
    }


@app.post("/memory/v1/memories/remember")
async def memory_remember(request: Request):
    """Store an explicit semantic memory in the assistant's own voice."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)

    content = _required_memory_text(payload, "content", 50000)
    title = str(payload.get("title") or "").strip()
    if len(title) > 500:
        raise HTTPException(status_code=422, detail="title exceeds 500 characters")
    memory_kind = str(payload.get("memory_kind") or "relationship").strip().lower()
    voice_errors = validate_autobiographical_memory(content, title, memory_kind)
    if voice_errors:
        raise HTTPException(status_code=422, detail=voice_errors)
    try:
        importance = int(payload.get("importance", 5))
        emotional_weight = int(payload.get("emotional_weight", 0))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="importance and emotional_weight must be integers")
    if not 1 <= importance <= 10 or not 0 <= emotional_weight <= 10:
        raise HTTPException(status_code=422, detail="importance must be 1-10 and emotional_weight must be 0-10")
    conversation_id = str(payload.get("conversation_id") or "").strip()
    if len(conversation_id) > 200:
        raise HTTPException(status_code=422, detail="conversation_id exceeds 200 characters")

    memory_id = await save_memory(
        content=content,
        importance=importance,
        source_session=conversation_id,
        title=title,
        source="assistant_explicit",
        emotional_weight=emotional_weight,
        memory_space_id=context.memory_space_id,
        assistant_identity_id=context.assistant_identity_id,
        source_client=context.client_id,
        memory_kind=memory_kind,
        enforce_identity=True,
    )
    return {"memory_id": memory_id, "stored": True}


@app.post("/memory/v1/recall")
async def memory_recall(request: Request):
    """Recall from the one shared space, regardless of which door asks."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)
    query = _required_memory_text(payload, "query", 10000)
    try:
        limit = max(1, min(int(payload.get("limit", 10)), 50))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="limit must be an integer")
    memories = await search_memories(
        query,
        limit=limit,
        track_recall=True,
        memory_space_id=context.memory_space_id,
    )
    return {
        "identity_instruction": RECALL_IDENTITY_INSTRUCTION,
        "memories": memories,
    }


@app.post("/memory/v1/handoff")
async def memory_handoff(request: Request):
    """Resume the latest other conversation as the same self."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)
    current_conversation_id = str(payload.get("current_conversation_id") or "").strip() or None
    if current_conversation_id and len(current_conversation_id) > 200:
        raise HTTPException(status_code=422, detail="current_conversation_id exceeds 200 characters")
    try:
        tail_count = max(1, min(int(payload.get("tail_count", 6)), 50))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="tail_count must be an integer")
    try:
        handoff = await get_memory_handoff(
            current_conversation_id=current_conversation_id,
            tail_count=tail_count,
            memory_space_id=context.memory_space_id,
        )
    except ArchiveIdentifierConfigError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    return {
        "identity_instruction": RECALL_IDENTITY_INSTRUCTION,
        "handoff": handoff,
    }


@app.post("/memory/v1/history/search")
async def memory_history_search(request: Request):
    """Explicitly inspect archived history without restoring it to active recall."""
    context = _memory_client_context(request)
    payload = await request.json()
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="request body must be a JSON object")
    _reject_memory_identity_override(payload)
    query = str(payload.get("query") or "").strip()
    if len(query) > 10000:
        raise HTTPException(status_code=422, detail="query exceeds 10000 characters")
    try:
        limit = max(1, min(int(payload.get("limit", 20)), 200))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="limit must be an integer")
    memories = await search_archived_memories(
        query=query,
        limit=limit,
        memory_space_id=context.memory_space_id,
    )
    return {
        "identity_instruction": RECALL_IDENTITY_INSTRUCTION,
        "memories": memories,
    }


@app.get("/")
async def root_status():
    """根路由 — 返回系统状态 JSON（admin面板依赖此接口获取统计数据）"""
    memory_count = 0
    mem_enabled = await get_memory_enabled()
    if mem_enabled:
        try:
            memory_count = await get_all_memories_count()
        except Exception:
            pass
    return {
        "status": "running",
        "gateway": "Kiwi-Mem v1.3.0",
        "version": "Kiwi-Mem v1.3.0",
        "memory_enabled": mem_enabled,
        "memory_count": memory_count,
        # 前端 admin-panel 读 status.memories, 加别名避免显示 '-'
        "memories": memory_count,
        "max_inject": await get_max_inject(),
        "default_model": DEFAULT_MODEL,
        "extract_interval": await get_extract_interval(),
    }


@app.get("/favicon.ico")
async def favicon():
    """返回空favicon防止404"""
    return Response(status_code=204)


@app.post("/auth/verify")
async def auth_verify(request: Request):
    """Verify the tab-scoped admin Bearer secret against its configured digest."""
    try:
        authenticated = authenticate_admin_token(bearer_token(request.headers.get("authorization", "")))
    except AdminTokenConfigError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    if not authenticated:
        raise HTTPException(
            status_code=401,
            detail="invalid admin bearer token",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return {"status": "ok", "message": "admin token verified"}


@app.get("/api/status")
async def api_status():
    """系统状态（JSON）— 兼容旧接口，重定向到根路由"""
    return await root_status()


@app.get("/admin")
async def admin_panel():
    """返回管理面板 HTML（admin-panel/index.html）。

    v1.1 重构时这里被改成了返回 JSON, 直接访问 /admin 看到 {"status": "running"}
    而不是面板页面。改回 FileResponse, 文件不存在时降级返回 JSON。
    """
    from fastapi.responses import FileResponse
    panel_path = os.path.join(os.path.dirname(__file__), "admin-panel", "index.html")
    if os.path.exists(panel_path):
        return FileResponse(panel_path, media_type="text/html; charset=utf-8")
    return {"status": "running", "service": "kiwi-mem", "warning": "admin-panel/index.html not found"}


@app.get("/v1/models")
async def list_models():
    """对外暴露的模型列表，供前端客户端拉取下拉选项。

    口径优先级：
      1) 管理面板里配置的供应商模型（已启用供应商 + 实际保存的模型）——这与聊天
         路由 resolve_provider_for_model 完全一致，前端看到的就是真正能用的模型。
      2) 没有配置任何模型时（例如只用 .env 直连、未加供应商），回退到旧行为：
         从环境变量 API_BASE_URL 对应的服务商拉取 /models。
      3) 都拿不到时，用 DEFAULT_MODEL 兜底，保证至少能选一个发出去。
    """
    # ── 1) 管理面板配置的供应商模型（首选）──
    try:
        saved = await get_enabled_provider_models()
    except Exception as e:
        print(f"⚠️ 读取已配置模型失败，回退环境变量: {e}")
        saved = []

    if saved:
        data = []
        seen = set()
        for m in saved:
            mid = (m.get("model_id") or "").strip()
            # 同一模型可能被挂在多个供应商下，路由是 LIMIT 1，这里也去重保持列表干净
            if not mid or mid in seen:
                continue
            seen.add(mid)
            data.append({
                "id": mid,
                "object": "model",
                "created": 1700000000,
                "owned_by": m.get("provider_name") or "kiwi-mem",
            })
        if data:
            return {"object": "list", "data": data}

    # ── 2) 回退：从环境变量配置的服务商拉取（兼容纯 .env 部署）──
    try:
        # 从 API_BASE_URL 提取基础地址（去掉 /chat/completions 部分）
        base = API_BASE_URL.split("/chat/completions")[0].rstrip("/")
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{base}/models",
                headers={"Authorization": f"Bearer {API_KEY}"},
            )
            if resp.status_code == 200:
                return resp.json()
    except Exception as e:
        print(f"⚠️ 拉取模型列表失败: {e}")

    # ── 3) 兜底：默认模型 ──
    return {
        "object": "list",
        "data": [
            {
                "id": DEFAULT_MODEL,
                "object": "model",
                "created": 1700000000,
                "owned_by": "kiwi-mem",
            }
        ],
    }


# ============================================================
# 文件内容提取
# ============================================================

@app.post("/v1/files/extract")
async def extract_file_content(file: UploadFile = File(...)):
    """提取上传文件的文本内容（PDF/DOCX/XLSX/ZIP等）"""
    import io
    import tempfile
    
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    content_bytes = await file.read()
    
    try:
        extracted = ""
        file_type = ext
        
        # PDF
        if ext == "pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(io.BytesIO(content_bytes))
                pages = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"[第{i+1}页]\n{text.strip()}")
                extracted = "\n\n".join(pages) if pages else "(PDF 无法提取文字，可能是扫描件)"
            except Exception as e:
                extracted = f"(PDF 解析失败: {str(e)})"
        
        # Word DOCX
        elif ext == "docx":
            try:
                from docx import Document
                doc = Document(io.BytesIO(content_bytes))
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                extracted = "\n\n".join(paragraphs) if paragraphs else "(DOCX 无内容)"
            except Exception as e:
                extracted = f"(DOCX 解析失败: {str(e)})"
        
        # Excel XLSX
        elif ext in ("xlsx", "xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        if row_str.strip():
                            rows.append(row_str)
                    if rows:
                        sheets.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows[:500]))  # 限制行数
                extracted = "\n\n".join(sheets) if sheets else "(XLSX 无内容)"
                wb.close()
            except Exception as e:
                extracted = f"(XLSX 解析失败: {str(e)})"
        
        # ZIP — 列出文件列表 + 提取文本文件内容
        elif ext == "zip":
            import zipfile
            try:
                zf = zipfile.ZipFile(io.BytesIO(content_bytes))
                file_list = zf.namelist()
                text_extensions = {'.txt', '.md', '.py', '.js', '.jsx', '.ts', '.tsx', '.json', '.csv', '.html', '.css', '.xml', '.yaml', '.yml', '.toml', '.sh', '.sql', '.java', '.c', '.cpp', '.go', '.rs', '.rb', '.log', '.ini', '.cfg', '.env'}
                
                parts = [f"压缩包共 {len(file_list)} 个文件：\n" + "\n".join(f"  {f}" for f in file_list[:100])]
                
                # 提取小的文本文件
                for name in file_list[:20]:
                    name_lower = name.lower()
                    if any(name_lower.endswith(e) for e in text_extensions):
                        info = zf.getinfo(name)
                        if info.file_size < 50000:  # 小于50KB
                            try:
                                text = zf.read(name).decode("utf-8", errors="ignore")
                                parts.append(f"\n[文件: {name}]\n{text}")
                            except Exception:
                                pass
                
                extracted = "\n".join(parts)
                zf.close()
            except Exception as e:
                extracted = f"(ZIP 解析失败: {str(e)})"
        
        # 其他文本类格式尝试直接读
        else:
            try:
                extracted = content_bytes.decode("utf-8", errors="ignore")
                if not extracted.strip() or '\x00' in extracted[:200]:
                    extracted = f"(二进制文件，无法提取文本内容)"
                    file_type = "binary"
            except Exception:
                extracted = f"(无法解析该文件格式)"
                file_type = "binary"
        
        # 截断过长的内容
        if len(extracted) > 100000:
            extracted = extracted[:100000] + f"\n\n...(内容过长，已截断至约10万字符)"
        
        return {
            "filename": filename,
            "type": file_type,
            "size": len(content_bytes),
            "text": extracted,
        }
    
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": f"文件处理失败: {str(e)}"}
        )


def _apply_reasoning(body: dict, is_openrouter: bool, is_anthropic_fmt: bool, reasoning_effort, skip_prompt: bool = False):
    """统一决定一个出站请求体的思考链参数。转发路径与工具循环共用，保证两条路对所有供应商一致。

    reasoning_effort 是从请求体里 pop 出来的原值：off/auto/low/medium/high/None。
      - 功能调用(skip_prompt) 或用户选 'off' → 不开思考（并清掉任何 reasoning/reasoning_effort 残留）
      - OpenRouter / Anthropic 直连 → 写 reasoning={"enabled":True[, "effort"]}
          None（旧前端 / 非推理模型未发）视为默认开（向后兼容）；'auto' 开但不带 effort；
          'low'/'medium'/'high' 带 effort。
      - 其它 OpenAI 兼容供应商 → 只透传合法 effort(low/medium/high)；off/auto/未知值剥掉，
          避免严格供应商（如直连 OpenAI o 系列）对非法 reasoning_effort 报 400。
    """
    # 先清干净，避免 fresh body 残留或重复调用叠加
    body.pop("reasoning", None)
    body.pop("reasoning_effort", None)
    if skip_prompt or reasoning_effort == "off":
        return
    if is_openrouter or is_anthropic_fmt:
        cfg = {"enabled": True}
        if reasoning_effort in ("low", "medium", "high"):
            cfg["effort"] = reasoning_effort
        body["reasoning"] = cfg
    elif reasoning_effort in ("low", "medium", "high"):
        body["reasoning_effort"] = reasoning_effort


def _is_prompt_cache_enabled_value(value) -> bool:
    return value is None or str(value).lower() != "false"


async def _get_prompt_cache_ttl() -> str:
    try:
        return _normalize_prompt_cache_ttl(await get_config("prompt_cache_ttl"))
    except Exception as e:
        print(f"⚠️  Prompt 缓存 TTL 配置读取失败，使用默认 1h: {e}")
        return "1h"


def _apply_aihubmix_cache_headers(headers: dict, api_url: str) -> None:
    if "aihubmix" in (api_url or "").lower():
        headers["anthropic-beta"] = AIHUBMIX_EXTENDED_CACHE_TTL_BETA


async def _apply_openrouter_sticky_routing(body: dict, is_openrouter: bool, model: str, session_id: str):
    """Keep OpenRouter requests sticky without disabling its cache-aware routing."""
    if not is_openrouter:
        return

    if session_id:
        # OpenRouter accepts a top-level session_id up to 256 chars.
        body["session_id"] = str(session_id)[:256]

    model_lower = (model or "").lower()
    if "claude" not in model_lower and "anthropic" not in model_lower:
        return

    provider_order_enabled = await get_config_bool("openrouter_provider_order_enabled", fallback=False)
    cache_enabled_val = await get_config("prompt_cache_enabled")
    prompt_cache_enabled = _is_prompt_cache_enabled_value(cache_enabled_val)

    if provider_order_enabled and not prompt_cache_enabled and "provider" not in body:
        body["provider"] = {"order": ["Anthropic"], "allow_fallbacks": True}
        print("🔀 Provider 偏好：优先 Anthropic 直连（prompt_cache_enabled=false）")


@app.post("/v1/chat/completions")
async def chat_completions(request: Request):
    """核心转发接口"""
    # API_KEY 检查移到供应商路由的 else 分支：只有「既没匹配到供应商、又没有环境变量
    # API_KEY」时才报 500。否则面板里配了供应商、但 env API_KEY 留空的用户会被误拦。
    body = await request.json()
    messages = body.get("messages", [])
    archive_received_at = datetime.now(timezone.utc)
    archive_idempotency_key = str(request.headers.get("idempotency-key") or "").strip()
    archive_material = archive_idempotency_key or json.dumps(
        messages, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )
    if bool(body.get("is_regenerate")) and not archive_idempotency_key:
        # Regenerations intentionally create assistant variants; without a
        # caller-provided idempotency key they are separate attempts.
        archive_material += f"\x1fregenerate:{uuid.uuid4()}"
    archive_request_key = hashlib.sha256(archive_material.encode("utf-8")).hexdigest()
    
    # ---------- 提取用户最新消息 ----------
    user_message = ""
    for msg in reversed(messages):
        if msg.get("role") == "user" and not _is_compressed_summary_message(msg):
            content = msg.get("content", "")
            if isinstance(content, str):
                user_message = content
            elif isinstance(content, list):
                user_message = " ".join(
                    item.get("text", "") for item in content
                    if isinstance(item, dict) and item.get("type") == "text"
                )
            break
    
    # ---------- 构建 system prompt ----------
    # 内部请求（如压缩上下文）可跳过人设注入
    skip_prompt = body.pop('skip_system_prompt', False)
    
    # 读取前端传来的模板变量上下文
    template_ctx = {
        'model_name': body.get('model', ''),
        'user_name': body.pop('user_name', ''),
        'assistant_name': body.pop('assistant_name', ''),
    }
    
    # v5.8：项目 ID（前端传来，用于项目指令/记忆/文件注入）
    project_id = body.pop('project_id', None) or None
    # v6.0：前端对话 ID，用于无缝换窗时避免衔接到当前对话自身
    conversation_id = body.pop('conversation_id', None) or None
    is_regenerate = bool(body.pop('is_regenerate', False))

    # 先确定最终模型，后面的 prompt cache 判断要用它。
    # 如果客户端没传 model，这里会补上默认值，避免误判为“非 Claude”而跳过缓存。
    # 未传 model 时优先用面板保存的默认聊天模型，再回落到环境变量 DEFAULT_MODEL
    model = body.get("model") or await get_config("default_chat_model") or DEFAULT_MODEL
    body["model"] = model
    provider_model_id_raw = body.pop("provider_model_id", None)
    try:
        provider_model_id = int(provider_model_id_raw) if provider_model_id_raw not in (None, "") else None
    except (TypeError, ValueError):
        provider_model_id = None

    # ---------- 供应商路由（提前解析）----------
    # 提前解析供应商，使本次请求链路里所有「是否走 Anthropic / 是否 OpenRouter」的判断
    # 都基于权威来源 api_format，而不是靠 URL 或模型名字符串去猜（历史 bug 的根因）。
    # 解析出两个布尔，本函数后续统一使用：
    #   is_anthropic_fmt —— 该供应商是否走 Anthropic 原生格式（来自 DB 的 api_format）
    #   is_openrouter    —— 上游是否 OpenRouter（仅用于 OpenRouter 专属字段/请求头）
    try:
        provider_info = await resolve_provider_for_model(model, provider_model_id=provider_model_id)
    except Exception:
        provider_info = None

    api_format = "openai"  # 默认 OpenAI 格式
    if provider_info:
        resolved_model = provider_info.get("model_id")
        if resolved_model and resolved_model != model:
            model = resolved_model
            body["model"] = model
        chat_api_key = provider_info["api_key"]
        api_format = provider_info.get("api_format", "openai") or "openai"
        base = provider_info["api_base_url"].rstrip("/")
        if api_format == "anthropic":
            chat_api_url = get_anthropic_url(base)
            print(f"🔀 路由到供应商 [{provider_info['provider_name']}] (Anthropic 格式): {chat_api_url}")
        else:
            # 归一化：先剥掉可能已带的聊天端点后缀，再统一拼 /chat/completions，
            # 避免用户把 base 填成 .../chat/completions 或误填 .../messages 时拼出错误路径。
            for _suffix in ("/chat/completions", "/messages"):
                if base.endswith(_suffix):
                    base = base[: -len(_suffix)]
                    break
            chat_api_url = f"{base}/chat/completions"
            print(f"🔀 路由到供应商 [{provider_info['provider_name']}]: {base}")
    else:
        if not API_KEY:
            return JSONResponse(
                status_code=500,
                content={"error": "API_KEY 未设置，请在环境变量中配置"},
            )
        chat_api_key = API_KEY
        chat_api_url = API_BASE_URL

    is_anthropic_fmt = (api_format == "anthropic")
    is_openrouter = "openrouter" in chat_api_url.lower()

    mem_enabled = await get_memory_enabled()
    prompt_meta = {}
    cache_on = False
    cache_ttl = "1h"
    cache_enabled_val = None
    if not skip_prompt:
        # v5.6：计算用户消息数（用于无缝切窗判断是第几轮）
        user_msg_count = _count_real_user_messages(messages)
        if mem_enabled and user_message:
            enhanced_prompt, prompt_meta = await build_system_prompt_with_memories(user_message, user_msg_count=user_msg_count, project_id=project_id, conversation_id=conversation_id, is_regenerate=is_regenerate)
        else:
            # v5.4：即使记忆关闭，也从数据库优先读取 system prompt（降级到文件版本）
            enhanced_prompt = await get_active_system_prompt() or SYSTEM_PROMPT
            # 记忆关闭时也注入当前时间；时间属于运行时上下文，不进入 system 缓存前缀。
            if "<!-- CACHE_B1 -->" not in enhanced_prompt and "<!-- CACHE_BOUNDARY -->" not in enhanced_prompt:
                enhanced_prompt += "\n\n<!-- CACHE_B1 -->\n\n<!-- CACHE_B2 -->"
            _now_cst = datetime.now(TZ_CST)
            _weekdays_cn = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
            _append_runtime_context_section(
                prompt_meta,
                f"【当前时间】{_now_cst.strftime('%Y-%m-%d %H:%M')} {_weekdays_cn[_now_cst.weekday()]}",
            )
        
        if enhanced_prompt:
            # 替换模板变量
            enhanced_prompt = replace_template_variables(enhanced_prompt, template_ctx)

            _normalize_frontend_system_messages(messages, template_ctx, prompt_meta)

            has_system = any(msg.get("role") == "system" for msg in messages)
            if has_system:
                for i, msg in enumerate(messages):
                    if msg.get("role") == "system":
                        frontend_sys = msg["content"] or ""
                        if not isinstance(frontend_sys, str):
                            # 第三方前端可能发数组格式 content blocks 的 system message。
                            # 不做摘要重排：把后端 enhanced_prompt 作为独立 system 插在它前面，
                            # 前端原数组原样保留，交给下游合并（adapter 会把多条 system 并进顶层 system）。
                            messages.insert(i, {"role": "system", "content": enhanced_prompt})
                            break
                        # 稳定的前端 system（如 skill prompt）插到 B2 前，随静态区缓存。
                        # 旧 system 摘要已在上方规范化成 user 消息，避免污染 B2。
                        frontend_sys = replace_template_variables(frontend_sys, template_ctx)
                        messages[i]["content"] = _merge_frontend_system_prompt(enhanced_prompt, frontend_sys)
                        break
            else:
                messages.insert(0, {"role": "system", "content": enhanced_prompt})
            
            # ---- Prompt Caching：把 system message 拆成多段 content block ----
            # 是否支持显式 cache_control 断点。Anthropic 原生格式直接支持；
            # OpenAI 兼容格式只给已验证会透传 Claude cache_control 的供应商放行。
            _model_lower = (model or "").lower()
            _api_url_lower = (chat_api_url or "").lower()
            is_claude_target = "claude" in _model_lower or "anthropic" in _model_lower
            is_aihubmix = "aihubmix" in _api_url_lower
            supports_explicit_cache = is_anthropic_fmt or (
                is_claude_target and (is_openrouter or is_aihubmix)
            )
            cache_enabled_val = await get_config("prompt_cache_enabled")
            cache_on = supports_explicit_cache and _is_prompt_cache_enabled_value(cache_enabled_val)
            if cache_on:
                cache_ttl = await _get_prompt_cache_ttl()
            cache_markers_present = (
                "<!-- CACHE_B1 -->" in enhanced_prompt
                or "<!-- CACHE_BOUNDARY -->" in enhanced_prompt
            )

            if cache_on:
                cache_applied = False
                for i, msg in enumerate(messages):
                    if msg.get("role") == "system" and isinstance(msg.get("content"), str):
                        content = msg["content"]

                        if "<!-- CACHE_B1 -->" in content:
                            seg1, rest = content.split("<!-- CACHE_B1 -->", 1)

                            if "<!-- CACHE_B2 -->" in rest:
                                seg2, rest2 = rest.split("<!-- CACHE_B2 -->", 1)
                            else:
                                seg2 = rest
                                rest2 = ""

                            dynamic = rest2

                            # 空段不打断点：seg2 在默认配置下可能为空，Anthropic 不接受空 text block。
                            content_blocks = []
                            for seg in (seg1, seg2):
                                text = (seg or "").strip()
                                if text:
                                    content_blocks.append({
                                        "type": "text",
                                        "text": text,
                                        "cache_control": _cache_control_block(cache_ttl),
                                    })

                            if dynamic.strip():
                                content_blocks.append({"type": "text", "text": dynamic.strip()})

                            messages[i]["content"] = content_blocks
                            bp_count = sum(1 for b in content_blocks if "cache_control" in b)
                            sizes = " + ".join(f"~{len(b['text'])}字" for b in content_blocks)
                            print(f"💾 Prompt 缓存：{bp_count} 个断点（{sizes}）")
                            cache_applied = True

                        elif "<!-- CACHE_BOUNDARY -->" in content:
                            # 旧版单断点兼容
                            static_part, dynamic_part = content.split("<!-- CACHE_BOUNDARY -->", 1)
                            static_part = static_part.rstrip()
                            dynamic_part = dynamic_part.strip()
                            
                            content_blocks = [
                                {"type": "text", "text": static_part, "cache_control": _cache_control_block(cache_ttl)}
                            ]
                            if dynamic_part:
                                content_blocks.append({"type": "text", "text": dynamic_part})
                            
                            messages[i]["content"] = content_blocks
                            print(f"💾 Prompt 缓存（旧版）：静态 ~{len(static_part)}字")
                            cache_applied = True
                        break
                if not cache_applied:
                    print(f"💾 Prompt 缓存跳过：未找到可拆分的 system 缓存边界（api_format={api_format}, model={model}）")
            elif cache_markers_present:
                if cache_enabled_val is not None and str(cache_enabled_val).lower() == 'false':
                    cache_skip_reason = "prompt_cache_enabled=false"
                else:
                    cache_skip_reason = (
                        "provider/model not enabled for explicit cache "
                        f"(api_format={api_format}, is_openrouter={is_openrouter}, "
                        f"is_aihubmix={is_aihubmix}, is_claude={is_claude_target}, model={model})"
                    )
                print(f"💾 Prompt 缓存跳过：{cache_skip_reason}")
    
    # 替换前端传来的 skill prompt 中的模板变量
    for msg in messages:
        if msg.get("role") == "system" and '{' in msg.get("content", ""):
            msg["content"] = replace_template_variables(msg["content"], template_ctx)
    
    body["messages"] = messages
    
    # ---------- 联网搜索 ----------
    tool_events = []  # 收集工具事件，通过 SSE 发给前端展示

    web_search_mode = body.pop("web_search", False)
    # 兼容布尔和字符串：true → 强制搜索, "auto" → 模型自行决定, false → 关闭
    if web_search_mode == "auto":
        do_search_force = False
        do_search_auto = True
    else:
        do_search_force = bool(web_search_mode)
        do_search_auto = False

    if do_search_force and user_message:
        try:
            search_engine = await get_config("search_engine") or ""
            search_api_key = await get_config("search_api_key") or ""
            search_max = await get_config_int("search_max_results", fallback=5)
            
            if search_engine:
                print(f"🌐 联网搜索: [{search_engine}] {user_message[:60]}...")
                search_results = await web_search(
                    query=user_message[:200],
                    engine=search_engine,
                    api_key=search_api_key,
                    max_results=search_max,
                )
                if search_results:
                    search_text = format_results_for_prompt(search_results, user_message[:100])
                    _append_runtime_context_section(prompt_meta, search_text)
                    print(f"🌐 搜索完成，获得 {len(search_results)} 条结果")
                    tool_events.append({
                        "type": "search", "engine": search_engine,
                        "query": user_message[:100], "count": len(search_results),
                        "results": [r.to_dict() for r in search_results[:10]],
                    })
                else:
                    print(f"🌐 搜索无结果")
            else:
                print(f"⚠️ 联网搜索已请求但未配置搜索引擎")
        except Exception as e:
            print(f"❌ 联网搜索出错: {e}")
    
    # ---------- 模型处理 ----------
    # model 已在 prompt cache 判断前标准化，避免缺省 model 请求跳过缓存。
    
    # ---------- 供应商路由 ----------
    # 已在前面（确定 model 后）提前解析：provider_info / api_format / chat_api_key /
    # chat_api_url / is_anthropic_fmt / is_openrouter 均已就绪，这里不再重复解析。

    # ---------- MCP 工具调用 ----------
    mcp_mode_raw = body.pop("mcp_mode", None)
    # Explicit request-body MCP servers are a compatibility path and are not governed by mcp_mode.
    mcp_servers = body.pop("mcp_servers", [])
    if mcp_mode_raw is None:
        mcp_mode_raw = await get_config("mcp_mode")
    mcp_mode = str(mcp_mode_raw or "auto").strip().lower()
    if mcp_mode not in ("off", "auto", "manual"):
        mcp_mode = "auto"
    
    # ---------- 生成 session ID ----------
    # 优先用前端传来的 conversation_id：工具抽屉的"手动展开工具"状态按 session 存，
    # 设计上是"下一轮对话生效"。session_id 必须跨轮稳定，否则每轮新 uuid 会丢掉
    # 上一轮展开的类别，_drawer_request_tools 永远不会真正生效。
    # Bug #3：前端不传 conversation_id 时，退回用「首条用户消息」的 hash —— 它在同一段
    # 对话的多轮间稳定（不像 uuid 每轮都变），让上面说的“跨轮稳定”在无 conversation_id
    # 时也成立。
    if conversation_id:
        session_id = conversation_id
    else:
        first_user = next(
            (
                _content_text_for_prompt(m.get("content"))
                for m in messages
                if m.get("role") == "user" and not _is_compressed_summary_message(m)
            ),
            "",
        ) or ""
        session_id = "auto-" + hashlib.md5(first_user.encode("utf-8")).hexdigest()[:8]

    archive_context = None
    if not skip_prompt:
        try:
            archive_context = await _archive_gateway_user(
                session_id=session_id,
                request_key=archive_request_key,
                user_message=user_message,
                model=model,
                project_id=project_id,
                is_regenerate=is_regenerate,
                occurred_at=archive_received_at,
            )
        except Exception as exc:
            print(f"🚨 完整聊天归档失败（user）: {type(exc).__name__}: {exc}")
            return JSONResponse(
                status_code=503,
                content={"error": "chat archive is enabled but the incoming message could not be stored"},
            )
    
    # 请求 LLM 在流式响应中包含 token 用量
    if body.get("stream"):
        body.setdefault("stream_options", {})["include_usage"] = True
    
    # 思考链参数：统一交给 _apply_reasoning 处理（转发路径与工具循环同一套规则，对各类供应商分流）。
    # 由 to_anthropic_request 把 reasoning.enabled 转成 Anthropic extended thinking。
    reasoning_effort = body.pop("reasoning_effort", None)
    _apply_reasoning(body, is_openrouter, is_anthropic_fmt, reasoning_effort, skip_prompt)
    
    # ---------- Prompt 缓存（v5.5 → v5.7 修正）----------
    # cache_control 现在在 system message 的 content block 上加，不在 body 顶层
    # （旧代码在 body 加 cache_control 是无效的，OpenRouter 需要 content block 级标记）
    
    # ---------- OpenRouter 黏性路由 ----------
    # session_id 让 OpenRouter 从第一轮开始按对话黏住路由；缓存开启时不再
    # 自动写 provider.order，避免绕开 OpenRouter 的 cache-aware sticky routing。
    await _apply_openrouter_sticky_routing(body, is_openrouter, model, session_id)

    # ---------- 转发请求 ----------
    if is_anthropic_fmt:
        headers = to_anthropic_headers(chat_api_key)
    else:
        headers = {
            "Authorization": f"Bearer {chat_api_key}",
            "Content-Type": "application/json",
        }
        if is_openrouter:
            headers["HTTP-Referer"] = EXTRA_REFERER
            headers["X-Title"] = EXTRA_TITLE
        _apply_aihubmix_cache_headers(headers, chat_api_url)
    
    is_stream = body.get("stream", False)
    
    # 🔍 调试：记录思考链相关参数
    if body.get("reasoning") or body.get("reasoning_effort") or body.get("include_reasoning"):
        print(f"🔍 [思考链参数] reasoning={body.get('reasoning')}, reasoning_effort={body.get('reasoning_effort')}, include_reasoning={body.get('include_reasoning')}")
    
    # ========== 收集工具（v6.3：抽屉模式 / 传统模式 二选一） ==========
    openai_tools = []
    tool_map = {}

    reminder_tools_enabled = await get_config_bool("reminder_tools_enabled", fallback=True)
    drawer_enabled = await get_config_bool("tool_drawer_enabled", fallback=False)

    if drawer_enabled:
        # ---- 抽屉模式：向量路由按需展开内部工具 + 外部 MCP 双轨 ----
        # Lazy init：toggle 启动时为 false、运行时打开的场景下，lifespan 没跑过
        # init_drawer，此时 CATEGORIES 为空会让 route_tools 返回 0 工具，叠加
        # 下面的 `not drawer_enabled` 门控会让传统工具也消失。这里幂等调用兜底。
        try:
            from tool_drawer import init_drawer as _drawer_init
            await _drawer_init()
        except Exception as e:
            print(f"⚠️ 工具抽屉 lazy init 失败: {e}")
        try:
            from tool_drawer import route_tools as _drawer_route
            user_embedding = prompt_meta.get("user_embedding") if prompt_meta else None
            drawer_tools, drawer_map = await _drawer_route(
                user_message=user_message,
                session_id=session_id,
                user_embedding=user_embedding,
                mem_enabled=mem_enabled,
                search_enabled=bool(do_search_auto),
                project_id=project_id,
                mcp_mode=mcp_mode,
                reminder_tools_enabled=reminder_tools_enabled,
            )
            openai_tools.extend(drawer_tools)
            tool_map.update(drawer_map)
        except Exception as e:
            print(f"❌ 工具抽屉路由失败: {e}")
        # Request-body MCP servers are explicit third-party input and bypass mcp_mode by design.
        if mcp_servers:
            try:
                mcp_tools, mcp_map = await get_tools_for_servers(mcp_servers)
                openai_tools.extend(mcp_tools)
                tool_map.update(mcp_map)
            except Exception as e:
                print(f"❌ 外部 MCP 工具获取失败: {e}")

    # ---- 传统模式：原有逐项注册（drawer_enabled=False 时执行）----
    # MCP 工具
    if not drawer_enabled and mcp_servers:
        try:
            mcp_tools, mcp_map = await get_tools_for_servers(mcp_servers)
            openai_tools.extend(mcp_tools)
            tool_map.update(mcp_map)
        except Exception as e:
            print(f"❌ MCP 工具获取失败: {e}")

    # 联网搜索 auto 模式：注册为 function tool，让模型自行决定是否调用
    if not drawer_enabled and do_search_auto:
        search_engine = await get_config("search_engine") or ""
        if search_engine:
            openai_tools.append({
                "type": "function",
                "function": {
                    "name": "_gateway_web_search",
                    "description": "搜索互联网获取实时信息。仅在用户明确要求联网搜索、或需要最新新闻/天气/实时数据/你不确定的事实时调用。闲聊、角色扮演、创意写作等不需要调用。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {
                                "type": "string",
                                "description": "搜索关键词，用简洁的搜索引擎友好格式",
                            }
                        },
                        "required": ["query"],
                    },
                },
            })
            # 标记为网关内置工具（不走 MCP，本地执行）
            tool_map["_gateway_web_search"] = {"type": "gateway_builtin", "handler": "web_search"}
            print(f"🌐 联网搜索已注册为工具（auto 模式，引擎: {search_engine}）")
        else:
            print(f"⚠️ 联网搜索 auto 模式已请求但未配置搜索引擎")

    # v5.8：对话搜索工具（始终可用，让模型能主动搜索过去的对话）
    if not drawer_enabled and mem_enabled:
        openai_tools.append({
            "type": "function",
            "function": {
                "name": "_gateway_search_conversations",
                "description": "搜索过去的对话记录。当用户提到'我们之前聊过''上次说的''之前讨论的'等回忆性表达，或者你需要查找过去对话中的具体细节时调用。输入搜索关键词，返回匹配的对话片段和上下文。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "搜索关键词，用简洁的内容关键词（如'用药方案''生日''项目部署'），不要用'我们讨论过'之类的元描述",
                        },
                        "limit": {
                            "type": "integer",
                            "description": "最多返回几条匹配（默认10）",
                        },
                    },
                    "required": ["query"],
                },
            },
        })
        tool_map["_gateway_search_conversations"] = {"type": "gateway_builtin", "handler": "search_conversations", "project_id": project_id}
        print(f"🔍 对话搜索工具已注册")

    # 提醒系统工具：配置级常驻，独立于记忆开关。
    if not drawer_enabled and reminder_tools_enabled:
        _reminder_tools = [
        {
            "type": "function",
            "function": {
                "name": "_gateway_create_reminder",
                "description": "为用户创建一条提醒。当用户说'提醒我...'、'...之后叫我...'、'每天...点提醒我...'时调用。title 用简洁的中文描述，notes 用来记录上下文信息以便提醒时参考。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {"type": "string", "description": "提醒标题，简短描述（如'吃药''给妈妈打电话'）"},
                        "notes": {"type": "string", "description": "备注信息，提醒时作为上下文参考（如'妈妈上周说周末要搬东西'）"},
                        "trigger_time": {"type": "string", "description": "触发时间，ISO 8601 格式（如'2026-03-31T23:00:00+08:00'）。相对时间请转换为绝对时间。"},
                        "repeat_type": {"type": "string", "enum": ["once", "daily", "weekly", "hourly"], "description": "重复类型：once=一次性, daily=每天, weekly=每周, hourly=每N小时"},
                        "repeat_config": {"type": "object", "description": "循环配置（hourly时传{hours:N}）"},
                    },
                    "required": ["title", "trigger_time"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_list_reminders",
                "description": "查看用户当前的所有活跃提醒。当用户问'我设了哪些提醒'、'有什么提醒'时调用。",
                "parameters": {"type": "object", "properties": {}},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_complete_reminder",
                "description": "标记一条提醒为已完成。当用户表示事情已经做完（如'回来了''做完了''学完了'），且当前有相关的待触发提醒时调用。一次性提醒会被标记完成，循环提醒不受影响。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "reminder_id": {"type": "string", "description": "要完成的提醒 ID"},
                    },
                    "required": ["reminder_id"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_delete_reminder",
                "description": "删除一条提醒（包括循环提醒）。当用户说'取消那个提醒'、'以后不用提醒我...了'时调用。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "reminder_id": {"type": "string", "description": "要删除的提醒 ID"},
                    },
                    "required": ["reminder_id"],
                },
            },
        },
    ]
        openai_tools.extend(_reminder_tools)
        for t in _reminder_tools:
            tool_map[t["function"]["name"]] = {"type": "gateway_builtin", "handler": "reminder"}
        print("⏰ 提醒工具已注册（常驻）")

    tools_cache_applied = False
    if cache_on and openai_tools and is_stream:
        tools_cache_applied = _apply_tools_cache_breakpoint(openai_tools, tool_map, cache_ttl=cache_ttl)
        if tools_cache_applied:
            print("💾 Prompt 缓存：tools meta-tool 断点已设置")

    runtime_sections = prompt_meta.get("runtime_context_sections", []) if prompt_meta else []
    if runtime_sections:
        if _inject_runtime_context_into_last_user(messages, runtime_sections):
            print(f"🧩 运行时上下文已注入最后一条 user 消息（{len(runtime_sections)} 段）")
        else:
            print("⚠️  运行时上下文注入失败：未找到可注入的 user 消息")

    message_cache_idx = None
    if cache_on:
        message_cache_idx, message_cache_reason = _apply_messages_cache_breakpoint(messages, is_regenerate=is_regenerate, cache_ttl=cache_ttl)
        if message_cache_idx is not None:
            print(f"💾 Prompt 缓存：messages 断点已设置在 index={message_cache_idx}")
        else:
            print(f"💾 Prompt 缓存：messages 断点跳过（{message_cache_reason}）")

        total_breakpoints = (
            (_count_cache_control_entries_in_tools(openai_tools) if is_stream else 0)
            + _count_cache_control_entries_in_messages(messages)
        )
        if total_breakpoints > 4:
            if message_cache_idx is not None:
                _clear_message_cache_control(messages[message_cache_idx])
                total_breakpoints = (
                    (_count_cache_control_entries_in_tools(openai_tools) if is_stream else 0)
                    + _count_cache_control_entries_in_messages(messages)
                )
                print(f"⚠️  Prompt 缓存：断点数超过 4，已撤销 messages 断点（当前 {total_breakpoints}）")
            else:
                print(f"⚠️  Prompt 缓存：断点数超过 4（当前 {total_breakpoints}），请检查配置")
        else:
            print(f"💾 Prompt 缓存：总断点数 {total_breakpoints}/4")
    body["messages"] = messages

    # ========== Tool Call 模式（MCP 和/或 auto 搜索） ==========
    if openai_tools and is_stream:
        print(f"🔧 工具模式: 共 {len(openai_tools)} 个工具可用")

        return StreamingResponse(
            _stream_with_tools(
                messages=messages,
                tools=openai_tools,
                tool_map=tool_map,
                model=model,
                temperature=body.get("temperature", 0.7),
                top_p=body.get("top_p"),
                max_tokens=body.get("max_tokens"),
                tool_events=tool_events,
                session_id=session_id,
                user_message=user_message,
                mem_enabled=mem_enabled,
                api_url=chat_api_url,
                api_key=chat_api_key,
                project_id=project_id,
                prompt_meta=prompt_meta,
                api_format=api_format,
                is_regenerate=is_regenerate,
                reasoning_effort=reasoning_effort,
                skip_prompt=skip_prompt,
                archive_context=archive_context,
            ),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )

    # ========== 正常转发模式 ==========
    if is_stream:
        return StreamingResponse(
            stream_and_capture(headers, body, session_id, user_message, model, tool_events, api_url=chat_api_url, project_id=project_id, prompt_meta=prompt_meta, api_format=api_format, api_key=chat_api_key, is_regenerate=is_regenerate, mem_enabled=mem_enabled, archive_context=archive_context),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )
    else:
        # 非流式：Anthropic 格式需要转换请求和响应
        send_body = to_anthropic_request(body) if api_format == "anthropic" else body
        async with httpx.AsyncClient(timeout=300) as client:
            response = await client.post(chat_api_url, headers=headers, json=send_body)

            if response.status_code == 200:
                resp_data = response.json()
                # Anthropic 响应转回 OpenAI 格式
                if api_format == "anthropic":
                    resp_data = from_anthropic_response(resp_data, model)
                assistant_msg = ""
                try:
                    assistant_msg = resp_data["choices"][0]["message"]["content"]
                except (KeyError, IndexError):
                    pass
                dream_triggered = detect_dream_trigger(assistant_msg)
                
                if mem_enabled and user_message and assistant_msg:
                    _emo = merge_emotion_levels(detect_emotion_from_user_msg(user_message), detect_emotion_from_response(assistant_msg))
                    _spawn_background_task(
                        process_memories_background(session_id, user_message, assistant_msg, model, emotion_level=_emo, project_id=project_id, is_regenerate=is_regenerate)
                    )
                if archive_context and assistant_msg:
                    try:
                        await _archive_gateway_assistant(archive_context, assistant_msg)
                    except Exception as exc:
                        print(f"🚨 完整聊天归档失败（assistant）: {type(exc).__name__}: {exc}")
                        resp_data["kiwi_archive_warning"] = "assistant response was not archived"
                
                if dream_triggered:
                    print(f"🌙 检测到 Dream 标记，后台启动 Dream（非流式响应无 SSE 事件）...")
                    _launch_dream_from_marker()
                return JSONResponse(status_code=200, content=resp_data)
            else:
                try:
                    err_content = response.json()
                except Exception:
                    err_content = {"error": response.text[:500]}
                return JSONResponse(status_code=response.status_code, content=err_content)


async def _execute_gateway_tool(tool_name: str, arguments: dict, tool_info: dict) -> tuple:
    """
    执行网关内置工具（联网搜索、提醒系统等）。
    返回 (result_text, extra_metadata) 元组，extra_metadata 用于 SSE 事件附加信息。
    """
    extra = {}

    if tool_name == "_gateway_web_search":
        query = arguments.get("query", "")
        if not query:
            return "搜索关键词为空", extra
        try:
            search_engine = await get_config("search_engine") or ""
            search_api_key = await get_config("search_api_key") or ""
            search_max = await get_config_int("search_max_results", fallback=5)

            # 未配置搜索引擎时明确告知，不能让它落到 web_search 返回空、伪装成「搜了但无结果」
            # ——否则模型会以为搜过了没找到，进而编造答案，掩盖了「根本没配引擎」的真因。
            if not search_engine:
                print(f"⚠️ [auto] 模型请求联网搜索，但未配置搜索引擎")
                return "联网搜索未配置搜索引擎，无法搜索。请在管理面板配置后再试。", extra

            print(f"🌐 [auto] 模型请求联网搜索: [{search_engine}] {query[:80]}")
            results = await web_search(
                query=query[:200],
                engine=search_engine,
                api_key=search_api_key,
                max_results=search_max,
            )
            if results:
                extra = {
                    "engine": search_engine,
                    "query": query[:100],
                    "count": len(results),
                    "results": [r.to_dict() for r in results[:10]],
                }
                text = format_results_for_prompt(results, query[:100])
                print(f"🌐 [auto] 搜索完成，{len(results)} 条结果")
                return text, extra
            else:
                print(f"🌐 [auto] 搜索无结果")
                return f"搜索「{query}」无结果。", extra
        except Exception as e:
            print(f"❌ [auto] 搜索出错: {e}")
            return f"搜索失败: {e}", extra

    # ── v5.8：对话搜索工具 ──

    if tool_name in ("_gateway_search_conversations", "gateway_search_conversations"):
        query = arguments.get("query", "")
        if not query:
            return "搜索关键词为空", extra
        try:
            from database import search_chat_messages
            search_limit = arguments.get("limit", 10)
            search_project_id = tool_info.get("project_id")
            results = await search_chat_messages(query, project_id=search_project_id, limit=search_limit, context_size=2)
            
            title_matches = results.get("title_matches", [])
            msg_matches = results.get("message_matches", [])
            
            if not title_matches and not msg_matches:
                print(f"🔍 对话搜索 '{query}' → 无结果")
                return f"在过去的对话中没有找到与「{query}」相关的内容。", extra
            
            # 格式化结果给模型
            lines = []
            
            if title_matches:
                lines.append(f"## 标题匹配（{len(title_matches)} 个对话）")
                for t in title_matches:
                    dt = t.get("date", "")[:10] if t.get("date") else ""
                    lines.append(f"- [{t['title']}]（{dt}）")
            
            if msg_matches:
                lines.append(f"\n## 消息内容匹配（{len(msg_matches)} 个对话）")
                for conv in msg_matches:
                    dt = conv.get("date", "")[:10] if conv.get("date") else ""
                    lines.append(f"\n### {conv['title']}（{dt}）")
                    for match in conv["matches"][:3]:
                        for msg in match.get("context", []):
                            role = "用户" if msg["role"] == "user" else "助手"
                            marker = "→ " if msg.get("is_match") else "  "
                            content = msg["content"][:200]
                            lines.append(f"{marker}{role}: {content}")
                        lines.append("")
            
            text = "\n".join(lines)
            total = len(title_matches) + sum(len(c.get("matches", [])) for c in msg_matches)
            print(f"🔍 对话搜索 '{query}' → 标题{len(title_matches)}条 + 消息{len(msg_matches)}个对话")
            
            extra = {
                "query": query[:100],
                "title_count": len(title_matches),
                "message_count": len(msg_matches),
            }
            return text, extra
        except Exception as e:
            print(f"❌ 对话搜索出错: {e}")
            return f"搜索失败: {e}", extra

    # ── 提醒系统工具 ──

    if tool_name == "_gateway_create_reminder":
        try:
            title = arguments.get("title", "")
            if not title:
                return "提醒标题不能为空", extra
            reminder_data = {
                "title": title,
                "notes": arguments.get("notes", ""),
                "trigger_time": arguments.get("trigger_time", ""),
                "repeat_type": arguments.get("repeat_type", "once"),
                "repeat_config": arguments.get("repeat_config", {}),
            }
            result = await create_reminder(reminder_data)
            repeat_label = {"once": "一次性", "daily": "每天", "weekly": "每周", "hourly": "循环"}.get(reminder_data["repeat_type"], "一次性")
            print(f"⏰ 提醒已创建: [{repeat_label}] {title}")
            return json.dumps({
                "success": True,
                "message": f"提醒已创建：{title}（{repeat_label}）",
                "reminder": result,
            }, ensure_ascii=False), extra
        except Exception as e:
            print(f"❌ 创建提醒失败: {e}")
            return f"创建提醒失败: {e}", extra

    if tool_name == "_gateway_list_reminders":
        try:
            reminders = await get_reminders(include_completed=False)
            if not reminders:
                return json.dumps({"success": True, "message": "当前没有活跃的提醒", "reminders": []}, ensure_ascii=False), extra
            lines = []
            for r in reminders:
                repeat_label = {"once": "一次性", "daily": "每天", "weekly": "每周", "hourly": "循环"}.get(r.get("repeat_type", "once"), "")
                status = "✅" if r.get("enabled") else "⏸️"
                lines.append(f"{status} [{r['id']}] {r['title']}（{repeat_label}，{r['trigger_time']}）")
                if r.get("notes"):
                    lines.append(f"   备注: {r['notes']}")
            return json.dumps({
                "success": True,
                "message": f"共 {len(reminders)} 条活跃提醒",
                "details": "\n".join(lines),
                "reminders": reminders,
            }, ensure_ascii=False), extra
        except Exception as e:
            return f"查询提醒失败: {e}", extra

    if tool_name == "_gateway_complete_reminder":
        try:
            rid = arguments.get("reminder_id", "")
            if not rid:
                return "请提供提醒 ID", extra
            # 查找提醒信息，循环提醒用 fire_reminder 计算下次触发时间
            all_rems = await get_reminders(include_completed=False)
            rem = next((r for r in all_rems if r["id"] == rid), None)
            if not rem:
                return json.dumps({"success": False, "message": "提醒不存在或已完成"}, ensure_ascii=False), extra
            ok = await fire_reminder(rid, rem.get("repeat_type", "once"), rem.get("repeat_config"))
            if ok:
                action = "已标记为完成" if rem.get("repeat_type") == "once" else "已触发，下次将自动提醒"
                print(f"⏰ 提醒已处理: {rid} ({action})")
                return json.dumps({"success": True, "message": f"提醒「{rem.get('title', rid)}」{action}"}, ensure_ascii=False), extra
            return json.dumps({"success": False, "message": "操作失败"}, ensure_ascii=False), extra
        except Exception as e:
            return f"完成提醒失败: {e}", extra

    if tool_name == "_gateway_delete_reminder":
        try:
            rid = arguments.get("reminder_id", "")
            if not rid:
                return "请提供提醒 ID", extra
            ok = await delete_reminder(rid)
            if ok:
                print(f"⏰ 提醒已删除: {rid}")
                return json.dumps({"success": True, "message": f"提醒已删除"}, ensure_ascii=False), extra
            return json.dumps({"success": False, "message": "提醒不存在"}, ensure_ascii=False), extra
        except Exception as e:
            return f"删除提醒失败: {e}", extra

    return f"未知的内置工具: {tool_name}", extra


async def _stream_with_tools(messages, tools, tool_map, model, temperature, tool_events, session_id, user_message, mem_enabled, api_url=None, api_key=None, project_id=None, prompt_meta=None, api_format="openai", is_regenerate: bool = False, reasoning_effort: str = None, skip_prompt: bool = False, top_p=None, max_tokens=None, archive_context=None):
    """
    工具 + 流式模式：tool call 轮次用非流式（需要完整看 tool_calls），
    最终回复直接输出已获得的内容（模拟流式），不再重复请求 LLM。
    工具执行采用并发策略：同服务器复用连接，跨服务器并行。
    """
    import httpx as _httpx

    _request_meta_context = tool_map.get("_drawer_request_tools", {})
    if _request_meta_context.get("type") == "meta":
        _request_disabled_categories = set(_request_meta_context.get("disabled_categories") or ())
    else:
        _request_disabled_categories = set()

    # Legacy compatibility only: keep old model-emitted return calls inside the
    # drawer meta path, even if an external MCP tried to claim the same name.
    tool_map["_drawer_return_tools"] = {"type": "meta", "handler": "drawer_meta", "legacy": True}

    _api_url = api_url or API_BASE_URL
    _api_key = api_key or API_KEY

    # 与主请求链路同口径：用 api_format（权威来源）+ URL 判断收敛成两个布尔
    _is_anthropic_fmt = (api_format == "anthropic")
    _is_openrouter = "openrouter" in _api_url.lower()

    # 先发送衔接提示（如果有无缝切窗）
    if prompt_meta and prompt_meta.get("handoff"):
        yield f"data: {json.dumps({'ev_handoff': prompt_meta['handoff']}, ensure_ascii=False)}\n\n"

    # 先发送已有的 tool_events（比如强制搜索结果）
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"

    if _is_anthropic_fmt:
        headers = to_anthropic_headers(_api_key)
    else:
        headers = {
            "Authorization": f"Bearer {_api_key}",
            "Content-Type": "application/json",
        }
        if _is_openrouter:
            headers["HTTP-Referer"] = EXTRA_REFERER
            headers["X-Title"] = EXTRA_TITLE
        _apply_aihubmix_cache_headers(headers, _api_url)

    current_messages = list(messages)
    max_rounds = 30

    for round_num in range(max_rounds):
        # ── tool call 轮：非流式请求，检测是否有工具调用 ──
        body = {
            "model": model,
            "messages": current_messages,
            "tools": tools,
            "tool_choice": "auto",
            "temperature": temperature,
            "stream": False,
        }
        # 条目九：透传采样参数（None 时不带该字段，交给上游默认）；top_p=0.0 合法，故用 is not None
        if top_p is not None:
            body["top_p"] = top_p
        if max_tokens is not None:
            body["max_tokens"] = max_tokens
        # 与转发路径同一套规则：尊重用户思考强度，并对各类供应商分流
        # （OpenRouter/Anthropic 用 reasoning 字段；其它 OpenAI 兼容供应商透传合法 effort）。
        _apply_reasoning(body, _is_openrouter, _is_anthropic_fmt, reasoning_effort, skip_prompt)
        await _apply_openrouter_sticky_routing(body, _is_openrouter, model, session_id)

        # Anthropic 格式转换
        send_body = to_anthropic_request(body) if api_format == "anthropic" else body

        print(f"🔄 Tool loop round {round_num + 1}: {len(tools)} tools, {len(current_messages)} msgs (format={api_format})")

        async with _httpx.AsyncClient(timeout=120) as client:
            resp = await client.post(_api_url, headers=headers, json=send_body)
            if resp.status_code != 200:
                print(f"❌ LLM 请求失败: {resp.status_code} {resp.text[:1000]}")

            if resp.status_code != 200 and api_format == "anthropic" and isinstance(send_body, dict) and send_body.get("thinking"):
                retry_body = dict(send_body)
                retry_body.pop("thinking", None)
                retry_body["temperature"] = temperature
                print("↩️ Anthropic 工具轮降级重试：移除 thinking")
                resp = await client.post(_api_url, headers=headers, json=retry_body)
                send_body = retry_body
                if resp.status_code != 200:
                    print(f"❌ LLM 降级重试失败（no thinking）: {resp.status_code} {resp.text[:1000]}")

            if resp.status_code != 200 and api_format == "anthropic" and isinstance(send_body, dict) and send_body.get("tools"):
                retry_body = dict(send_body)
                retry_body.pop("tools", None)
                retry_body.pop("tool_choice", None)
                print("↩️ Anthropic 工具轮降级重试：移除 tools/tool_choice")
                resp = await client.post(_api_url, headers=headers, json=retry_body)
                if resp.status_code != 200:
                    print(f"❌ LLM 降级重试失败（no tools）: {resp.status_code} {resp.text[:1000]}")
                else:
                    send_body = retry_body

            if resp.status_code != 200:
                yield f"data: {json.dumps({'choices': [{'delta': {'content': f'⚠️ 模型请求失败 ({resp.status_code})'}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"
                return
            data = resp.json()
            # Anthropic 响应转回 OpenAI 格式
            if api_format == "anthropic":
                data = from_anthropic_response(data, model)

        choice = data.get("choices", [{}])[0]
        message = choice.get("message", {})
        tool_calls = message.get("tool_calls", [])

        print(f"🔄 Round {round_num + 1}: tool_calls={len(tool_calls)}, has_content={bool(message.get('content'))}")

        if not tool_calls:
            # ── 无 tool_calls：直接用非流式结果，模拟流式输出 ──
            # v5.4 优化：不再重发流式请求，省掉一次完整的模型调用延迟
            final_text = message.get("content", "")
            usage_data = data.get("usage")

            if round_num == 0:
                print(f"⚡ 第一轮无工具调用，直接复用结果输出（省去二次请求）")
            else:
                print(f"✅ 工具调用后最终回复：直接输出 {len(final_text)} 字符")

            assistant_msg = final_text
            dream_triggered = detect_dream_trigger(assistant_msg)

            # ---- 收尾补救（数据必活，与 stream_and_capture 同一套）----
            # 模拟流式（yield + sleep）与思考链 yield 都是取消点；spawn 若写在流式之后，断连时永不执行
            # → 丢消息/丢提取。把 spawn 收进 try/finally 的 finally（不 await；create_task 安全）。
            # final_text 在模拟流式前已完整，断连也存完整消息。
            _tool_finalize_spawned = False
            _tool_dream_spawned = False
            mem_task = None

            def _tool_spawn_finalize():
                nonlocal _tool_finalize_spawned, mem_task
                if _tool_finalize_spawned:
                    return
                _tool_finalize_spawned = True
                if not assistant_msg:
                    return
                _emo = merge_emotion_levels(detect_emotion_from_user_msg(user_message), detect_emotion_from_response(assistant_msg))
                mem_task = _spawn_background_task(
                    _finalize_stream_memories(mem_enabled, session_id, user_message, assistant_msg, model, _emo, project_id, is_regenerate, archive_context)
                )

            def _tool_spawn_dream():
                nonlocal _tool_dream_spawned
                if _tool_dream_spawned:
                    return
                _tool_dream_spawned = True
                if dream_triggered:
                    _spawn_background_task(_dream_fallback_after_grace("auto"))

            try:
                # 处理思考链
                reasoning = message.get("reasoning_content") or message.get("reasoning") or ""
                if reasoning and isinstance(reasoning, str):
                    for i in range(0, len(reasoning), 40):
                        yield f"data: {json.dumps({'choices': [{'delta': {'reasoning_content': reasoning[i:i+40]}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"

                # 模拟流式输出正文
                if final_text:
                    chunk_size = 20
                    for i in range(0, len(final_text), chunk_size):
                        yield f"data: {json.dumps({'choices': [{'delta': {'content': final_text[i:i+chunk_size]}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
                        await asyncio.sleep(0.008)

                finish_payload = {'choices': [{'delta': {}, 'finish_reason': 'stop'}], 'model': model}
                if usage_data:
                    finish_payload['usage'] = usage_data
                yield f"data: {json.dumps(finish_payload, ensure_ascii=False)}\n\n"
            finally:
                _tool_spawn_finalize()
                _tool_spawn_dream()

            # 下面仅正常完成执行（断连时异常已带走控制权）：推 ev_memory / ev_dream / [DONE]
            if mem_task is not None:
                try:
                    mem_result = await asyncio.shield(mem_task)
                except asyncio.CancelledError:
                    raise
                except Exception as e:
                    print(f"⚠️ 收尾记忆 task 出错（不影响流收尾）: {e}")
                    mem_result = None
                if mem_result and mem_result.get("action") != "skip":
                    yield f"data: {json.dumps({'ev_memory': mem_result}, ensure_ascii=False)}\n\n"
            # Dream 事件必须在 [DONE] 之前
            if dream_triggered:
                print(f"🌙 检测到 Dream 标记，通知前端启动 Dream（后端已排兜底）...")
                yield f"data: {json.dumps({'ev_dream': {'triggered': True}}, ensure_ascii=False)}\n\n"
            # [DONE] 作为流的最后一个事件
            yield "data: [DONE]\n\n"
            return

        # ── 有 tool_calls → 并行执行工具 ──
        current_messages.append({
            "role": "assistant",
            "content": message.get("content") or None,
            "tool_calls": tool_calls,
        })

        # 解析所有工具调用
        parsed = []
        for tc in tool_calls:
            tc_id = tc.get("id", "")
            func = tc.get("function", {})
            tc_name = func.get("name", "")
            tc_args_str = func.get("arguments", "{}")
            try:
                tc_args = json.loads(tc_args_str)
            except json.JSONDecodeError:
                tc_args = {}
            parsed.append({"id": tc_id, "name": tc_name, "args": tc_args})

        # 分组：网关内置工具 vs MCP 远程工具
        # v5.8：兼容模型可能吃掉工具名前缀 _ 的情况
        def _resolve_tool_name(name):
            if name in tool_map:
                return name
            if f"_{name}" in tool_map:
                return f"_{name}"
            return name
        
        for p in parsed:
            p["name"] = _resolve_tool_name(p["name"])
        
        # v6.3：四类工具分发 — gateway / drawer / meta / mcp
        # 抽屉模式下 tool_map 会出现 drawer 和 meta 类型，传统模式只有 gateway_builtin/MCP
        gw_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") == "gateway_builtin"]
        drawer_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") == "drawer"]
        meta_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") == "meta"]
        external_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") == "external_mcp"]
        mcp_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") not in ("gateway_builtin", "drawer", "meta", "external_mcp")]

        tool_results = {}   # { call_id: result_text }
        tool_extras = {}    # { call_id: extra_metadata } — 并发安全，每个 call 独立
        approved_categories = set()  # request-local; never infer from shared session state

        # 网关工具（联网搜索等）：各自并发
        async def _run_gw(p):
            tool_info = tool_map.get(p["name"], {})
            result_text, extra_meta = await _execute_gateway_tool(p["name"], p["args"], tool_info)
            tool_results[p["id"]] = result_text
            tool_extras[p["id"]] = extra_meta

        # 抽屉工具：调 mcp_server 函数直跑（绕开 MCP 协议），跑完通知抽屉
        async def _run_drawer(p):
            try:
                from tool_drawer import execute_drawer_tool, record_tool_use
                result_text, extra_meta = await execute_drawer_tool(p["name"], p["args"])
                tool_results[p["id"]] = result_text
                tool_extras[p["id"]] = extra_meta or {}
                record_tool_use(session_id, p["name"])
            except Exception as e:
                tool_results[p["id"]] = f"[drawer_error] {p['name']}: {e}"
                tool_extras[p["id"]] = {}

        # Meta 工具（_drawer_request_tools / list_tool_categories）：单独走 handle_meta_tool
        async def _run_meta(p):
            try:
                from tool_drawer import handle_meta_tool
                tool_info = tool_map.get(p["name"], {})
                result_text = await handle_meta_tool(
                    p["name"],
                    p["args"],
                    session_id,
                    disabled_categories=tool_info.get("disabled_categories"),
                    mcp_mode=tool_info.get("mcp_mode"),
                    pinned_external=tool_info.get("pinned_external"),
                    approved_categories=approved_categories,
                )
                tool_results[p["id"]] = result_text
                tool_extras[p["id"]] = {}
            except Exception as e:
                tool_results[p["id"]] = f"[meta_error] {p['name']}: {e}"
                tool_extras[p["id"]] = {}


        # 外部 MCP 工具：抽屉负责暴露，执行时还原原始工具名
        async def _run_external_mcp(p):
            try:
                from tool_drawer import record_tool_use
                info = tool_map.get(p["name"], {})
                origin_name = info.get("origin_name") or p["name"]
                single_tool_map = {
                    origin_name: {
                        "url": info.get("server_url") or info.get("url", ""),
                        "transport": info.get("transport", "streamable_http"),
                        "server_name": info.get("server_name", ""),
                    }
                }
                result_text = await call_tool(origin_name, p["args"], single_tool_map)
                tool_results[p["id"]] = result_text
                tool_extras[p["id"]] = {}
                record_tool_use(session_id, p["name"])
            except Exception as e:
                tool_results[p["id"]] = f"工具调用失败 [{p['name']}]: {e}"
                tool_extras[p["id"]] = {}

        # 构建并发任务列表
        tasks = [_run_gw(p) for p in gw_parsed]
        tasks.extend(_run_drawer(p) for p in drawer_parsed)
        tasks.extend(_run_meta(p) for p in meta_parsed)
        tasks.extend(_run_external_mcp(p) for p in external_parsed)

        # MCP 工具：同服务器复用连接，不同服务器并发
        if mcp_parsed:
            async def _run_mcp():
                r = await call_tools_batch(mcp_parsed, tool_map)
                tool_results.update(r)
            tasks.append(_run_mcp())

        # 所有工具并发执行
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

        # 若本轮通过 _drawer_request_tools 新展开了抽屉类别，把这些类别的工具增量补进当轮
        # tools/tool_map，让模型「展开 → 同一次回复内下一步就能调用」，不必等下一轮对话。
        if meta_parsed:
            try:
                from tool_drawer import build_tools_for_category
                _newly = []
                for _meta_call in meta_parsed:
                    if _meta_call["name"] != "_drawer_request_tools":
                        continue
                    _cat = _meta_call["args"].get("category")
                    if (
                        _cat in approved_categories
                        and _cat not in _request_disabled_categories
                        and _cat not in _newly
                    ):
                        _newly.append(_cat)
                if _newly:
                    _existing = {t["function"]["name"] for t in tools if t.get("function")}
                    for _cat in _newly:
                        _schemas, _tmap = build_tools_for_category(_cat, project_id=project_id)
                        for _s in _schemas:
                            _nm = _s.get("function", {}).get("name")
                            if _nm and _nm not in _existing:
                                tools.append(_s)
                                _existing.add(_nm)
                        tool_map.update(_tmap)
                    print(f"🗃️  循环内补入展开类别 {_newly}，现共 {len(tools)} 个工具")
            except Exception as _merge_e:
                print(f"⚠️  循环内补入展开工具失败: {_merge_e}")

        print(f"⚡ {len(parsed)} 个工具调用并发完成")

        # 发送工具事件给前端 + 加入消息历史
        for p in parsed:
            result_text = tool_results.get(p["id"], "执行失败")

            evt_type = "search" if p["name"] in ("_gateway_web_search", "gateway_web_search") else "tool_call"
            result_limit = 8000 if tool_map.get(p["name"], {}).get("type") == "external_mcp" else 2000
            evt = {
                "type": evt_type, "name": p["name"],
                "arguments": p["args"],
                "result": result_text[:result_limit] if result_text else "",
            }
            sr = tool_extras.get(p["id"], {})
            if sr:
                evt.update(sr)
            yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"

            current_messages.append({
                "role": "tool",
                "tool_call_id": p["id"],
                "content": result_text[:8000] if result_text else "",
            })

    # 循环结束还没出结果
    yield f"data: {json.dumps({'choices': [{'delta': {'content': '⚠️ 工具调用轮次过多，已停止'}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
    yield "data: [DONE]\n\n"


async def _simulate_stream(text: str, model: str, tool_events: list = None):
    """将完整文本模拟为 SSE 流式输出（tool call 完成后使用）"""
    # 先发送工具事件
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"
    
    chunk_size = 20
    for i in range(0, len(text), chunk_size):
        chunk = text[i:i + chunk_size]
        data = {
            "choices": [{
                "delta": {"content": chunk},
                "finish_reason": None,
            }],
            "model": model,
        }
        yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
        await asyncio.sleep(0.02)  # 模拟流式延迟
    
    # 发送结束标记
    yield "data: [DONE]\n\n"


def _estimate_tokens(text: str) -> int:
    """估算文本的 token 数（中文 ~1.5 char/token，英文 ~4 char/token）"""
    if not text:
        return 0
    cjk = sum(1 for ch in text if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef')
    other = len(text) - cjk
    return max(1, round(cjk / 1.5 + other / 4))


async def _dream_is_running() -> bool:
    """当前是否有 Dream 在跑（兜底启动前的幂等判断）。查不到状态时保守返回 False。"""
    try:
        from database import get_dream_status
        status = await get_dream_status()
        return bool(status.get("current"))
    except Exception:
        return False


def _launch_dream_detached(trigger_type: str = "manual"):
    """detached 启动一次 Dream：run_dream 跑在独立后台 task 里（消费事件仅记日志），
    不依赖任何客户端连接，客户端断连也能把梦做完。
    run_dream 自带 _dream_running 同步并发保护，重复启动会被其拒绝（第二个 yield error 后返回）。"""
    async def _bg_dream():
        try:
            from dream import run_dream
            async for event in run_dream(trigger_type=trigger_type):
                if event["type"] == "error":
                    print(f"   🌙 Dream 出错: {event['data']}")
                elif event["type"] == "complete":
                    print(f"   🌙 Dream 完成: {event['data']}")
        except Exception as e:
            print(f"   🌙 Dream 异常: {e}")
    _spawn_background_task(_bg_dream())


def _launch_dream_from_marker():
    """从聊天回复里的 Dream 标记启动一次后台 Dream（非流式路径用）。"""
    _launch_dream_detached("manual")


# 聊天里检测到 dream 标记后的 Dream 兜底策略：给在线前端一个宽限期，让它照旧通过
# /dream/start 启动并展示"陪着看梦"的进度（既有体验不动）；宽限期后若仍无 Dream 在跑
# （多半客户端已断连、前端没机会启动），后端兜底 detached 启动，保证"去睡吧然后关 App"
# 也能把梦做完。run_dream 单次至少一次 LLM 调用、耗时远超宽限期，故不会误判成"没在跑"而双启。
DREAM_HANDOFF_GRACE_SECONDS = 15


async def _dream_fallback_after_grace(trigger_type: str = "auto", grace: float = DREAM_HANDOFF_GRACE_SECONDS):
    """收尾 task 内的 Dream 兜底：独立后台协程，不随聊天连接取消。"""
    try:
        await asyncio.sleep(grace)
    except asyncio.CancelledError:
        raise
    if await _dream_is_running():
        return  # 在线前端已接手启动，或上次 Dream 仍在跑：不重复启动
    print("🌙 Dream 兜底：宽限期内前端未接手（多半已断连），后端 detached 启动")
    _launch_dream_detached(trigger_type)


def _gateway_archive_turn_id(session_id: str, request_key: str) -> str:
    seed = f"{session_id}\x1f{request_key}".encode("utf-8")
    return hashlib.sha256(seed).hexdigest()[:40]


async def _archive_gateway_user(
    *,
    session_id: str,
    request_key: str,
    user_message: str,
    model: str,
    project_id: str = None,
    is_regenerate: bool = False,
    occurred_at: datetime = None,
) -> dict | None:
    """Persist the incoming visible user utterance before contacting the model."""
    if not user_message or not await get_chat_archive_enabled():
        return None
    turn_id = _gateway_archive_turn_id(session_id, request_key)
    # A regeneration belongs to the same transcript but does not repeat the
    # already visible user utterance.  Its assistant variant is still appended.
    if not is_regenerate:
        await ingest_memory_event(
            event_id=f"gateway:{turn_id}:user",
            source_client="kiwi_gateway",
            conversation_id=session_id,
            role="user",
            content=user_message,
            occurred_at=occurred_at or datetime.now(timezone.utc),
            metadata={"transport": "gateway", "model": model, "project_id": project_id},
        )
    return {
        "turn_id": turn_id,
        "conversation_id": session_id,
        "model": model,
        "project_id": project_id,
        "is_regenerate": bool(is_regenerate),
    }


async def _archive_gateway_assistant(archive_context: dict | None, assistant_msg: str) -> dict | None:
    """Append one visible assistant response; content hash keeps variants distinct."""
    if not archive_context or not assistant_msg:
        return None
    clean_content = strip_dream_tag(strip_emotion_tag(assistant_msg))
    if not clean_content:
        return None
    content_hash = hashlib.sha256(clean_content.encode("utf-8")).hexdigest()[:16]
    result = await ingest_memory_event(
        event_id=f"gateway:{archive_context['turn_id']}:assistant:{content_hash}",
        source_client="kiwi_gateway",
        conversation_id=archive_context["conversation_id"],
        role="assistant",
        content=clean_content,
        occurred_at=datetime.now(timezone.utc),
        metadata={
            "transport": "gateway",
            "model": archive_context.get("model"),
            "project_id": archive_context.get("project_id"),
            "is_regenerate": bool(archive_context.get("is_regenerate")),
        },
    )
    return result


async def _finalize_stream_memories(mem_enabled, session_id, user_message, assistant_msg, model, emotion_level, project_id, is_regenerate, archive_context=None):
    """流式收尾的记忆工作（消息入库 + 情绪权重 + 按周期提取）。

    独立成协程，供上层用 _spawn_background_task 提为独立后台 task：客户端断连时
    生成器被取消，但本 task 作为独立任务仍跑完——消息入库、记忆提取、计数器清零都在
    process_memories_background 内原子完成，消除"断连丢消息 / 丢提取 / 计数器清了但没提取"的窗口。

    mem_enabled 由调用方传入（转发路径内联 get_memory_enabled()，工具路径用其 mem_enabled 参数，
    后者已含 skip_prompt 覆盖），保持与各自原有判定完全一致。
    """
    if archive_context and assistant_msg:
        try:
            await _archive_gateway_assistant(archive_context, assistant_msg)
        except Exception as exc:
            # The user utterance was already committed before the upstream call,
            # so a failed assistant append is visible as an incomplete turn and
            # can be replayed by the client instead of silently fabricating data.
            print(f"🚨 完整聊天归档失败（assistant）: {type(exc).__name__}: {exc}")
    if mem_enabled and user_message and assistant_msg:
        return await process_memories_background(
            session_id, user_message, assistant_msg, model,
            emotion_level=emotion_level, project_id=project_id, is_regenerate=is_regenerate,
        )
    return None

async def stream_and_capture(headers: dict, body: dict, session_id: str, user_message: str, model: str, tool_events: list = None, api_url: str = None, project_id: str = None, prompt_meta: dict = None, api_format: str = "openai", api_key: str = None, is_regenerate: bool = False, mem_enabled: bool = True, archive_context=None):
    """流式响应 + 捕获完整回复 + 工具事件"""
    _api_url = api_url or API_BASE_URL

    # 先发送衔接提示（如果有无缝切窗）
    if prompt_meta and prompt_meta.get("handoff"):
        yield f"data: {json.dumps({'ev_handoff': prompt_meta['handoff']}, ensure_ascii=False)}\n\n".encode("utf-8")

    # 先发送工具事件
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n".encode("utf-8")

    full_response = []
    _logged_first_delta = False
    _reasoning_chunks = 0

    # ---- 收尾补救（数据必活）----
    # 把"入库+提取"与"Dream 兜底"两个 spawn 收进闭包，由下方 try/finally 保证在任何退出路径
    # （正常完成 / 客户端断连取消 / 上游错误 return）都同步补出生。这堵住 Codex 实测②的丢数据点：
    # 客户端在正文最后一个字后立刻断开时，取消发生在"流收完 → 执行 spawn"之间的 await，
    # 裸 shield 保护不到"尚未 spawn"的窗口。finally 里绝不 await；create_task 是同步调用，安全。
    finalize_spawned = False
    dream_fb_spawned = False
    mem_task = None

    def _spawn_finalize_if_needed():
        nonlocal finalize_spawned, mem_task
        if finalize_spawned:
            return
        _asmsg = "".join(full_response)
        if not _asmsg:
            finalize_spawned = True
            return
        _emo = merge_emotion_levels(detect_emotion_from_user_msg(user_message), detect_emotion_from_response(_asmsg))
        mem_task = _spawn_background_task(
            _finalize_stream_memories(mem_enabled, session_id, user_message, _asmsg, model, _emo, project_id, is_regenerate, archive_context)
        )
        finalize_spawned = True

    def _spawn_dream_fallback_if_needed():
        nonlocal dream_fb_spawned
        if dream_fb_spawned:
            return
        dream_fb_spawned = True
        if detect_dream_trigger("".join(full_response)):
            _spawn_background_task(_dream_fallback_after_grace("auto"))

    try:
        # Anthropic 格式：转换请求体，使用流式适配器
        if api_format == "anthropic":
            send_body = to_anthropic_request(body)
            send_body["stream"] = True
            _headers = to_anthropic_headers(api_key or API_KEY)
            _headers["Accept-Encoding"] = "identity"

            async with httpx.AsyncClient(timeout=300) as client:
                async with client.stream("POST", _api_url, headers=_headers, json=send_body) as response:
                    if response.status_code != 200:
                        error_body = b""
                        async for chunk in response.aiter_bytes():
                            error_body += chunk
                        print(f"❌ Anthropic 流式请求失败 [{response.status_code}]: {error_body[:500].decode('utf-8', errors='ignore')}")
                        err_msg = f"⚠️ 请求失败 ({response.status_code})"
                        err_payload = json.dumps({'choices': [{'delta': {'content': err_msg}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)
                        yield f"data: {err_payload}\n\n".encode("utf-8")
                        yield b"data: [DONE]\n\n"
                        return

                    # 按 SSE 事件（\n\n）缓冲转发，并抑制上游 [DONE]——由本函数末尾统一发，保证它是流的最后一个事件
                    _ev_buf = ""
                    async for openai_chunk in anthropic_stream_to_openai(response, model):
                        _ev_buf += openai_chunk.decode("utf-8", errors="ignore").replace("\r\n", "\n")
                        while "\n\n" in _ev_buf:
                            event, _ev_buf = _ev_buf.split("\n\n", 1)
                            event = event.strip()
                            if not event or event == "data: [DONE]":
                                continue
                            yield (event + "\n\n").encode("utf-8")
                            try:
                                for line in event.split("\n"):
                                    line = line.strip()
                                    if line.startswith("data: "):
                                        data = json.loads(line[6:])
                                        delta = data.get("choices", [{}])[0].get("delta", {})
                                        content = delta.get("content", "")
                                        if content:
                                            full_response.append(content)
                                        if delta.get("reasoning_content"):
                                            _reasoning_chunks += 1
                            except Exception:
                                pass
                    _tail = _ev_buf.strip()
                    if _tail and _tail != "data: [DONE]":
                        yield (_tail + "\n\n").encode("utf-8")
        else:
            # OpenAI 格式：直接转发
            buffer = ""
            async with httpx.AsyncClient(timeout=300) as client:
                async with client.stream("POST", _api_url, headers=headers, json=body) as response:
                    if response.status_code != 200:
                        error_body = b""
                        async for chunk in response.aiter_bytes():
                            error_body += chunk
                        print(f"❌ 流式请求失败 [{response.status_code}]: {error_body[:500].decode('utf-8', errors='ignore')}")
                        err_msg = f"⚠️ 请求失败 ({response.status_code})"
                        err_payload = json.dumps({'choices': [{'delta': {'content': err_msg}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)
                        yield f"data: {err_payload}\n\n".encode("utf-8")
                        yield b"data: [DONE]\n\n"
                        return

                    # 按 SSE 事件（\n\n）缓冲转发，并抑制上游 [DONE]——由本函数末尾统一发，保证它是流的最后一个事件。
                    # 以「整事件」为单位转发（而非裸字节），既能干净拦掉 [DONE]，又不破坏事件分帧。
                    async for chunk in response.aiter_bytes(chunk_size=256):
                        buffer += chunk.decode("utf-8", errors="ignore").replace("\r\n", "\n")
                        while "\n\n" in buffer:
                            event, buffer = buffer.split("\n\n", 1)
                            event = event.strip()
                            if not event or event == "data: [DONE]":
                                continue
                            yield (event + "\n\n").encode("utf-8")
                            for line in event.split("\n"):
                                line = line.strip()
                                if not line.startswith("data: "):
                                    continue
                                try:
                                    data = json.loads(line[6:])
                                    delta = data.get("choices", [{}])[0].get("delta", {})

                                    # 🔍 调试日志：记录第一个有效delta的所有字段
                                    if not _logged_first_delta and delta:
                                        keys = list(delta.keys())
                                        if keys and keys != ['role']:
                                            print(f"🔍 [流式调试] 首个delta字段: {keys}, 模型: {model}")
                                            for k in ('reasoning_content', 'reasoning', 'reasoning_details'):
                                                if k in delta:
                                                    sample = str(delta[k])[:100]
                                                    print(f"🔍 [流式调试] {k} 示例: {sample}")
                                            _logged_first_delta = True

                                    if delta.get('reasoning_content') or delta.get('reasoning') or delta.get('reasoning_details'):
                                        _reasoning_chunks += 1

                                    content = delta.get("content", "")
                                    if content:
                                        full_response.append(content)
                                except (json.JSONDecodeError, KeyError, IndexError):
                                    pass
                    _tail = buffer.strip()
                    if _tail and _tail != "data: [DONE]":
                        yield (_tail + "\n\n").encode("utf-8")
    finally:
        # 断连（取消/GeneratorExit）或正常完成都在此把收尾 task 与 Dream 兜底同步补出生。
        _spawn_finalize_if_needed()
        _spawn_dream_fallback_if_needed()

    # 下面仅在正常完成时执行（断连时异常已带走控制权）：把 ev_memory / ev_dream / [DONE] 推给客户端
    assistant_msg = "".join(full_response)

    # 🔍 流式完成汇总
    print(f"🔍 [流式完成] 模型={model}, 正文={len(assistant_msg)}字, 思考链chunks={_reasoning_chunks}")
    if _reasoning_chunks == 0 and '<think>' in assistant_msg:
        print(f"🔍 [流式完成] ⚠️ 思考链在正文中（<think>标签），前端需要解析")

    dream_triggered = detect_dream_trigger(assistant_msg)

    # 记忆行照旧实时出现在聊天里：mem_task 已在 finally 里 spawn；仍用 shield（await 被取消不连带取消 mem_task）
    if mem_task is not None:
        try:
            mem_result = await asyncio.shield(mem_task)
        except asyncio.CancelledError:
            raise
        except Exception as e:
            print(f"⚠️ 收尾记忆 task 出错（不影响流收尾）: {e}")
            mem_result = None
        if mem_result and mem_result.get("action") != "skip":
            yield f"data: {json.dumps({'ev_memory': mem_result}, ensure_ascii=False)}\n\n".encode("utf-8")

    # Dream 触发：在 [DONE] 之前推送，前端据此启动 Dream 并展示进度（既有"陪看"体验不动）
    if dream_triggered:
        print(f"🌙 检测到 Dream 标记，通知前端启动 Dream（后端已排兜底）...")
        yield f"data: {json.dumps({'ev_dream': {'triggered': True}}, ensure_ascii=False)}\n\n".encode("utf-8")

    # [DONE] 作为流的最后一个事件（上游/适配器的 [DONE] 已在上面被抑制）
    yield b"data: [DONE]\n\n"

# ============================================================
# 记忆管理接口
# ============================================================

@app.get("/debug/memories")
async def debug_memories(
    q: str = "",
    limit: int = 20,
    offset: int = 0,
    sort: str = "newest",
    category_id: int = None,
    min_importance: int = None,
    track_recall: bool = False,
):
    """查看和搜索记忆（支持分类筛选、分页、排序、重要度过滤）。

    返回字段名为 memories（前端约定），输出每条包含 is_permanent 和 heat
    （前端用于显示锁定 🔒 和热度 🔥 badge）。
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用（设置 MEMORY_ENABLED=true 开启）"}

    # 限制查询范围，防止过大请求消耗资源
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    try:
        if q:
            # 搜索路径：search_memories 已经返回 is_permanent / heat
            memories = await search_memories(q, limit=limit + offset, track_recall=track_recall)
            if category_id is not None:
                memories = [m for m in memories if m.get("category_id") == category_id]
            if min_importance is not None:
                memories = [m for m in memories if m.get("importance", 0) >= min_importance]
            memories = memories[offset:offset + limit]
        else:
            # 非搜索路径：get_recent_memories 不算 heat, 这里在端点层补上
            memories = await get_recent_memories(limit=limit + offset, category_id=category_id)
            memories = [dict(m) for m in memories]
            if min_importance is not None:
                memories = [m for m in memories if m.get("importance", 0) >= min_importance]
            # 排序
            if sort == "oldest":
                memories.sort(key=lambda m: m.get("created_at") or "")
            elif sort == "importance":
                memories.sort(key=lambda m: m.get("importance", 0), reverse=True)
            elif sort == "heat":
                # heat 此路径不算，按 access_count 近似
                memories.sort(key=lambda m: m.get("access_count", 0) or 0, reverse=True)
            memories = memories[offset:offset + limit]

        # 总数复用与列表相同的 WHERE（含类型排除、valid_until、分类过滤），作分页分母
        from database import get_memories_count
        total = await get_memories_count(category_id=category_id)

        return {
            "total_memories": total,
            "query": q or "(最近记忆)",
            "memories": [
                {
                    "id": m["id"],
                    "title": m.get("title", ""),
                    "content": m["content"],
                    "importance": m["importance"],
                    "is_permanent": m.get("is_permanent", False),
                    "heat": m.get("heat", 0),
                    "created_at": str(m["created_at"]),
                    "memory_type": m.get("memory_type", "fragment"),
                    "category_id": m.get("category_id"),
                    "category_name": m.get("category_name", ""),
                    "category_color": m.get("category_color", ""),
                    "source": m.get("source", "ai_extracted"),
                    "resolution": m.get("resolution", 1.0),
                }
                for m in memories
            ],
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/memory-archive")
async def debug_memory_archive(q: str = "", limit: int = 20):
    """显式检索已经退出日常召回的冷记忆与历史版本。"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        memories = await search_archived_memories(q, limit=limit)
        return {
            "query": q or "(最近归档)",
            "total_memories": len(memories),
            "memories": [
                {
                    **memory,
                    "created_at": str(memory.get("created_at")),
                    "valid_until": str(memory.get("valid_until")) if memory.get("valid_until") else None,
                    "archived_at": str(memory.get("archived_at")) if memory.get("archived_at") else None,
                }
                for memory in memories
            ],
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/memories/{memory_id}/versions")
async def debug_memory_versions(memory_id: int):
    """查看一条记忆的不可变原文与所有派生版本。"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        versions = await get_memory_versions(memory_id)
        if not versions:
            return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在或没有版本"})
        return {
            "memory_id": memory_id,
            "versions": [
                {
                    **version,
                    "created_at": str(version.get("created_at")),
                }
                for version in versions
            ],
        }
    except Exception as e:
        return {"error": str(e)}


@app.delete("/debug/memories/{memory_id}")
async def delete_single_memory(memory_id: int, force: bool = False):
    """删除单条记忆；锁定记忆默认受保护，显式 force 才放行。"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        status = await delete_memory(memory_id, force=force)
        if status == "deleted":
            total = await get_all_memories_count()
            return {"status": "deleted", "memory_id": memory_id, "remaining": total}
        if status == "protected":
            return JSONResponse(
                status_code=403,
                content={"error": f"记忆 #{memory_id} 已锁定，请先解锁或显式使用 force=true"},
            )
        return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在"})
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/batch-delete")
async def batch_delete_memories(request: Request):
    """批量删除记忆；只有 JSON 布尔 true 可强制删除锁定记忆。"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        body = await request.json()
        ids = body.get("ids", [])
        force = body.get("force") is True
        if not ids:
            return {"error": "ids 不能为空"}
        result = await batch_delete_memories_guarded(ids, force=force)
        total = await get_all_memories_count()
        response = {"status": "deleted", "remaining": total, **result}
        if result["rejected"]:
            response["rejected_reason"] = "锁定记忆需先解锁或显式使用 JSON force=true"
        return response
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/batch-update")
async def batch_update_memories(request: Request):
    """批量更新记忆字段（importance / category_id / is_permanent）"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        body = await request.json()
        ids = body.get("ids", [])
        if not ids:
            return {"error": "ids 不能为空"}
        
        importance = body.get("importance")
        category_id = body.get("category_id", "UNSET")
        is_permanent = body.get("is_permanent")
        
        pool = await get_pool()
        async with pool.acquire() as conn:
            # 构建动态 SET 子句
            sets = []
            vals = []
            idx = 1
            if importance is not None:
                sets.append(f"importance = ${idx}")
                vals.append(importance)
                idx += 1
            if category_id != "UNSET":
                sets.append(f"category_id = ${idx}")
                vals.append(category_id)
                idx += 1
            if is_permanent is not None:
                sets.append(f"is_permanent = ${idx}")
                permanent_value = bool(is_permanent)
                vals.append(permanent_value)
                idx += 1
                sets.append("lock_source = 'user'" if permanent_value else "lock_source = NULL")
            
            if not sets:
                return {"error": "没有提供更新字段"}
            
            vals.append(ids)
            sql = f"UPDATE memories SET {', '.join(sets)} WHERE id = ANY(${idx}::int[])"
            await conn.execute(sql, *vals)
        
        return {"status": "updated", "count": len(ids)}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/debug/memories")
async def clear_memories(request: Request):
    """清空全部记忆；必须同时通过严格 force 与确认短语双闸。"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict) or not (
            body.get("force") is True
            and body.get("confirm") == "DELETE_ALL_MEMORIES"
        ):
            return JSONResponse(
                status_code=400,
                content={
                    "error": "清空全部记忆需同时携带 JSON force=true 与 confirm='DELETE_ALL_MEMORIES'"
                },
            )
        count = await clear_all_memories()
        return {"status": "cleared", "deleted_count": count}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/memory-heat")
async def debug_memory_heat(limit: int = 50):
    """
    记忆热度报告（v5.2）
    查看每条记忆的热度、召回次数、情绪浓度、查询多样性
    """
    try:
        from database import get_memory_heat_report
        report = await get_memory_heat_report(limit=min(limit, 200))
        
        # 统计摘要
        if report:
            hot = sum(1 for r in report if r["heat"] > 0.7)
            warm = sum(1 for r in report if 0.3 < r["heat"] <= 0.7)
            cold = sum(1 for r in report if r["heat"] <= 0.3)
            total_recalls = sum(r["access_count"] for r in report)
            emotional = sum(1 for r in report if r["emotional_weight"] > 0)
        else:
            hot = warm = cold = total_recalls = emotional = 0
        
        return {
            "summary": {
                "total": len(report),
                "hot": hot,
                "warm": warm,
                "cold": cold,
                "total_recalls": total_recalls,
                "emotional_memories": emotional,
            },
            "memories": report,
        }
    except Exception as e:
        return {"error": str(e)}


@app.put("/debug/memories/{memory_id}")
async def update_single_memory(memory_id: int, request: Request):
    """
    更新单条记忆
    请求体示例：{"content": "新内容", "importance": 8}
    可以只传其中一个字段
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        body = await request.json()
        content = body.get("content")
        importance = body.get("importance")
        title = body.get("title")
        # category_id: None清除分类, int设置分类, 不传不改
        cat_id = body.get("category_id", "UNSET")
        
        success = await update_memory(memory_id, content=content, importance=importance, title=title, category_id=cat_id)
        if success:
            return {"status": "updated", "memory_id": memory_id}
        else:
            return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在或没有提供更新内容"})
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories")
async def add_memory_manual(request: Request):
    """
    手动添加记忆
    请求体示例：{"content": "知知喜欢喝奶茶", "memory_kind": "user_fact", "importance": 7}
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        body = await request.json()
        content = body.get("content", "")
        importance = body.get("importance", 5)
        title = body.get("title", "")
        category_id = body.get("category_id")
        memory_kind = str(body.get("memory_kind") or "").strip().lower() or None
        
        if not content:
            return JSONResponse(status_code=400, content={"error": "content 不能为空"})

        prepared, voice_errors = prepare_generated_memory(content, title, memory_kind)
        if voice_errors:
            return JSONResponse(status_code=422, content={"error": "记忆违反第一人称身份契约", "details": voice_errors})
        new_id = await save_memory(
            content=prepared["content"],
            importance=importance,
            source_session="manual",
            title=prepared["title"],
            category_id=category_id,
            source="user_explicit",
            memory_kind=prepared["memory_kind"],
            enforce_identity=True,
        )
        total = await get_all_memories_count()
        return {"status": "added", "id": new_id, "content": prepared["content"], "importance": importance, "title": prepared["title"], "total": total}
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/{memory_id}/toggle-permanent")
async def toggle_memory_permanent(memory_id: int):
    """切换记忆的锁定状态"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT id, COALESCE(is_permanent, false) as is_permanent FROM memories WHERE id = $1",
                memory_id
            )
            if not row:
                return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在"})
            
            new_val = not row["is_permanent"]
            await conn.execute(
                "UPDATE memories SET is_permanent = $1, lock_source = $2 WHERE id = $3",
                new_val, "user" if new_val else None, memory_id
            )
            status = "locked" if new_val else "unlocked"
            print(f"🔒 记忆 #{memory_id} {'锁定' if new_val else '解锁'}")
            return {"status": status, "memory_id": memory_id, "is_permanent": new_val}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/migrate-embeddings")
async def api_migrate_embeddings():
    """为所有缺少向量的记忆生成 embedding"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        result = await migrate_embeddings()
        return result
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/embedding-stats")
async def api_embedding_stats():
    """查看 embedding 覆盖率"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        stats = await get_embedding_stats()
        return stats
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/extract-now")
async def api_extract_now(request: Request):
    """
    手动触发记忆提取（从最近对话中提取记忆）
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        # 解析 project_id
        project_id = None
        try:
            body = await request.json()
            project_id = body.get("project_id")
            session_id = str(body.get("session_id") or "").strip() or None
        except Exception:
            session_id = None
        # 手动提取也只选择一个明确 session；未指定时选择最近活跃 session。
        extract_interval = await get_extract_interval()
        session_id = session_id or await get_latest_conversation_session_id()
        recent_msgs = await get_recent_messages(session_id, limit=extract_interval * 2) if session_id else []
        if not recent_msgs:
            return {"status": "ok", "action": "extract", "saved": 0, "skipped": 0, "message": "没有最近的对话可提取"}
        
        messages_for_extraction = [
            {"role": row["role"], "content": row["content"]}
            for row in recent_msgs
        ]
        
        # 获取对比用的已有记忆
        user_text = " ".join(r["content"] for r in recent_msgs if r["role"] == "user")
        related = await search_memories(user_text[:500], limit=50, track_recall=False, project_id=project_id)
        recent = await get_recent_memories(limit=30, project_id=project_id)
        seen = set()
        existing_contents = []
        for content in [r["content"] for r in related] + [r["content"] for r in recent]:
            if content not in seen:
                seen.add(content)
                existing_contents.append(content)
        
        # 获取分类
        try:
            all_cats = await get_all_categories()
            cat_names = [c["name"] for c in all_cats]
        except Exception:
            cat_names = []
        
        from config import get_config
        db_memory_model = await get_config("default_memory_model")
        db_memory_prompt = await get_config("prompt_memory_extract")
        
        new_memories = await extract_memories(
            messages_for_extraction,
            existing_memories=existing_contents,
            categories=cat_names,
            model_override=db_memory_model if db_memory_model else None,
            prompt_override=db_memory_prompt if db_memory_prompt else None,
        )
        
        # 过滤 + 去重 + 保存
        META_BLACKLIST = [
            "记忆库", "记忆系统", "检索", "没有被记录", "没有被提取",
            "记忆遗漏", "尚未被记录", "写入不完整", "检索功能",
            "系统没有返回", "关键词匹配", "语义匹配", "语义检索",
            "阈值", "数据库", "seed", "导入", "部署",
            "bug", "debug", "端口", "网关",
        ]
        saved_count = 0
        skipped_count = 0
        session_id = "manual-" + str(uuid.uuid4())[:8]
        
        for mem in new_memories:
            if any(kw in mem["content"] for kw in META_BLACKLIST):
                continue
            is_dup, _ = await check_memory_duplicate(mem["content"], new_title=mem.get("title", ""), project_id=project_id)
            if is_dup:
                skipped_count += 1
                continue
            cat_id = None
            cat_hint = mem.get("category", "")
            if cat_hint:
                cat_id = await match_category_by_name(cat_hint)
            await save_memory(
                content=mem["content"],
                importance=mem["importance"],
                source_session=session_id,
                title=mem.get("title", ""),
                category_id=cat_id,
                source="manual_extracted",
                emotional_weight=mem.get("emotional_weight", 0),
                project_id=project_id,
                memory_kind=mem.get("memory_kind"),
                enforce_identity=True,
            )
            saved_count += 1
        
        total = await get_all_memories_count()
        return {"status": "ok", "action": "extract", "saved": saved_count, "skipped": skipped_count, "total": total}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/daily-digest")
async def api_daily_digest(date: str = None):
    """
    手动触发每日记忆整理
    ?date=2026-03-02  指定日期整理
    不传 date 则整理昨天的
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        from daily_digest import run_daily_digest
        from config import get_config
        db_digest_model = await get_config("default_digest_model")
        db_digest_prompt = await get_config("prompt_daily_digest")
        result = await run_daily_digest(
            target_date=date,
            model_override=db_digest_model if db_digest_model else None,
            prompt_override=db_digest_prompt if db_digest_prompt else None,
        )
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 日历记忆页面接口（v5.0 记忆桥）
# ============================================================

@app.get("/admin/day-page")
async def api_generate_day_page(date: str = None):
    """手动触发日页面生成 ?date=2026-04-01"""
    try:
        from daily_digest import generate_day_page
        result = await generate_day_page(target_date=date)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/week-summary")
async def api_generate_week_summary(start: str = None, end: str = None):
    """手动触发周总结 ?start=2026-03-31&end=2026-04-06"""
    try:
        from daily_digest import generate_week_summary
        if not start or not end:
            from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            # 默认上周一到上周日
            days_since_monday = now.weekday()
            last_monday = now - td(days=days_since_monday + 7)
            last_sunday = last_monday + td(days=6)
            start = last_monday.strftime("%Y-%m-%d")
            end = last_sunday.strftime("%Y-%m-%d")
        result = await generate_week_summary(start, end)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/month-summary")
async def api_generate_month_summary(month: str = None):
    """手动触发月总结 ?month=2026-03"""
    try:
        from daily_digest import generate_month_summary
        if not month:
            from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            last_month_end = now.replace(day=1) - td(days=1)
            month = last_month_end.strftime("%Y-%m")
        # 解析月份
        year, mon = month.split("-")
        import calendar as cal_mod
        last_day = cal_mod.monthrange(int(year), int(mon))[1]
        start = f"{month}-01"
        end = f"{month}-{last_day:02d}"
        result = await generate_month_summary(start, end, month)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/quarter-summary")
async def api_generate_quarter_summary(quarter: str = None):
    """手动触发季度总结 ?quarter=2026-Q1"""
    try:
        from daily_digest import generate_period_summary
        import calendar as cal_mod
        from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls

        if not quarter:
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            cur_q = (now.month - 1) // 3 + 1
            target_q = cur_q - 1
            target_y = now.year
            if target_q <= 0:
                target_q = 4
                target_y -= 1
            quarter = f"{target_y}-Q{target_q}"

        parts = quarter.split("-Q")
        if len(parts) != 2:
            return {"error": f"无效格式: {quarter!r}，需要 YYYY-QN（如 2026-Q1）"}
        year = int(parts[0])
        q = int(parts[1])
        if q < 1 or q > 4:
            return {"error": f"无效季度: Q{q}，需要 Q1-Q4"}
        q_start_month = (q - 1) * 3 + 1
        q_end_month = q_start_month + 2
        start = f"{year}-{q_start_month:02d}-01"
        q_end_day = cal_mod.monthrange(year, q_end_month)[1]
        end = f"{year}-{q_end_month:02d}-{q_end_day:02d}"
        label = f"{year}Q{q}"

        result = await generate_period_summary(start, end, "quarter", label, "月总结")
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/year-summary")
async def api_generate_year_summary(year: str = None):
    """手动触发年度总结 ?year=2025"""
    try:
        from daily_digest import generate_period_summary
        from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls

        if not year:
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            year = str(now.year - 1)

        y = int(year)
        start = f"{y}-01-01"
        end = f"{y}-12-31"

        result = await generate_period_summary(start, end, "year", str(y), "季度总结")
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 自动上下文压缩摘要（v6.1）
# 压缩在前端执行，后端只负责存储/读取摘要，为无缝换窗 v2 预留
# ============================================================

@app.post("/admin/compression-summary")
async def api_save_compression_summary(request: Request):
    """前端压缩成功后调此端点存储摘要（为无缝换窗 v2 预留）"""
    try:
        body = await request.json()
        conv_id = body.get("conversation_id")
        summary = body.get("summary", "")
        if not conv_id or not summary:
            return JSONResponse(status_code=400, content={"error": "缺少 conversation_id 或 summary"})
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO compression_summaries (conversation_id, summary, model, summary_type, msg_count) VALUES ($1, $2, $3, $4, $5)",
                conv_id,
                summary,
                body.get("model", ""),
                body.get("summary_type", "auto"),
                body.get("msg_count", 0),
            )
        return JSONResponse(content={"ok": True})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/admin/compression-summaries")
async def api_get_compression_summaries(conversation_id: str):
    """读取某对话的全部压缩摘要（按时间正序），供无缝换窗 v2 使用"""
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT summary, model, summary_type, msg_count, compressed_at FROM compression_summaries WHERE conversation_id = $1 ORDER BY compressed_at ASC",
                conversation_id,
            )
        return JSONResponse(content=[
            {
                "summary": r["summary"],
                "model": r["model"],
                "summary_type": r["summary_type"],
                "msg_count": r["msg_count"],
                "compressed_at": str(r["compressed_at"]),
            }
            for r in rows
        ])
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/calendar/{date}")
async def api_get_calendar_day(date: str, type: str = "day"):
    """获取指定日期的日历页面"""
    try:
        from database import get_calendar_page
        page = await get_calendar_page(date, type)
        if not page:
            return {"status": "ok", "page": None}
        # 序列化 date 对象
        page["date"] = str(page["date"])
        if page.get("created_at"):
            page["created_at"] = page["created_at"].isoformat()
        if page.get("updated_at"):
            page["updated_at"] = page["updated_at"].isoformat()
        return {"status": "ok", "page": page}
    except Exception as e:
        return {"error": str(e)}


@app.get("/calendar")
async def api_get_calendar_range(start: str = None, end: str = None, type: str = None):
    """获取一段时间的日历页面 ?start=2026-03-25&end=2026-04-01&type=day"""
    try:
        from database import get_calendar_range
        if not start or not end:
            # 默认最近7天
            from datetime import timedelta as td, timezone as tz_mod
            TZ = tz_mod(td(hours=8))
            from datetime import datetime as dt_cls
            now = dt_cls.now(TZ)
            if not end:
                end = now.strftime("%Y-%m-%d")
            if not start:
                start = (now - td(days=7)).strftime("%Y-%m-%d")
        pages = await get_calendar_range(start, end, type)
        for p in pages:
            p["date"] = str(p["date"])
            if p.get("created_at"):
                p["created_at"] = p["created_at"].isoformat()
            if p.get("updated_at"):
                p["updated_at"] = p["updated_at"].isoformat()
        return {"status": "ok", "pages": pages, "count": len(pages)}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/calendar/{date}")
async def api_save_calendar_page(date: str, req: Request):
    """用户手动编辑/创建日历页面"""
    try:
        from database import update_calendar_page_user_edit
        body = await req.json()
        content = body.get("content", "")
        title = body.get("title", "")
        page_type = body.get("type", "day")
        # 用户编辑只写 diary（content）与 title；页面已有的 AI 内容
        # （sections/keywords/summary/digest/model_used）一律保留，不再被空值覆盖。
        page_id = await update_calendar_page_user_edit(
            date_str=date,
            page_type=page_type,
            diary=content,
            title=title,
        )
        return {"status": "ok", "id": page_id}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/calendar/{date}")
async def api_delete_calendar_page(date: str, type: str = "day"):
    """删除指定日期的日历页面"""
    try:
        from database import delete_calendar_page
        ok = await delete_calendar_page(date, type)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 评论接口（v5.0 记忆桥通用）
# ============================================================

@app.post("/comments")
async def api_create_comment(req: Request):
    """创建评论"""
    try:
        from database import create_comment
        body = await req.json()
        comment = await create_comment(
            target_type=body["target_type"],
            target_id=body["target_id"],
            content=body["content"],
            author=body.get("author", "user"),
            parent_id=body.get("parent_id"),
        )
        if comment and comment.get("created_at"):
            comment["created_at"] = comment["created_at"].isoformat()
        return {"status": "ok", "comment": comment}
    except Exception as e:
        return {"error": str(e)}


@app.get("/comments")
async def api_get_comments(target_type: str, target_id: int):
    """获取评论列表 ?target_type=day_page&target_id=1"""
    try:
        from database import get_comments
        comments = await get_comments(target_type, target_id)
        for c in comments:
            if c.get("created_at"):
                c["created_at"] = c["created_at"].isoformat()
        return {"status": "ok", "comments": comments}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/comments/{comment_id}")
async def api_delete_comment(comment_id: int):
    """删除评论"""
    try:
        from database import delete_comment
        ok = await delete_comment(comment_id)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# Dream 接口（v5.1）
# ============================================================

@app.post("/dream/start")
async def api_dream_start(req: Request):
    """
    触发 Dream，返回 SSE 事件流
    Body: {"trigger_type": "manual"} (可选)
    """
    from starlette.responses import StreamingResponse
    from dream import run_dream

    body = {}
    try:
        body = await req.json()
    except Exception:
        pass

    trigger = body.get("trigger_type", "manual")

    async def event_generator():
        async for event in run_dream(trigger_type=trigger):
            event_type = event.get("type", "message")
            data = event.get("data", "")
            if isinstance(data, dict):
                data = json.dumps(data, ensure_ascii=False)
            yield f"event: {event_type}\ndata: {data}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/dream/start-detached")
async def api_dream_start_detached(req: Request):
    """detached 触发 Dream：立即返回 JSON，不开 SSE，Dream 在后台独立跑完（不依赖本连接）。

    供 MCP trigger_dream 工具等"只要把梦发起、不需要盯进度"的场景使用，避免读一行 SSE
    就断连把刚起来的梦掐死。已有 Dream 在跑时返回 already_running（幂等）。
    """
    body = {}
    try:
        body = await req.json()
    except Exception:
        pass
    trigger = body.get("trigger_type", "manual")

    if await _dream_is_running():
        return {"status": "already_running"}
    _launch_dream_detached(trigger)
    return {"status": "started"}


@app.post("/dream/stop")
async def api_dream_stop():
    """中断正在进行的 Dream"""
    from dream import stop_dream
    return await stop_dream()


@app.post("/admin/dream/force-stop")
async def api_dream_force_stop():
    """强制清理卡住的 Dream（直接更新数据库状态）"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await conn.execute("""
            UPDATE dream_logs SET status = 'interrupted', finished_at = NOW(),
                dream_narrative = COALESCE(dream_narrative, '') || '\n[手动强制中断]'
            WHERE status = 'running'
        """)
    return {"status": "ok", "message": f"已强制中断所有 running 状态的 Dream", "result": str(result)}


@app.get("/dream/status")
async def api_dream_status():
    """获取当前 Dream 状态"""
    try:
        from database import get_dream_status, get_unprocessed_memories
        from config import get_config
        status = await get_dream_status()
        unprocessed = await get_unprocessed_memories()
        last_dream_date = await get_config("last_dream_date")
        drowsy_threshold = int(await get_config("dream_drowsy_threshold") or "30")

        # 序列化时间
        for key in ("current", "last_completed"):
            if status.get(key):
                for field in ("started_at", "finished_at"):
                    if status[key].get(field):
                        status[key][field] = status[key][field].isoformat()

        return {
            "status": "ok",
            **status,
            "unprocessed_count": len(unprocessed),
            # 前端 admin-panel 读 unprocessed_fragments / is_dreaming, 这里加别名保证两套字段都能用
            "unprocessed_fragments": len(unprocessed),
            "is_dreaming": status.get("is_running", False),
            "drowsy_threshold": drowsy_threshold,
            "is_drowsy": len(unprocessed) >= drowsy_threshold,
            "last_dream_date": last_dream_date,
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/dream/history")
async def api_dream_history(limit: int = 10):
    """获取 Dream 执行历史"""
    try:
        from database import get_dream_history
        history = await get_dream_history(limit)
        for h in history:
            for field in ("started_at", "finished_at"):
                if h.get(field):
                    h[field] = h[field].isoformat()
        return {"status": "ok", "history": history}
    except Exception as e:
        return {"error": str(e)}


@app.get("/dream/scenes")
async def api_get_scenes():
    """获取所有活跃的记忆场景"""
    try:
        from database import get_active_scenes
        scenes = await get_active_scenes()
        for s in scenes:
            for field in ("created_at", "updated_at"):
                if s.get(field):
                    s[field] = s[field].isoformat()
        return {"status": "ok", "scenes": scenes, "count": len(scenes)}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/dream/{dream_id}")
async def api_delete_dream(dream_id: int):
    """软删除一条 Dream 日志，并在同一事务内软删除其来源场景。"""
    try:
        from database import get_pool
        pool = await get_pool()
        async with pool.acquire() as conn:
            async with conn.transaction():
                row = await conn.fetchrow(
                    """
                    UPDATE dream_logs
                    SET deleted = TRUE
                    WHERE id = $1 AND NOT deleted
                    RETURNING id
                    """,
                    dream_id,
                )
                if not row:
                    return {"error": f"Dream #{dream_id} 不存在或已删除"}
                await conn.execute(
                    """
                    UPDATE mem_scenes
                    SET status = 'deleted'
                    WHERE created_by_dream_id = $1
                      AND status <> 'deleted'
                    """,
                    dream_id,
                )
        return {"status": "ok", "deleted": dream_id}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/scene/{scene_id}")
async def api_update_scene(scene_id: int, req: Request):
    """用户手动编辑记忆场景（标题、叙事、远见）"""
    try:
        from database import update_mem_scene
        body = await req.json()
        kwargs = {}
        if "title" in body:
            kwargs["title"] = body["title"]
        if "narrative" in body:
            kwargs["narrative"] = body["narrative"]
        if "foresight" in body:
            kwargs["foresight"] = body["foresight"]
        if not kwargs:
            return {"error": "没有可更新的字段"}
        ok = await update_mem_scene(scene_id, **kwargs)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 动态配置管理接口（v3.1）
# ============================================================

@app.get("/admin/config")
async def api_get_config():
    """获取所有配置"""
    try:
        config = await get_all_config()
        return {"status": "ok", "config": config}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/config/{key}")
async def api_set_config(key: str, request: Request):
    """更新单个配置"""
    try:
        data = await request.json()
        value = str(data.get("value", ""))
        success = await set_config(key, value)
        if success:
            return {"status": "updated", "key": key, "value": value}
        else:
            return {"error": f"无效的配置项或值: {key}={value}"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# Prompt 出厂默认值管理（v5.6）
# ============================================================

# ============================================================
# 无缝换窗 v2：源对话概要压缩（同步调用）
# ============================================================

async def _compress_for_handoff(existing_summary: str, messages: list):
    """把源对话的（已有概要 + 待压消息）同步压成一段概要，失败返回 None。

    镜像前端 compressContext 做法——压缩 prompt 走 system、对话文本走 user，
    不使用 {messages} 占位符模板。
    """
    lines = []
    if existing_summary:
        lines.append(f"[之前的对话摘要]\n{existing_summary}\n")
    for m in messages:
        role_label = "用户" if m.get("role") == "user" else "助手"
        content = m.get("content", "") or ""
        if len(content) > 500:
            content = content[:500] + "…（截断）"
        if content:
            lines.append(f"{role_label}: {content}")

    transcript = "\n".join(lines).strip()
    if not transcript:
        return None

    use_model = (
        await get_config("handoff_summary_model")
        or await get_config("default_compress_model")
        or os.getenv("MEMORY_MODEL", "anthropic/claude-haiku-4")
    )
    compress_prompt = await get_config("prompt_compress") or (
        "请将以下对话内容压缩为简洁的摘要，保留关键信息、话题走向和情感基调。"
        "不要截断正在进行中的话题。"
    )

    try:
        from database import resolve_model_endpoint
        use_api_url, use_api_key, use_api_format = await resolve_model_endpoint(use_model)
    except Exception:
        use_api_url = os.getenv("MEMORY_API_BASE_URL", "") or os.getenv("API_BASE_URL", "https://openrouter.ai/api/v1/chat/completions")
        if not use_api_url.rstrip("/").endswith("/chat/completions"):
            use_api_url = f"{use_api_url.rstrip('/')}/chat/completions"
        use_api_key = os.getenv("MEMORY_API_KEY", "") or os.getenv("API_KEY", "")
        use_api_format = "openai"

    try:
        from anthropic_adapter import prepare_background_request, parse_background_response
        _body = {
            "model": use_model,
            "max_tokens": 2000,
            "messages": [
                {"role": "system", "content": compress_prompt},
                {"role": "user", "content": transcript},
            ],
        }
        _headers, _send_body = prepare_background_request(use_api_key, use_api_format, _body)
        async with httpx.AsyncClient(timeout=20) as client:
            response = await client.post(use_api_url, headers=_headers, json=_send_body)

        if response.status_code == 200:
            data = parse_background_response(response.json(), use_api_format)
            summary = data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            if summary:
                return summary

        print(f"⚠️  换窗概要压缩失败: HTTP {response.status_code}")
    except Exception as e:
        print(f"⚠️  换窗概要压缩失败: {e}")
    return None


def _get_factory_prompts() -> dict:
    """收集所有出厂默认 prompt（从各模块常量中读取）"""
    from memory_extractor import EXTRACTION_PROMPT, EMOTION_HIGH_INSTRUCTION
    from daily_digest import (
        DIGEST_PROMPT, DEFAULT_PROFILE_PROMPT, DAY_PAGE_PROMPT,
        WEEK_SUMMARY_PROMPT, MONTH_SUMMARY_PROMPT, PERIOD_SUMMARY_PROMPT,
    )
    from dream import DREAM_PROMPT
    return {
        "prompt_memory_extract":    EXTRACTION_PROMPT,
        "prompt_daily_digest":      DIGEST_PROMPT,
        "prompt_user_profile":      DEFAULT_PROFILE_PROMPT,
        "prompt_daily_digest_page": DAY_PAGE_PROMPT,
        "prompt_weekly_summary":    WEEK_SUMMARY_PROMPT,
        "prompt_monthly_summary":   MONTH_SUMMARY_PROMPT,
        "prompt_period_summary":    PERIOD_SUMMARY_PROMPT,
        "prompt_dream":             DREAM_PROMPT,
    }


@app.get("/admin/default-prompts")
async def api_get_default_prompts():
    """获取所有出厂默认 prompt（供前端「恢复默认」按钮使用）"""
    try:
        factory = _get_factory_prompts()
        return {"status": "ok", "prompts": factory}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/restore-prompt/{key}")
async def api_restore_prompt(key: str):
    """将指定 prompt 恢复为出厂默认值"""
    try:
        factory = _get_factory_prompts()
        if key not in factory:
            return {"error": f"未知的 prompt 配置项: {key}"}
        default_value = factory[key]
        success = await set_config(key, default_value)
        if success:
            print(f"🔄 已恢复默认 prompt: {key}")
            return {"status": "restored", "key": key, "length": len(default_value)}
        else:
            return {"error": f"写入失败: {key}"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 供应商管理 API
# ============================================================

@app.get("/admin/providers")
async def api_get_providers():
    """获取所有供应商。

    注意：必须脱敏 api_key —— 数据库里存的是 LLM 服务商的原始 key，
    直接 dict(p) 送给前端会让任何能调到这个接口的人拿到完整 key。
    前端 admin-panel 读的是 api_key_preview, 这里把 api_key 替换成
    preview 字段, 原始 key 不出后端。
    """
    try:
        providers = await get_all_providers()
        result = []
        for p in providers:
            sp = dict(p)
            raw_key = sp.pop("api_key", "") or ""
            if raw_key:
                if len(raw_key) > 12:
                    sp["api_key_preview"] = raw_key[:6] + "…" + raw_key[-4:]
                else:
                    sp["api_key_preview"] = "•" * min(len(raw_key), 8)
            else:
                sp["api_key_preview"] = ""
            if sp.get("created_at"):
                sp["created_at"] = sp["created_at"].isoformat()
            if sp.get("updated_at"):
                sp["updated_at"] = sp["updated_at"].isoformat()
            result.append(sp)
        return {"status": "ok", "providers": result}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/providers")
async def api_create_provider(request: Request):
    """创建供应商"""
    try:
        data = await request.json()
        name = data.get("name", "").strip()
        api_base_url = data.get("api_base_url", "").strip()
        api_key = data.get("api_key", "").strip()
        enabled = data.get("enabled", True)

        if not name:
            return {"error": "供应商名称不能为空"}
        if not api_base_url:
            return {"error": "API Base URL 不能为空"}

        provider = await create_provider(name, api_base_url, api_key, enabled, api_format=data.get("api_format", "openai"))
        return {"status": "created", "provider": provider}
    except Exception as e:
        print(f"⚠️  创建供应商失败: {e}")
        return {"error": str(e)}


@app.put("/admin/providers/{provider_id}")
async def api_update_provider(provider_id: int, request: Request):
    """更新供应商"""
    try:
        data = await request.json()
        # 编辑时留空的 api_key 视为「不修改」，避免把已存的 key 清空（前端会发空串）
        if not (data.get("api_key") or "").strip():
            data.pop("api_key", None)
        provider = await update_provider(provider_id, **data)
        if provider:
            return {"status": "updated", "provider": provider}
        return {"error": "供应商不存在"}
    except Exception as e:
        print(f"⚠️  更新供应商失败: {e}")
        return {"error": str(e)}


@app.delete("/admin/providers/{provider_id}")
async def api_delete_provider(provider_id: int):
    """删除供应商"""
    try:
        success = await delete_provider(provider_id)
        if success:
            return {"status": "deleted"}
        return {"error": "供应商不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/test-provider/{provider_id}")
async def api_test_provider(provider_id: int):
    """测试供应商 API Key 是否可用（调 /models 端点验证）"""
    try:
        provider = await get_provider(provider_id)
        if not provider:
            return {"ok": False, "error": "供应商不存在"}
        base = (provider.get("api_base_url") or "").rstrip("/")
        api_key = provider.get("api_key") or ""
        if not base or not api_key:
            return {"ok": False, "error": "缺少 API Base URL 或 API Key"}

        # 去掉聊天端点后缀，还原到基础地址再拼 /models
        # （供应商可能保存成完整的 .../v1/chat/completions 或 Anthropic 的 .../v1/messages）
        for suffix in ("/chat/completions", "/messages"):
            if base.endswith(suffix):
                base = base.rsplit(suffix, 1)[0]
                break

        api_format = (provider.get("api_format") or "openai") or "openai"
        if api_format == "anthropic":
            # Anthropic 用 x-api-key + anthropic-version，模型列表在 /v1/models
            headers = to_anthropic_headers(api_key)
            models_url = f"{base}/models" if base.endswith("/v1") else f"{base}/v1/models"
        else:
            headers = {"Authorization": f"Bearer {api_key}"}
            models_url = f"{base}/models"

        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(models_url, headers=headers)
        if resp.status_code == 200:
            data = resp.json()
            count = len(data.get("data", []))
            return {"ok": True, "message": f"连接成功，获取到 {count} 个模型"}
        elif resp.status_code == 401:
            return {"ok": False, "error": "API Key 无效（401）"}
        elif resp.status_code == 403:
            return {"ok": False, "error": "权限不足（403）"}
        else:
            return {"ok": False, "error": f"HTTP {resp.status_code}"}
    except httpx.TimeoutException:
        return {"ok": False, "error": "连接超时"}
    except httpx.ConnectError:
        return {"ok": False, "error": "无法连接到服务器"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _detect_provider_type(api_base_url: str) -> str:
    """根据供应商 URL 判断类型"""
    url = (api_base_url or '').lower()
    if 'aihubmix' in url:
        return 'aihubmix'
    elif 'openrouter' in url:
        return 'openrouter'
    return 'generic'


def _transform_aihubmix_model(m: dict) -> dict:
    """将 AIHubMix 新 API 格式转换为 OpenRouter 兼容格式，前端无需改动"""
    features = [f.strip() for f in (m.get('features') or '').split(',') if f.strip()]
    input_mods = [x.strip() for x in (m.get('input_modalities') or 'text').split(',') if x.strip()]
    model_type = (m.get('types') or 'llm').strip()

    # 映射 features → supported_parameters
    params = []
    if 'thinking' in features:
        params.append('reasoning')
    if 'tools' in features or 'function_calling' in features:
        params.append('tools')
    if 'web' in features:
        params.append('web')

    # 推断 output_modalities
    output_mods = ['text']
    if model_type == 'image_generation':
        output_mods = ['image']
    elif model_type == 'video':
        output_mods = ['video']

    pricing = m.get('pricing') or {}
    # 只有真正有定价数据时才转换，避免无定价模型被误判为免费
    transformed_pricing = {}
    if pricing and (pricing.get('input') is not None or pricing.get('output') is not None):
        transformed_pricing = {
            'prompt': str(pricing.get('input', '')),
            'completion': str(pricing.get('output', '')),
        }

    return {
        'id': m.get('model_id', ''),
        'name': m.get('model_id', ''),
        'description': m.get('desc', ''),
        'architecture': {
            'input_modalities': input_mods,
            'output_modalities': output_mods,
        },
        'supported_parameters': params,
        'context_length': m.get('context_length'),
        'max_output': m.get('max_output'),
        'pricing': transformed_pricing if transformed_pricing else None,
        '_is_embedding': model_type == 'embedding',
        '_is_rerank': model_type == 'rerank',
        '_ahm_type': model_type,       # 原始类型，供前端筛选
    }


@app.get("/admin/providers/{provider_id}/models")
async def api_get_provider_models(provider_id: int):
    """从供应商 API 拉取模型列表（代理，避免前端跨域）。同时拉取聊天模型和嵌入模型。"""
    try:
        provider = await get_provider(provider_id)
        if not provider:
            return {"error": "供应商不存在"}

        # 构造基础地址：去掉聊天端点后缀（OpenAI 的 /chat/completions、Anthropic 的 /messages）
        base = provider['api_base_url'].rstrip('/')
        for suffix in ('/chat/completions', '/messages'):
            if base.endswith(suffix):
                base = base.rsplit(suffix, 1)[0]
                break

        provider_type = _detect_provider_type(base)

        # Anthropic 原生供应商用 x-api-key + anthropic-version，模型列表在 /v1/models
        api_format = (provider.get('api_format') or 'openai') or 'openai'
        if api_format == 'anthropic':
            headers = to_anthropic_headers(provider['api_key'] or '')
        else:
            headers = {"Content-Type": "application/json"}
            if provider['api_key']:
                headers["Authorization"] = f"Bearer {provider['api_key']}"

        import httpx
        async with httpx.AsyncClient(timeout=30) as client:

            # ── AIHubMix：优先用新 API，失败降级旧 API ──
            if provider_type == 'aihubmix':
                # 新 API 地址：https://aihubmix.com/api/v1/models
                new_api_base = base.split('/v1')[0] if '/v1' in base else base
                try:
                    resp = await client.get(f"{new_api_base}/api/v1/models", headers=headers)
                    if resp.status_code == 200:
                        raw_models = resp.json().get("data", [])
                        models = [_transform_aihubmix_model(m) for m in raw_models]
                    else:
                        raise ValueError(f"新 API 返回 {resp.status_code}")
                except Exception:
                    # 降级到旧 /v1/models 接口
                    resp = await client.get(f"{base}/models", headers=headers)
                    if resp.status_code != 200:
                        return {"error": f"供应商返回 {resp.status_code}", "detail": resp.text[:500]}
                    models = resp.json().get("data", [])
                    provider_type = 'generic'  # 降级后按通用处理

            # ── OpenRouter / Anthropic / 通用：走 /models 接口 ──
            else:
                # Anthropic 模型列表在 /v1/models（base 已去后缀，可能以 /v1 结尾）
                if api_format == 'anthropic':
                    models_url = f"{base}/models" if base.endswith('/v1') else f"{base}/v1/models"
                else:
                    models_url = f"{base}/models"
                resp = await client.get(models_url, headers=headers)
                if resp.status_code != 200:
                    return {"error": f"供应商返回 {resp.status_code}", "detail": resp.text[:500]}
                chat_models = resp.json().get("data", [])

                # 尝试拉取嵌入模型（不是所有供应商都支持，失败不影响）
                embed_models = []
                try:
                    embed_resp = await client.get(f"{base}/embeddings/models", headers=headers)
                    if embed_resp.status_code == 200:
                        embed_data = embed_resp.json().get("data", [])
                        chat_ids = {m.get("id") for m in chat_models}
                        for m in embed_data:
                            if m.get("id") not in chat_ids:
                                m["_is_embedding"] = True
                                embed_models.append(m)
                except Exception:
                    pass

                models = chat_models + embed_models

        return {
            "status": "ok",
            "provider_id": provider_id,
            "provider_name": provider['name'],
            "provider_type": provider_type,
            "count": len(models),
            "models": models,
        }
    except httpx.TimeoutException:
        return {"error": "请求超时，请检查供应商地址"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 供应商已保存模型管理 API
# ============================================================

@app.get("/admin/all-saved-models")
async def api_get_all_saved_models():
    """获取所有供应商的已保存模型（含供应商名称，用于默认模型选择器）"""
    try:
        models = await get_all_saved_models()
        return {"status": "ok", "models": models}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/providers/{provider_id}/saved-models")
async def api_get_saved_models(provider_id: int):
    """获取供应商已保存的模型列表"""
    try:
        models = await get_provider_models(provider_id)
        return {"status": "ok", "models": models}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/providers/{provider_id}/saved-models")
async def api_add_saved_model(provider_id: int, request: Request):
    """添加模型到供应商"""
    try:
        data = await request.json()
        model_id = data.get("model_id", "").strip()
        if not model_id:
            return {"error": "model_id 不能为空"}

        model = await add_provider_model(
            provider_id=provider_id,
            model_id=model_id,
            display_name=data.get("display_name", ""),
            model_type=data.get("model_type", "chat"),
            input_modes=data.get("input_modes", "text"),
            output_modes=data.get("output_modes", "text"),
            capabilities=data.get("capabilities", ""),
            api_format=data.get("api_format"),
        )
        if model:
            return {"status": "created", "model": model}
        return {"error": "模型已存在"}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/saved-models/{model_pk_id}")
async def api_update_saved_model(model_pk_id: int, request: Request):
    """更新已保存模型的配置"""
    try:
        data = await request.json()
        model = await update_provider_model(model_pk_id, **data)
        if model:
            return {"status": "updated", "model": model}
        return {"error": "模型不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/saved-models/{model_pk_id}")
async def api_delete_saved_model(model_pk_id: int):
    """删除已保存的模型"""
    try:
        success = await delete_provider_model(model_pk_id)
        if success:
            return {"status": "deleted"}
        return {"error": "模型不存在"}
    except Exception as e:
        return {"error": str(e)}

        
# ============================================================
# 记忆分类管理 API（v3.7）
# ============================================================

@app.get("/admin/categories")
async def api_get_categories():
    """获取所有分类（含记忆计数）"""
    try:
        categories = await get_all_categories()
        return {"status": "ok", "categories": categories}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/categories")
async def api_create_category(request: Request):
    """创建分类"""
    try:
        data = await request.json()
        name = data.get("name", "").strip()
        if not name:
            return {"error": "分类名称不能为空"}
        category = await create_category(
            name=name,
            color=data.get("color", "#6B7280"),
            icon=data.get("icon", "📁"),
            sort_order=data.get("sort_order", 0),
        )
        return {"status": "created", "category": category}
    except Exception as e:
        if "unique" in str(e).lower():
            return {"error": "分类名称已存在"}
        return {"error": str(e)}


@app.put("/admin/categories/{category_id}")
async def api_update_category(category_id: int, request: Request):
    """更新分类"""
    try:
        data = await request.json()
        category = await update_category(category_id, **data)
        if category:
            return {"status": "updated", "category": category}
        return {"error": "分类不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/categories/{category_id}")
async def api_delete_category(category_id: int):
    """删除分类"""
    try:
        success = await delete_category(category_id)
        if success:
            return {"status": "deleted"}
        return {"error": "分类不存在"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 联网搜索 API（v3.8）
# ============================================================

@app.get("/admin/search-engines")
async def api_get_search_engines():
    """获取所有支持的搜索引擎列表"""
    return {"engines": get_engine_list()}


@app.get("/admin/search-config")
async def api_get_search_config():
    """获取当前搜索配置（api_key 脱敏，不回显明文）"""
    engine = await get_config("search_engine") or ""
    api_key = await get_config("search_api_key") or ""
    max_results = await get_config_int("search_max_results", fallback=5)
    masked = (api_key[:4] + "…" + api_key[-3:]) if len(api_key) > 10 else ("•" * min(len(api_key), 8))
    return {
        "engine": engine,
        "api_key": masked,
        "api_key_set": bool(api_key),
        "max_results": max_results,
    }


@app.put("/admin/search-config")
async def api_set_search_config(request: Request):
    """更新搜索配置"""
    try:
        data = await request.json()
        if "engine" in data:
            await set_config("search_engine", data["engine"])
        # 仅当传入了非空、且不是脱敏回显的值时才覆盖（留空＝保持原 key）
        _sk = (data.get("api_key") or "")
        if _sk and "…" not in _sk and "•" not in _sk:
            await set_config("search_api_key", _sk)
        if "max_results" in data:
            await set_config("search_max_results", str(data["max_results"]))
        return {"status": "updated"}
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": str(e)})


@app.post("/admin/search-test")
async def api_search_test(request: Request):
    """测试搜索（调试用）"""
    try:
        data = await request.json()
        query = data.get("query", "")
        engine = data.get("engine") or await get_config("search_engine") or ""
        api_key = data.get("api_key") or await get_config("search_api_key") or ""
        max_results = data.get("max_results", 5)
        
        if not query:
            return JSONResponse(status_code=400, content={"error": "query 不能为空"})
        if not engine:
            return JSONResponse(status_code=400, content={"error": "未配置搜索引擎"})
        
        results = await web_search(query=query, engine=engine, api_key=api_key, max_results=max_results)
        return {
            "engine": engine,
            "query": query,
            "count": len(results),
            "results": [r.to_dict() for r in results],
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# MCP 客户端管理 API（v3.8）
# ============================================================

@app.post("/admin/mcp/list-tools")
async def api_mcp_list_tools(request: Request):
    """获取指定 MCP 服务器的工具列表"""
    try:
        data = await request.json()
        servers = data.get("servers", [])
        if not servers:
            return {"tools": [], "tool_map": {}}
        
        openai_tools, tool_map = await get_tools_for_servers(servers)
        return {
            "count": len(openai_tools),
            "tools": [t["function"] for t in openai_tools],
            "tool_map": {k: v["server_name"] for k, v in tool_map.items()},
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/admin/mcp/clear-cache")
async def api_mcp_clear_cache(request: Request):
    """清除 MCP 工具缓存"""
    try:
        data = await request.json()
        url = data.get("url")
        clear_tool_cache(url)
        try:
            from tool_drawer import force_refresh_external_drawers
            await force_refresh_external_drawers()
        except Exception as e:
            print(f"⚠️ 外部 MCP 抽屉刷新失败: {e}")
        return {"status": "cleared", "url": url or "all"}
    except Exception:
        clear_tool_cache()
        try:
            from tool_drawer import force_refresh_external_drawers
            await force_refresh_external_drawers()
        except Exception as e:
            print(f"⚠️ 外部 MCP 抽屉刷新失败: {e}")
        return {"status": "cleared", "url": "all"}


# ============================================================
# 供应商余额查询 API（多供应商通用）
# ============================================================

async def _query_openrouter_credits(api_key: str):
    """查询 OpenRouter 余额"""
    result = {}
    async with httpx.AsyncClient(timeout=10) as client:
        resp1 = await client.get(
            "https://openrouter.ai/api/v1/auth/key",
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if resp1.status_code == 200:
            d = resp1.json().get("data", {})
            result["usage"] = d.get("usage", 0)
            result["limit"] = d.get("limit")
            result["limit_remaining"] = d.get("limit_remaining")
        
        resp2 = await client.get(
            "https://openrouter.ai/api/v1/credits",
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if resp2.status_code == 200:
            d2 = resp2.json().get("data", {})
            result["total_credits"] = d2.get("total_credits", 0)
            result["total_usage"] = d2.get("total_usage", 0)
            result["balance"] = round(d2.get("total_credits", 0) - d2.get("total_usage", 0), 6)
    return result


async def _query_generic_credits(base_url: str, api_key: str):
    """尝试 OpenAI 兼容的余额查询（/v1/dashboard/billing/subscription）"""
    base = base_url.rstrip("/").split("/chat/completions")[0].rstrip("/")
    # 去掉末尾的 /v1 以拿到根域名
    root = base.rsplit("/v1", 1)[0] if "/v1" in base else base
    result = {}
    async with httpx.AsyncClient(timeout=10) as client:
        # 方式1：new-api / one-api 风格的 /v1/dashboard/billing/subscription
        try:
            resp = await client.get(
                f"{root}/v1/dashboard/billing/subscription",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            if resp.status_code == 200:
                d = resp.json()
                hard_limit = d.get("hard_limit_usd") or d.get("system_hard_limit_usd", 0)
                # 过滤掉 new-api 返回的"无限额度"假数字（通常是 1亿）
                if hard_limit and hard_limit < 100000:
                    result["total_credits"] = hard_limit
        except Exception:
            pass
        
        # 方式2：/v1/dashboard/billing/usage
        try:
            from datetime import datetime, timedelta
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            start = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
            resp2 = await client.get(
                f"{root}/v1/dashboard/billing/usage?start_date={start}&end_date={today}",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            if resp2.status_code == 200:
                d2 = resp2.json()
                usage_cents = d2.get("total_usage", 0)
                result["total_usage"] = round(usage_cents / 100, 6) if usage_cents > 1 else usage_cents
        except Exception:
            pass
        
        if "total_credits" in result:
            total_usage = result.get("total_usage", 0)
            result["balance"] = round(result["total_credits"] - total_usage, 6)
    
    return result


@app.get("/admin/credits")
async def api_get_credits():
    """查询所有已启用供应商的余额"""
    try:
        providers = await get_all_providers()
        enabled = [p for p in providers if p.get("enabled")]
        
        if not enabled:
            # 没有配置供应商，用全局环境变量兜底（向后兼容）。
            # 非 OpenRouter 的环境变量供应商也走通用查询，不再只认 OpenRouter。
            if API_KEY:
                if "openrouter" in API_BASE_URL.lower():
                    result = await _query_openrouter_credits(API_KEY)
                    name = "OpenRouter"
                else:
                    result = await _query_generic_credits(API_BASE_URL, API_KEY)
                    name = "环境变量供应商"
                if result:
                    result["provider_name"] = name
                    return {"providers": [result]}
            return {"providers": []}
        
        results = []
        for p in enabled:
            base = p.get("api_base_url", "")
            key = p.get("api_key", "")
            if not key:
                continue
            
            entry = {"provider_id": p["id"], "provider_name": p["name"]}
            
            if "openrouter" in base.lower():
                data = await _query_openrouter_credits(key)
            else:
                data = await _query_generic_credits(base, key)
            
            entry.update(data)
            if data:  # 只返回有数据的
                results.append(entry)
        
        return {"providers": results}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# System Prompt 管理 API（v3.7）
# ============================================================

@app.get("/admin/system-prompt")
async def api_get_system_prompt():
    """获取当前 system prompt"""
    try:
        prompt = await get_active_system_prompt()
        # 判断来源
        db_prompt = await get_system_prompt_from_db()
        source = "database" if db_prompt is not None else "file"
        return {"status": "ok", "content": prompt, "source": source, "length": len(prompt)}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/system-prompt")
async def api_set_system_prompt(request: Request):
    """保存 system prompt 到数据库"""
    try:
        data = await request.json()
        content = data.get("content", "")
        await set_system_prompt_in_db(content)
        return {"status": "updated", "length": len(content)}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 用户画像 API
# ============================================================

@app.post("/admin/update-profile-now")
async def api_update_profile_now():
    """手动触发用户画像更新"""
    try:
        from daily_digest import update_user_profile
        result = await update_user_profile()
        return result
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# v5.8：对话搜索 API
# ============================================================

@app.get("/search/messages")
async def api_search_messages(q: str = "", project_id: str = None, limit: int = 20):
    """
    搜索对话消息内容和标题。
    
    参数：
    - q: 搜索关键词
    - project_id: 项目ID过滤（'none' 表示只搜无项目的对话）
    - limit: 最多返回多少条匹配
    """
    try:
        from database import search_chat_messages
        results = await search_chat_messages(q, project_id=project_id, limit=limit)
        return results
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 云端同步 API（v4.1）
# ============================================================

# ──── 对话 ────

@app.get("/sync/conversations")
async def api_sync_get_conversations():
    """获取对话列表（不含消息体）"""
    try:
        convs = await sync_get_conversations()
        return {"conversations": [_serialize_datetimes(c) for c in convs]}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/sync/conversations/{conv_id}")
async def api_sync_get_conversation(conv_id: str):
    """获取单个对话 + 全部消息"""
    try:
        conv = await sync_get_conversation(conv_id)
        if not conv:
            return JSONResponse(status_code=404, content={"error": "对话不存在"})
        # datetime 对象需要序列化为 ISO 字符串，否则 JSONResponse 会崩溃
        return _serialize_datetimes(conv)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/sync/conversations/{conv_id}")
async def api_sync_upsert_conversation(conv_id: str, request: Request):
    """创建或更新对话（含消息）"""
    try:
        data = await request.json()
        data["id"] = conv_id
        messages = data.pop("messages", None)
        await sync_upsert_conversation(data)
        if messages is not None:
            await sync_upsert_messages(conv_id, messages)
        return {"status": "ok"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/sync/conversations/{conv_id}")
async def api_sync_delete_conversation(conv_id: str):
    """删除对话"""
    try:
        deleted = await sync_delete_conversation(conv_id)
        return {"deleted": deleted}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 项目 ────

@app.get("/sync/projects")
async def api_sync_get_projects():
    """获取所有项目"""
    try:
        projs = await sync_get_projects()
        return {"projects": [_serialize_datetimes(p) for p in projs]}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/sync/projects/{proj_id}")
async def api_sync_upsert_project(proj_id: str, request: Request):
    """创建或更新项目"""
    try:
        data = await request.json()
        data["id"] = proj_id
        await sync_upsert_project(data)
        return {"status": "ok"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/sync/projects/{proj_id}")
async def api_sync_delete_project(proj_id: str):
    """删除项目"""
    try:
        deleted = await sync_delete_project(proj_id)
        # v5.8：删除项目时清理文件块
        try:
            from database import delete_all_file_chunks
            await delete_all_file_chunks(proj_id)
        except Exception:
            pass
        return {"deleted": deleted}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── v5.8：项目文件分块处理 ────

@app.post("/projects/{proj_id}/files/{file_id}/process")
async def api_process_file_chunks(proj_id: str, file_id: str, request: Request):
    """
    接收文件文本内容，分块 + 生成嵌入 + 存入数据库。
    前端上传文件后调用。
    body: { "file_name": "xxx.txt", "text_content": "..." }
    """
    try:
        data = await request.json()
        file_name = data.get("file_name", "")
        text_content = data.get("text_content", "")
        
        if not text_content or not text_content.strip():
            return {"chunks": 0, "message": "无文本内容"}
        
        from database import save_file_chunks, delete_file_chunks
        # 先删除旧的块（如果文件重新上传）
        await delete_file_chunks(proj_id, file_id)
        # 分块 + 嵌入 + 存储
        count = await save_file_chunks(proj_id, file_id, file_name, text_content)
        return {"chunks": count, "file_id": file_id, "file_name": file_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/projects/{proj_id}/files/{file_id}/chunks")
async def api_delete_file_chunks(proj_id: str, file_id: str):
    """删除某个文件的所有块"""
    try:
        from database import delete_file_chunks
        count = await delete_file_chunks(proj_id, file_id)
        return {"deleted": count}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 批量导入（localStorage → 数据库） ────

@app.post("/sync/import")
async def api_sync_import(request: Request):
    """一次性导入所有 localStorage 数据"""
    try:
        data = await request.json()
        conversations = data.get("conversations", [])
        projects = data.get("projects", [])
        result = await sync_import_all(conversations, projects)
        print(f"📦 云端同步导入完成：{result}")
        return {"status": "ok", **result}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 用户/助手配置同步（复用 config 表） ────

@app.get("/sync/settings")
async def api_sync_get_settings():
    """获取所有同步配置（头像、昵称、助手设置等）"""
    result = {}
    for key in SYNC_SETTING_KEYS:
        val = await get_config(key)
        result[key] = val or ""
    return result


@app.put("/sync/settings")
async def api_sync_put_settings(request: Request):
    """批量更新同步配置"""
    try:
        data = await request.json()
        updated, rejected = [], []
        for key, value in data.items():
            if key not in SYNC_SETTING_KEYS:
                rejected.append(key)
                continue
            ok = await set_config(key, str(value) if value is not None else "")
            (updated if ok else rejected).append(key)
        # rejected = 同步白名单拒收，或白名单内键被 set_config 校验拒收。
        # 必须如实回报，否则前端批量同步对"被静默丢弃的设置"毫无察觉
        # （例如网关版本过旧、还没有某个新配置键时，对应设置会一直存不进却无任何提示）。
        if rejected:
            print(f"⚠️  /sync/settings 拒收配置项（未落库）: {rejected}")
        return {"status": "ok", "updated": updated, "rejected": rejected}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 数据导出（备份 zip） ────

@app.get("/sync/export")
async def api_sync_export():
    """导出全部数据为 zip"""
    import io
    import zipfile
    from datetime import datetime, timezone

    try:
        # 收集所有数据
        convs_raw = await sync_get_conversations()
        # 为每个对话加载消息
        convs_full = []
        for c in convs_raw:
            full = await sync_get_conversation(c["id"])
            if full:
                # datetime 转 ISO 字符串
                convs_full.append(_serialize_datetimes(full))
            else:
                convs_full.append(_serialize_datetimes(c))

        projs_raw = await sync_get_projects()
        projs = [_serialize_datetimes(p) for p in projs_raw]

        # 记忆
        pool = await get_pool()
        async with pool.acquire() as conn:
            mem_rows = await conn.fetch("SELECT * FROM memories ORDER BY created_at DESC")
            version_rows = await conn.fetch("""
                SELECT * FROM memory_versions ORDER BY memory_id, version
            """)
            event_rows = await conn.fetch("""
                SELECT * FROM memory_events ORDER BY occurred_at, event_id
            """)
            stored_archive_fingerprint = await conn.fetchval(
                "SELECT value FROM gateway_config WHERE key='archive_id_hmac_key_fingerprint'"
            )
        memories = [_serialize_datetimes(dict(r)) for r in mem_rows]
        memory_versions = [_serialize_datetimes(dict(r)) for r in version_rows]
        memory_events = [_serialize_datetimes(dict(r)) for r in event_rows]
        current_archive_ids = bool(memory_events) and all(
            is_opaque_archive_identifier(event.get("event_id"), "event")
            and is_opaque_archive_identifier(event.get("conversation_id"), "conversation")
            for event in memory_events
        )
        archive_fingerprint = archive_identifier_key_fingerprint() if current_archive_ids else ""
        if current_archive_ids and stored_archive_fingerprint != archive_fingerprint:
            raise ArchiveIdentifierConfigError(
                "archive HMAC key does not match the fingerprint recorded with existing events"
            )
        archive_manifest = {
            "identifier_version": 2 if current_archive_ids else 1,
            "key_fingerprint": archive_fingerprint,
        }

        # 配置
        all_config = await get_all_config()
        config_flat = {}
        for k, v in all_config.items():
            config_flat[k] = v.get("value", "") if isinstance(v, dict) else v

        # 同步设置
        settings = {}
        for key in SYNC_SETTING_KEYS:
            val = await get_config(key)
            settings[key] = val or ""

        # 打包 zip
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("conversations.json", json.dumps(convs_full, ensure_ascii=False, indent=2))
            zf.writestr("projects.json", json.dumps(projs, ensure_ascii=False, indent=2))
            zf.writestr("memories.json", json.dumps(memories, ensure_ascii=False, indent=2))
            zf.writestr("memory_versions.json", json.dumps(memory_versions, ensure_ascii=False, indent=2))
            zf.writestr("memory_events.json", json.dumps(memory_events, ensure_ascii=False, indent=2))
            zf.writestr("archive_manifest.json", json.dumps(archive_manifest, ensure_ascii=False, indent=2))
            zf.writestr("config.json", json.dumps(config_flat, ensure_ascii=False, indent=2))
            zf.writestr("settings.json", json.dumps(settings, ensure_ascii=False, indent=2))
        buf.seek(0)

        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        filename = f"kiwi-mem-backup-{ts}.zip"

        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


def _serialize_datetimes(obj):
    """递归将 datetime 对象转为 ISO 字符串"""
    from datetime import datetime as _dt
    if isinstance(obj, dict):
        return {k: _serialize_datetimes(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_serialize_datetimes(v) for v in obj]
    elif isinstance(obj, _dt):
        return obj.isoformat()
    return obj


def _parse_backup_datetime(value):
    """Parse an exported ISO timestamp without inventing a value for nulls."""
    from datetime import datetime as _dt
    if not value:
        return None
    if isinstance(value, _dt):
        return value
    try:
        return _dt.fromisoformat(str(value).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None


# ──── 数据导入（从备份 zip 恢复） ────

@app.post("/sync/import-backup")
async def api_sync_import_backup(file: UploadFile = File(...)):
    """从备份 zip 恢复数据"""
    import io
    import zipfile

    try:
        content = await file.read()
        buf = io.BytesIO(content)

        if not zipfile.is_zipfile(buf):
            return JSONResponse(status_code=400, content={"error": "不是有效的 zip 文件"})

        buf.seek(0)
        result = {"conversations": 0, "messages": 0, "projects": 0, "memories": 0, "memory_versions": 0, "memory_events": 0, "settings": 0, "config": 0}

        with zipfile.ZipFile(buf, 'r') as zf:
            names = zf.namelist()

            archived_events = json.loads(zf.read("memory_events.json")) if "memory_events.json" in names else []
            trusted_opaque_archive = False
            if archived_events:
                current_flags = [
                    is_opaque_archive_identifier(event.get("event_id"), "event")
                    and is_opaque_archive_identifier(event.get("conversation_id"), "conversation")
                    for event in archived_events
                ]
                if any(current_flags) and not all(current_flags):
                    return JSONResponse(status_code=409, content={"error": "backup mixes legacy and current archive identifiers"})
                trusted_opaque_archive = all(current_flags)
                if trusted_opaque_archive:
                    if "archive_manifest.json" not in names:
                        return JSONResponse(status_code=409, content={"error": "current archive backup is missing its key manifest"})
                    manifest = json.loads(zf.read("archive_manifest.json"))
                    if (
                        int(manifest.get("identifier_version") or 0) != 2
                        or manifest.get("key_fingerprint") != archive_identifier_key_fingerprint()
                    ):
                        return JSONResponse(status_code=409, content={"error": "backup archive HMAC key does not match this server"})

            # 导入项目
            if "projects.json" in names:
                projs = json.loads(zf.read("projects.json"))
                for p in projs:
                    await sync_upsert_project(p)
                    result["projects"] += 1

            # 导入对话 + 消息
            if "conversations.json" in names:
                convs = json.loads(zf.read("conversations.json"))
                for conv in convs:
                    messages = conv.pop("messages", [])
                    await sync_upsert_conversation(conv)
                    if messages:
                        await sync_upsert_messages(conv["id"], messages)
                        result["messages"] += len(messages)
                    result["conversations"] += 1

            # 导入记忆
            memory_id_map = {}
            if "memories.json" in names:
                mems = json.loads(zf.read("memories.json"))
                for mem in mems:
                    try:
                        new_memory_id = await save_memory(
                            content=mem.get("content", ""),
                            importance=mem.get("importance", 5),
                            title=mem.get("title", ""),
                            category_id=mem.get("category_id"),
                            source=mem.get("source", "backup_import"),
                            source_session=mem.get("source_session", ""),
                            emotional_weight=mem.get("emotional_weight", 0) or 0,
                            project_id=mem.get("project_id"),
                        )
                        memory_id_map[mem.get("id")] = new_memory_id
                        pool = await get_pool()
                        async with pool.acquire() as conn:
                            await conn.execute("""
                                UPDATE memories
                                SET memory_type = $2,
                                    is_permanent = $3,
                                    lock_source = $4,
                                    created_at = COALESCE($5, created_at),
                                    last_accessed = COALESCE($6, last_accessed),
                                    valid_from = COALESCE($7, valid_from),
                                    valid_until = $8,
                                    softened_at = $9,
                                    resolution = COALESCE($10, resolution),
                                    dream_processed_at = $11,
                                    archived_at = $12,
                                    archive_reason = $13
                                WHERE id = $1
                            """, new_memory_id,
                                 mem.get("memory_type", "fragment"),
                                 bool(mem.get("is_permanent", False)),
                                 mem.get("lock_source"),
                                 _parse_backup_datetime(mem.get("created_at")),
                                 _parse_backup_datetime(mem.get("last_accessed")),
                                 _parse_backup_datetime(mem.get("valid_from")),
                                 _parse_backup_datetime(mem.get("valid_until")),
                                 _parse_backup_datetime(mem.get("softened_at")),
                                 mem.get("resolution", 1.0),
                                 _parse_backup_datetime(mem.get("dream_processed_at")),
                                 _parse_backup_datetime(mem.get("archived_at")),
                                 mem.get("archive_reason"))
                        result["memories"] += 1
                    except Exception:
                        pass  # 跳过重复或无效记忆

            # 版本在所有主记忆建立 ID 映射后恢复。save_memory 自动创建的
            # version=1 会由备份中的不可变原始版本覆盖。
            if "memory_versions.json" in names and memory_id_map:
                versions = json.loads(zf.read("memory_versions.json"))
                pool = await get_pool()
                async with pool.acquire() as conn:
                    for version in versions:
                        new_memory_id = memory_id_map.get(version.get("memory_id"))
                        if not new_memory_id:
                            continue
                        await conn.execute("""
                            INSERT INTO memory_versions
                                (memory_id, version, content, title, resolution,
                                 embedding, change_type, created_at)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, COALESCE($8, NOW()))
                            ON CONFLICT (memory_id, version) DO UPDATE SET
                                content = EXCLUDED.content,
                                title = EXCLUDED.title,
                                resolution = EXCLUDED.resolution,
                                embedding = EXCLUDED.embedding,
                                change_type = EXCLUDED.change_type,
                                created_at = EXCLUDED.created_at
                        """, new_memory_id, version.get("version", 1),
                             version.get("content", ""), version.get("title", ""),
                             version.get("resolution", 1.0), version.get("embedding"),
                             version.get("change_type", "original"),
                             _parse_backup_datetime(version.get("created_at")))
                        result["memory_versions"] += 1

            # 完整对话原文是 append-only 证据：备份恢复只做幂等插入，
            # 同 event_id 的不同内容会触发冲突，不得覆盖旧原文。
            if "memory_events.json" in names:
                for event in archived_events:
                    restored = await ingest_memory_event(
                        event_id=str(event.get("event_id") or ""),
                        source_client=str(event.get("source_client") or "backup_import"),
                        conversation_id=str(event.get("conversation_id") or ""),
                        role=str(event.get("role") or ""),
                        content=str(event.get("content") or ""),
                        occurred_at=_parse_backup_datetime(event.get("occurred_at")) or datetime.now(timezone.utc),
                        metadata=event.get("metadata") if isinstance(event.get("metadata"), dict) else {},
                        memory_space_id=str(event.get("memory_space_id") or "zhizhi_grey"),
                        assistant_identity_id=str(event.get("assistant_identity_id") or "grey_knox"),
                        trusted_opaque_ids=trusted_opaque_archive,
                    )
                    if restored.get("created"):
                        result["memory_events"] += 1

            # 导入同步设置
            if "settings.json" in names:
                settings = json.loads(zf.read("settings.json"))
                for key, val in settings.items():
                    if val:
                        await set_config(key, str(val))
                        result["settings"] += 1

            # 导入 gateway 配置
            if "config.json" in names:
                config = json.loads(zf.read("config.json"))
                for key, val in config.items():
                    if val:
                        ok = await set_config(key, str(val))
                        if ok:
                            result["config"] += 1

        print(f"📦 备份导入完成：{result}")
        return {"status": "ok", **result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 数据重置 ────

@app.delete("/sync/reset")
async def api_sync_reset(request: Request):
    """重置全部聊天数据（对话+项目+同步设置），记忆和 gateway 配置保留"""
    try:
        data = await request.json()
        confirm = data.get("confirm")
        if confirm != "RESET_ALL_DATA":
            return JSONResponse(status_code=400, content={"error": "需要确认码 confirm='RESET_ALL_DATA'"})

        pool = await get_pool()
        async with pool.acquire() as conn:
            # 删除所有消息和对话（级联）
            deleted_convs = await conn.execute("DELETE FROM chat_conversations")
            deleted_projs = await conn.execute("DELETE FROM chat_projects")
            # compression_summaries 无外键，不会随对话级联删除，手动清空
            await conn.execute("DELETE FROM compression_summaries")

            # 清除同步设置
            for key in SYNC_SETTING_KEYS:
                await conn.execute("DELETE FROM gateway_config WHERE key = $1", key)

        print("⚠️ 数据重置完成")
        return {
            "status": "ok",
            "message": "所有聊天数据已重置",
            "deleted_conversations": deleted_convs,
            "deleted_projects": deleted_projs,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 提醒系统 API（v4.2）
# ============================================================

@app.get("/reminders")
async def api_get_reminders(all: bool = False):
    """获取提醒列表（默认只返回活跃的）"""
    try:
        reminders = await get_reminders(include_completed=all)
        return JSONResponse(content=reminders)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/reminders")
async def api_create_reminder(request: Request):
    """手动创建提醒"""
    try:
        body = await request.json()
        result = await create_reminder(body)
        return JSONResponse(content=result)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# 注意：/reminders/due 和 /reminders/{rid}/fire 必须在 /reminders/{rid} 之前定义，
# 否则 "due" 和 "xxx/fire" 会被 {rid} 路径参数捕获

@app.get("/reminders/due")
async def api_get_due_reminders():
    """获取所有到期的提醒（前端轮询用）"""
    try:
        due = await get_due_reminders()
        return JSONResponse(content=due)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/reminders/{rid}/fire")
async def api_fire_reminder(rid: str):
    """标记提醒已触发（前端调用）"""
    try:
        reminders = await get_reminders(include_completed=True)
        reminder = next((r for r in reminders if r["id"] == rid), None)
        if not reminder:
            return JSONResponse(status_code=404, content={"error": "提醒不存在"})
        ok = await fire_reminder(rid, reminder.get("repeat_type", "once"), reminder.get("repeat_config"))
        return JSONResponse(content={"ok": ok, "repeat_type": reminder.get("repeat_type")})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.put("/reminders/{rid}")
async def api_update_reminder(rid: str, request: Request):
    """更新提醒"""
    try:
        body = await request.json()
        ok = await update_reminder(rid, body)
        if ok:
            return JSONResponse(content={"ok": True})
        return JSONResponse(status_code=404, content={"error": "提醒不存在"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/reminders/{rid}")
async def api_delete_reminder(rid: str):
    """删除提醒"""
    try:
        ok = await delete_reminder(rid)
        if ok:
            return JSONResponse(content={"ok": True})
        return JSONResponse(status_code=404, content={"error": "提醒不存在"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 挂载 MCP Server（Streamable HTTP）
# ============================================================
#
# 记忆系统：/memory/mcp
#   工具：search_memory, save_memory, get_recent, trigger_digest

app.mount("/memory", get_mcp_app())
app.mount("/calendar", get_calendar_mcp_app())

# 管理面板静态资源（css/js/assets）。必须在所有 /admin/* API 路由之后挂载，
# 这样显式 API 路由优先匹配，本挂载只接管 /admin/css、/admin/js 等静态文件。
# /admin 返回 index.html（上面的 @app.get("/admin")），/admin/ 由 html=True 提供；
# 页面内用绝对路径 /admin/... 引用资源。
from fastapi.staticfiles import StaticFiles as _StaticFiles
_panel_dir = os.path.join(os.path.dirname(__file__), "admin-panel")
if os.path.isdir(_panel_dir):
    app.mount("/admin", _StaticFiles(directory=_panel_dir, html=True), name="admin-panel")


# ============================================================
# 启动入口
# ============================================================

if __name__ == "__main__":
    import uvicorn
    print(f"🚀 AI Memory Gateway 启动中... 端口 {PORT}")
    print(f"📝 人设长度：{len(SYSTEM_PROMPT)} 字符")
    print(f"🤖 默认模型：{DEFAULT_MODEL}")
    print(f"🔗 API 地址：{API_BASE_URL}")
    print(f"🧠 记忆系统：{'开启' if MEMORY_ENABLED else '关闭'}")
    if MEMORY_ENABLED:
        print(f"📊 记忆提取间隔：每 {MEMORY_EXTRACT_INTERVAL} 轮")
    uvicorn.run(app, host="0.0.0.0", port=PORT)
